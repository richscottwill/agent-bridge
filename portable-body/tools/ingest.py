#!/usr/bin/env python3
"""
WW Dashboard Ingester — Weekly Performance Analysis Tool

Reads the AB SEM WW Dashboard xlsx file and produces:
1. Structured JSON data extract (for programmatic use)
2. Markdown callout drafts per market (AU, MX, and all WW markets)
3. WW summary with trends, anomalies, and projections

Usage:
    python3 ingest.py <path_to_xlsx> [--week YYYY_WNN] [--markets AU,MX] [--output-dir DIR]

If --week is omitted, auto-detects the latest week with data.
If --markets is omitted, processes all markets.
"""

import openpyxl
import datetime
import json
import sys
import os
import argparse
from collections import defaultdict, OrderedDict
from pathlib import Path

# ── Column layout for daily market tabs ──
# A=Weeks, B=Date, C=Month, D=Cost, E=Clicks, F=Impressions, G=Reg
# (H=spacer)
# I=Brand Cost, J=Brand Clicks, K=Brand Imp, L=Brand Reg
# (M=spacer)
# N=NB Cost, O=NB Clicks, P=NB Imp, Q=NB Reg

DAILY_COLS = {
    'week': 0, 'date': 1, 'month': 2,
    'cost': 3, 'clicks': 4, 'impressions': 5, 'regs': 6,
    'brand_cost': 8, 'brand_clicks': 9, 'brand_imp': 10, 'brand_regs': 11,
    'nb_cost': 13, 'nb_clicks': 14, 'nb_imp': 15, 'nb_regs': 16,
}

ALL_MARKETS = ['US', 'CA', 'UK', 'DE', 'FR', 'IT', 'ES', 'JP', 'AU', 'MX']

# Weekly tab metric blocks — auto-detected at runtime (see _detect_weekly_blocks)
# These are fallback values only, used if auto-detection fails
WEEKLY_METRICS_FALLBACK = {
    'spend': 1, 'regs': 40, 'cpa': 79, 'clicks': 118,
    'impressions': 157, 'cpc': 196, 'cvr': 235, 'ctr': 274,
}

# Market row offsets within each Weekly metric block (0-indexed from block start)
# Each market has 3 rows: total, brand, nb
WEEKLY_MARKET_ORDER = ['US', 'JP', 'UK', 'DE', 'FR', 'IT', 'ES', 'CA', 'MX', 'AU']


def safe_float(v):
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def safe_int(v):
    if v is None:
        return 0
    try:
        return int(round(float(v)))
    except (ValueError, TypeError):
        return 0


def pct(a, b):
    """Percentage change from b to a."""
    if b == 0:
        return None
    return (a - b) / b


def fmt_pct(v):
    if v is None:
        return 'N/A'
    return f'{v:+.1%}'


def fmt_dollar(v):
    if abs(v) >= 1000:
        return f'${v:,.0f}'
    return f'${v:,.2f}'


def fmt_int(v):
    return f'{v:,}'



class DashboardIngester:
    def __init__(self, xlsx_path):
        self.xlsx_path = xlsx_path
        self.wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        self.sheet_names = self.wb.sheetnames

    def _detect_daily_cols(self, market):
        """Auto-detect column layout for a daily market tab by reading the header row.

        Returns a dict mapping metric names to column indices.
        Falls back to DAILY_COLS if header detection fails.
        """
        if market not in self.sheet_names:
            return DAILY_COLS
        ws = self.wb[market]
        header = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        # Build keyword-to-metric mapping
        col_keywords = {
            'week': ['week'],
            'date': ['date'],
            'month': ['month'],
            'cost': ['cost', 'spend'],
            'clicks': ['click'],
            'impressions': ['impression', 'impr'],
            'regs': ['reg', 'registration', 'conv'],
            'brand_cost': ['brand cost', 'brand spend'],
            'brand_clicks': ['brand click'],
            'brand_imp': ['brand imp', 'brand impression'],
            'brand_regs': ['brand reg', 'brand conv'],
            'nb_cost': ['nb cost', 'non-brand cost', 'non brand cost', 'nb spend'],
            'nb_clicks': ['nb click', 'non-brand click', 'non brand click'],
            'nb_imp': ['nb imp', 'non-brand imp', 'non brand imp'],
            'nb_regs': ['nb reg', 'non-brand reg', 'non brand reg', 'nb conv'],
        }

        detected = {}
        for col_idx, cell in enumerate(header):
            cell_str = str(cell).strip().lower() if cell else ''
            if not cell_str:
                continue
            for metric, keywords in col_keywords.items():
                if metric not in detected:
                    for kw in keywords:
                        if kw in cell_str:
                            detected[metric] = col_idx
                            break

        # Also scan for any additional columns beyond the known set
        extra_cols = {}
        for col_idx, cell in enumerate(header):
            cell_str = str(cell).strip().lower() if cell else ''
            if not cell_str or col_idx in detected.values():
                continue
            # Capture CPC, CVR, CTR, CPA columns if present
            for metric in ['cpc', 'cvr', 'ctr', 'cpa']:
                if metric in cell_str and metric not in detected and metric not in extra_cols:
                    prefix = ''
                    if 'brand' in cell_str:
                        prefix = 'brand_'
                    elif 'nb' in cell_str or 'non-brand' in cell_str or 'non brand' in cell_str:
                        prefix = 'nb_'
                    extra_cols[f'{prefix}{metric}'] = col_idx

        # Merge detected + extras, fall back to DAILY_COLS for anything missing
        result = dict(DAILY_COLS)  # start with defaults
        result.update(detected)
        result.update(extra_cols)
        return result

    def _read_daily_tab(self, market):
        """Read daily data from a market tab. Returns list of day dicts.

        Auto-detects column layout from header row. Reads all rows including hidden ones.
        """
        if market not in self.sheet_names:
            return []
        ws = self.wb[market]
        cols = self._detect_daily_cols(market)

        days = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            week_label = str(row[cols['week']]) if cols['week'] < len(row) and row[cols['week']] else ''
            date_col = cols['date']
            date_val = row[date_col] if date_col < len(row) else None
            if not week_label or not date_val:
                continue
            if isinstance(date_val, datetime.datetime):
                date_str = date_val.strftime('%Y-%m-%d')
            else:
                date_str = str(date_val)

            def get_float(key):
                idx = cols.get(key)
                if idx is not None and idx < len(row):
                    return safe_float(row[idx])
                return 0.0

            def get_int(key):
                idx = cols.get(key)
                if idx is not None and idx < len(row):
                    return safe_int(row[idx])
                return 0

            cost = get_float('cost')
            regs = get_int('regs')
            clicks = get_int('clicks')
            # Skip rows with no data (future dates)
            if cost == 0 and regs == 0 and clicks == 0:
                continue

            month_col = cols.get('month')
            month_val = str(row[month_col]) if month_col is not None and month_col < len(row) and row[month_col] else ''

            day = {
                'week': week_label,
                'date': date_str,
                'month': month_val,
                'cost': cost,
                'clicks': clicks,
                'impressions': get_int('impressions'),
                'regs': regs,
                'brand_cost': get_float('brand_cost'),
                'brand_clicks': get_int('brand_clicks'),
                'brand_imp': get_int('brand_imp'),
                'brand_regs': get_int('brand_regs'),
                'nb_cost': get_float('nb_cost'),
                'nb_clicks': get_int('nb_clicks'),
                'nb_imp': get_int('nb_imp'),
                'nb_regs': get_int('nb_regs'),
            }
            # Add any extra detected metrics (CPC, CVR, CTR, CPA at daily level)
            for key in cols:
                if key not in day and key not in ('week', 'date', 'month'):
                    idx = cols[key]
                    if idx < len(row) and row[idx] is not None:
                        day[key] = safe_float(row[idx])
            days.append(day)
        return days

    def _aggregate_week(self, days, week_label):
        """Aggregate daily data for a given week label."""
        week_days = [d for d in days if d['week'] == week_label]
        if not week_days:
            return None
        agg = {
            'week': week_label,
            'date_range': f"{week_days[0]['date']} to {week_days[-1]['date']}",
            'num_days': len(week_days),
            'cost': sum(d['cost'] for d in week_days),
            'clicks': sum(d['clicks'] for d in week_days),
            'impressions': sum(d['impressions'] for d in week_days),
            'regs': sum(d['regs'] for d in week_days),
            'brand_cost': sum(d['brand_cost'] for d in week_days),
            'brand_clicks': sum(d['brand_clicks'] for d in week_days),
            'brand_imp': sum(d['brand_imp'] for d in week_days),
            'brand_regs': sum(d['brand_regs'] for d in week_days),
            'nb_cost': sum(d['nb_cost'] for d in week_days),
            'nb_clicks': sum(d['nb_clicks'] for d in week_days),
            'nb_imp': sum(d['nb_imp'] for d in week_days),
            'nb_regs': sum(d['nb_regs'] for d in week_days),
        }
        # Derived metrics
        agg['cpa'] = agg['cost'] / agg['regs'] if agg['regs'] > 0 else None
        agg['cpc'] = agg['cost'] / agg['clicks'] if agg['clicks'] > 0 else None
        agg['cvr'] = agg['regs'] / agg['clicks'] if agg['clicks'] > 0 else None
        agg['ctr'] = agg['clicks'] / agg['impressions'] if agg['impressions'] > 0 else None
        agg['brand_cpa'] = agg['brand_cost'] / agg['brand_regs'] if agg['brand_regs'] > 0 else None
        agg['brand_cpc'] = agg['brand_cost'] / agg['brand_clicks'] if agg['brand_clicks'] > 0 else None
        agg['brand_cvr'] = agg['brand_regs'] / agg['brand_clicks'] if agg['brand_clicks'] > 0 else None
        agg['nb_cpa'] = agg['nb_cost'] / agg['nb_regs'] if agg['nb_regs'] > 0 else None
        agg['nb_cpc'] = agg['nb_cost'] / agg['nb_clicks'] if agg['nb_clicks'] > 0 else None
        agg['nb_cvr'] = agg['nb_regs'] / agg['nb_clicks'] if agg['nb_clicks'] > 0 else None
        return agg

    def _get_unique_weeks(self, days):
        """Get ordered list of unique week labels from daily data."""
        seen = set()
        weeks = []
        for d in days:
            if d['week'] not in seen:
                seen.add(d['week'])
                weeks.append(d['week'])
        return weeks

    def _detect_latest_week(self, days):
        """Find the latest week that has a full 7 days of data."""
        weeks = self._get_unique_weeks(days)
        # Walk backwards to find the latest complete week
        for w in reversed(weeks):
            week_days = [d for d in days if d['week'] == w]
            if len(week_days) == 7 and sum(d['regs'] for d in week_days) > 0:
                return w
        # Fallback: latest week with any data
        for w in reversed(weeks):
            week_days = [d for d in days if d['week'] == w]
            if sum(d['regs'] for d in week_days) > 0:
                return w
        return weeks[-1] if weeks else None

    def _find_yoy_week(self, week_label):
        """Given '2026 W12', return '2025 W12'."""
        parts = week_label.split()
        if len(parts) == 2:
            year = int(parts[0])
            return f'{year - 1} {parts[1]}'
        return None

    def _get_week_number(self, week_label):
        """Extract week number from '2026 W12' -> 12."""
        parts = week_label.split()
        if len(parts) == 2 and parts[1].startswith('W'):
            return int(parts[1][1:])
        return None

    def _get_week_year(self, week_label):
        """Extract year from '2026 W12' -> 2026."""
        parts = week_label.split()
        if len(parts) == 2:
            return int(parts[0])
        return None

    def _read_ieccp(self, target_week):
        """Read ie%CCP values from the IECCP tab for the target week and prior week.

        The IECCP tab has multiple sections stacked vertically:
          - Rows 1-12:  CPA (per market) — NOT ie%CCP, skip these
          - Row 15:     "IECCP" header
          - Rows 16-26: ie%CCP ratios (the actual metric we want)
          - Row 28:     "IECCP Segment" header
          - Rows 30-48: CCP Segment (Brand/NB CCP per registration values)
          - Rows 52-70: CPA Segment (Brand/NB CPA)
          - Rows 75-93: CCP (Brand/NB CCP guidance values, e.g. MX Brand=90, NB=30)
          - Rows 96-107: CCP per Account (blended CCP per registration)

        ie%CCP formula: CPA / CCP_per_Account
          where CCP_per_Account = (Brand_CCP * Brand_Regs + NB_CCP * NB_Regs) / Total_Regs
        ie%CCP < 1.0 means spend is efficient (CPA < CCP). Target is 1.0 (100%).
        Values are ratios (0.0-1.0 range for healthy markets).

        BUG FIX (2026-03-30): Previously scanned from row 1, which matched CPA
        rows (rows 2-12) instead of IECCP rows (rows 16-26). CPA values like
        $65.59 were multiplied by 100 to produce nonsensical 6559% ie%CCP.
        Now scans only from the IECCP header (row 15) onward.
        """
        if 'IECCP' not in self.sheet_names:
            return {}
        ws = self.wb['IECCP']

        # Read all rows (no max_row/max_col limit)
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return {}

        # Row 1 = week headers (col 0 = label, cols 1+ = weeks)
        row1 = all_rows[0]
        tw_col = None
        lw_col = None
        for i, v in enumerate(row1):
            if v and str(v).strip() == target_week:
                tw_col = i
                lw_col = i - 1 if i > 0 else None
                break
        if tw_col is None:
            return {}

        # Find the IECCP header row to start scanning from the correct section
        ieccp_start = None
        for row_idx, row in enumerate(all_rows):
            cell = str(row[0]).strip() if row and row[0] else ''
            if cell == 'IECCP':
                ieccp_start = row_idx
                break

        if ieccp_start is None:
            return {}

        # Scan for market rows ONLY after the IECCP header
        market_labels = {
            'US': ['US', 'US SEM'],
            'JP': ['JP', 'JP SEM'],
            'UK': ['UK', 'UK SEM'],
            'DE': ['DE', 'DE SEM'],
            'FR': ['FR', 'FR SEM'],
            'IT': ['IT', 'IT SEM'],
            'ES': ['ES', 'ES SEM'],
            'CA': ['CA', 'CA SEM'],
            'MX': ['MX', 'MX SEM'],
            'AU': ['AU', 'AU SEM'],
        }

        # Scan rows after IECCP header, stop at next section header or empty gap
        market_row_map = {}
        for row_idx in range(ieccp_start + 1, min(ieccp_start + 15, len(all_rows))):
            row = all_rows[row_idx]
            cell = str(row[0]).strip() if row and row[0] else ''
            if not cell:
                continue
            # Stop if we hit the next section header
            if cell in ('IECCP Segment', 'CCP Segment', 'CPA Segment', 'CCP', 'CCP per Acc'):
                break
            for market, labels in market_labels.items():
                if market not in market_row_map and cell in labels:
                    market_row_map[market] = row_idx
                    break

        result = {}
        for market, row_idx in market_row_map.items():
            row = all_rows[row_idx]
            tw_val = safe_float(row[tw_col]) if tw_col < len(row) else None
            lw_val = safe_float(row[lw_col]) if lw_col and lw_col < len(row) else None
            if tw_val:
                result[market] = {'tw': tw_val, 'lw': lw_val}

        # Also pick up WW and EU5 aggregates from the same section
        for row_idx in range(ieccp_start + 1, min(ieccp_start + 15, len(all_rows))):
            row = all_rows[row_idx]
            cell = str(row[0]).strip() if row and row[0] else ''
            if cell in ('WW', 'WW SEM'):
                tw_val = safe_float(row[tw_col]) if tw_col < len(row) else None
                lw_val = safe_float(row[lw_col]) if lw_col and lw_col < len(row) else None
                if tw_val:
                    result['WW'] = {'tw': tw_val, 'lw': lw_val}
            elif cell in ('EU5', 'EU5 SEM'):
                tw_val = safe_float(row[tw_col]) if tw_col < len(row) else None
                lw_val = safe_float(row[lw_col]) if lw_col and lw_col < len(row) else None
                if tw_val:
                    result['EU5'] = {'tw': tw_val, 'lw': lw_val}

        return result

    def _detect_weekly_blocks(self):
        """Auto-detect metric block positions on the Weekly tab.

        Scans column A for rows containing 'US SEM' or 'US' to find the start
        of each metric block. Identifies the metric by scanning header rows above.

        Returns dict: {metric_name: block_start_row (1-indexed)}
        """
        if 'Weekly' not in self.sheet_names:
            return WEEKLY_METRICS_FALLBACK
        ws = self.wb['Weekly']

        # Read all of column A to find block starts (HA386 = col 209, row 386)
        all_col_a = []
        for row in ws.iter_rows(min_col=1, max_col=1, min_row=1, max_row=ws.max_row, values_only=True):
            all_col_a.append(str(row[0]).strip() if row[0] else '')

        # Find all rows where col A contains 'US SEM' or exactly 'US' (first market in each block)
        us_labels = ['US SEM', 'US']
        block_starts = []
        for i, cell in enumerate(all_col_a):
            if cell in us_labels:
                block_starts.append(i + 1)  # convert to 1-indexed

        if not block_starts:
            return WEEKLY_METRICS_FALLBACK

        # For each block start, look upward for the metric name in header rows
        metric_keywords = {
            'spend': ['spend', 'cost'],
            'regs': ['reg', 'registration'],
            'cpa': ['cpa', 'cost per acq', 'cost per reg'],
            'clicks': ['click'],
            'impressions': ['impression', 'impr'],
            'cpc': ['cpc', 'cost per click'],
            'cvr': ['cvr', 'conversion rate', 'conv rate'],
            'ctr': ['ctr', 'click through', 'click-through'],
        }

        blocks = {}
        for bs in block_starts:
            metric_name = None
            # Look up to 5 rows above for a header
            for lookback in range(1, 6):
                check_idx = bs - 1 - lookback  # 0-indexed
                if check_idx < 0:
                    break
                cell = all_col_a[check_idx].lower()
                if not cell:
                    continue
                for metric, keywords in metric_keywords.items():
                    if metric not in blocks and any(kw in cell for kw in keywords):
                        metric_name = metric
                        break
                if metric_name:
                    break

            if not metric_name:
                # Fallback: assign by position order
                ordered = ['spend', 'regs', 'cpa', 'clicks', 'impressions', 'cpc', 'cvr', 'ctr']
                idx = len(blocks)
                if idx < len(ordered):
                    metric_name = ordered[idx]

            if metric_name and metric_name not in blocks:
                blocks[metric_name] = bs

        # If we found fewer blocks than expected, merge with fallback
        if len(blocks) < len(WEEKLY_METRICS_FALLBACK):
            for metric, row in WEEKLY_METRICS_FALLBACK.items():
                if metric not in blocks:
                    blocks[metric] = row

        return blocks

    def _find_weekly_market_row(self, block_start_1indexed, market):
        """Find the 1-indexed row for a market within a Weekly metric block.

        Markets appear in 3-row groups (total, Brand, NB) in WEEKLY_MARKET_ORDER.
        Returns (total_row, brand_row, nb_row) all 1-indexed.
        """
        if market not in WEEKLY_MARKET_ORDER:
            return None
        idx = WEEKLY_MARKET_ORDER.index(market)
        total = block_start_1indexed + (idx * 3)
        return (total, total + 1, total + 2)

    def _read_weekly_tab(self, market, target_week, num_weeks=52):
        """Read the Weekly tab for a market's historical weekly data.

        Auto-detects metric block positions and reads the full column range.
        Returns a list of dicts with weekly metrics going back num_weeks from target_week.
        """
        if 'Weekly' not in self.sheet_names:
            return []
        ws = self.wb['Weekly']

        # Auto-detect block positions
        if not hasattr(self, '_weekly_blocks'):
            self._weekly_blocks = self._detect_weekly_blocks()
        BLOCK_ROWS = self._weekly_blocks

        if market not in WEEKLY_MARKET_ORDER:
            return []

        # Get week column headers from row 1 (read full width)
        row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        # Find target week column
        tw_col = None
        for i, v in enumerate(row1):
            if v and str(v).strip() == target_week:
                tw_col = i
                break
        if tw_col is None:
            return []

        # Determine column range
        start_col = max(1, tw_col - num_weeks + 1)
        week_cols = list(range(start_col, tw_col + 1))

        # Pre-read all needed rows (cache them)
        def get_row(row_num):
            return list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]

        # For each metric, read total/brand/nb rows
        rows_cache = {}
        for metric, block_start in BLOCK_ROWS.items():
            market_rows = self._find_weekly_market_row(block_start, market)
            if market_rows is None:
                continue
            total_row, brand_row, nb_row = market_rows
            try:
                rows_cache[f'{metric}_total'] = get_row(total_row)
                rows_cache[f'{metric}_brand'] = get_row(brand_row)
                rows_cache[f'{metric}_nb'] = get_row(nb_row)
            except Exception:
                pass  # Row may not exist for newer blocks

        # Build weekly data
        weeks = []
        for col in week_cols:
            week_label = str(row1[col]).strip() if col < len(row1) and row1[col] else ''
            if not week_label:
                continue

            def val(key, c):
                r = rows_cache.get(key)
                if r and c < len(r):
                    return safe_float(r[c])
                return 0.0

            w = {
                'week': week_label,
                'spend': val('spend_total', col),
                'brand_spend': val('spend_brand', col),
                'nb_spend': val('spend_nb', col),
                'regs': safe_int(val('regs_total', col)),
                'brand_regs': safe_int(val('regs_brand', col)),
                'nb_regs': safe_int(val('regs_nb', col)),
                'cpa': val('cpa_total', col) or None,
                'brand_cpa': val('cpa_brand', col) or None,
                'nb_cpa': val('cpa_nb', col) or None,
                'clicks': safe_int(val('clicks_total', col)),
                'brand_clicks': safe_int(val('clicks_brand', col)),
                'nb_clicks': safe_int(val('clicks_nb', col)),
                'impressions': safe_int(val('impressions_total', col)),
                'cpc': val('cpc_total', col) or None,
                'nb_cpc': val('cpc_nb', col) or None,
                'brand_cpc': val('cpc_brand', col) or None,
                'cvr': val('cvr_total', col) or None,
                'brand_cvr': val('cvr_brand', col) or None,
                'nb_cvr': val('cvr_nb', col) or None,
                'ctr': val('ctr_total', col) or None,
            }

            # Add any extra detected metrics beyond the standard 8
            for metric in BLOCK_ROWS:
                if metric not in ('spend', 'regs', 'cpa', 'clicks', 'impressions', 'cpc', 'cvr', 'ctr'):
                    w[metric] = val(f'{metric}_total', col) or None
                    w[f'brand_{metric}'] = val(f'{metric}_brand', col) or None
                    w[f'nb_{metric}'] = val(f'{metric}_nb', col) or None

            # Skip weeks with no data
            if w['regs'] > 0 or w['spend'] > 0:
                weeks.append(w)

        return weeks


    def _detect_monthly_metric_blocks(self):
        """Auto-detect metric block positions on the 2026 Monthly tab.

        Scans column A for rows containing 'SEM' to find the start of each
        metric block's actuals section. Each block has OP2 rows above and
        actuals rows below, with markets in 3-row groups (total, Brand, NB).

        Returns dict: {metric_name: {'op2_start': row, 'actuals_start': row, 'col_offset': col}}
        """
        if '2026 Monthly' not in self.sheet_names:
            return {}
        ws = self.wb['2026 Monthly']
        # Read full sheet (BD323 = col 56, row 323)
        all_rows = list(ws.iter_rows(min_row=1, max_row=323, max_col=56, values_only=True))
        self._monthly_all_rows = all_rows  # cache for reuse

        # Known metric keywords to look for in header/label cells
        metric_keywords = {
            'spend': ['spend', 'cost'],
            'regs': ['reg', 'registration'],
            'cpa': ['cpa', 'cost per'],
            'clicks': ['click'],
            'impressions': ['impression', 'impr'],
            'cpc': ['cpc', 'cost per click'],
            'cvr': ['cvr', 'conversion rate'],
            'ctr': ['ctr', 'click through', 'click-through'],
        }

        # Strategy: scan column A (col 0) for rows containing market SEM labels
        # like "US SEM". These mark the start of actuals sections. The metric
        # block identity comes from nearby header rows.
        blocks = {}
        market_label = 'US SEM'  # US is always first market in each block

        # Find all rows where col 0 contains 'US SEM' — each is an actuals block start
        actuals_starts = []
        for i, row in enumerate(all_rows):
            cell = str(row[0]).strip() if row and row[0] else ''
            if cell == market_label:
                actuals_starts.append(i + 1)  # convert to 1-indexed

        # For each actuals start, look upward for the metric block header
        for act_start in actuals_starts:
            # Look at rows above for a header clue (typically 1-3 rows up has the metric name)
            metric_name = None
            for lookback in range(1, 20):
                check_row_idx = act_start - 1 - lookback  # 0-indexed
                if check_row_idx < 0:
                    break
                row = all_rows[check_row_idx]
                # Check first few cells for metric keywords
                for cell_idx in range(min(3, len(row))):
                    cell = str(row[cell_idx]).strip().lower() if row[cell_idx] else ''
                    if not cell:
                        continue
                    for metric, keywords in metric_keywords.items():
                        if metric not in blocks and any(kw in cell for kw in keywords):
                            metric_name = metric
                            break
                    if metric_name:
                        break
                if metric_name:
                    break

            if not metric_name:
                # Fallback: assign by position order (spend, regs, cpa, clicks, imp, cpc, cvr, ctr)
                ordered = ['spend', 'regs', 'cpa', 'clicks', 'impressions', 'cpc', 'cvr', 'ctr']
                idx = len(blocks)
                if idx < len(ordered):
                    metric_name = ordered[idx]

            if metric_name and metric_name not in blocks:
                # Find OP2 section: look for rows above actuals that have budget data
                # OP2 is typically ~15 rows above actuals start
                op2_start = None
                for lookback in range(3, 20):
                    check_idx = act_start - 1 - lookback  # 0-indexed
                    if check_idx < 0:
                        break
                    row = all_rows[check_idx]
                    cell = str(row[0]).strip() if row and row[0] else ''
                    if cell == market_label:
                        # This is the OP2 section's US SEM row
                        op2_start = check_idx + 1  # 1-indexed
                        break

                blocks[metric_name] = {
                    'actuals_start': act_start,
                    'op2_start': op2_start,
                }

        return blocks

    def _find_monthly_market_row(self, block_start_1indexed, market):
        """Find the 1-indexed row for a market within a monthly metric block.

        Markets appear in 3-row groups (total, Brand, NB) in a fixed order.
        """
        market_order = ['US', 'JP', 'UK', 'DE', 'FR', 'IT', 'ES', 'CA', 'MX', 'AU']
        if market not in market_order:
            return None
        idx = market_order.index(market)
        return block_start_1indexed + (idx * 3)

    def _detect_monthly_col_layout(self):
        """Auto-detect column layout for monthly metric blocks.

        Scans header rows to find month columns (Jan-Dec) and their offsets.
        Returns dict with 'months' mapping month labels to column indices.
        """
        if not hasattr(self, '_monthly_all_rows'):
            return {'month_cols': {}}
        rows = self._monthly_all_rows

        monthly_months = [
            '2026 Jan', '2026 Feb', '2026 Mar', '2026 Apr', '2026 May', '2026 Jun',
            '2026 Jul', '2026 Aug', '2026 Sep', '2026 Oct', '2026 Nov', '2026 Dec',
        ]
        month_abbrs = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                       'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

        # Scan first 3 rows for month headers
        month_cols = {}
        for row_idx in range(min(3, len(rows))):
            row = rows[row_idx]
            for col_idx, cell in enumerate(row):
                cell_str = str(cell).strip() if cell else ''
                # Match "Jan", "Feb", etc. or "2026 Jan", etc.
                for j, abbr in enumerate(month_abbrs):
                    if cell_str == abbr or cell_str == monthly_months[j]:
                        if monthly_months[j] not in month_cols:
                            month_cols[monthly_months[j]] = col_idx

        # If we didn't find labeled headers, fall back to known positions
        if not month_cols:
            for j, month in enumerate(monthly_months):
                month_cols[month] = 1 + j  # cols 1-12

        return {'month_cols': month_cols, 'months': monthly_months}

    def _read_monthly_actuals(self):
        """Read 2026 Monthly tab for actual metrics by month, with Brand/NB split.

        Auto-detects metric blocks (spend, regs, CPA, clicks, impressions, CPC,
        CVR, CTR) by scanning the sheet structure. Reads up to BD323.
        """
        if '2026 Monthly' not in self.sheet_names:
            return {}

        blocks = self._detect_monthly_metric_blocks()
        if not blocks:
            return {}

        col_layout = self._detect_monthly_col_layout()
        month_cols = col_layout['month_cols']
        monthly_months = col_layout.get('months', [])
        rows = self._monthly_all_rows

        if not month_cols or not monthly_months:
            return {}

        actuals = {m: {} for m in ALL_MARKETS}

        for metric, block_info in blocks.items():
            act_start = block_info['actuals_start']

            for market in ALL_MARKETS:
                total_row_1idx = self._find_monthly_market_row(act_start, market)
                if total_row_1idx is None or total_row_1idx > len(rows):
                    continue

                total = rows[total_row_1idx - 1]
                brand = rows[total_row_1idx] if total_row_1idx < len(rows) else None
                nb = rows[total_row_1idx + 1] if total_row_1idx + 1 < len(rows) else None

                for month in monthly_months:
                    col = month_cols.get(month)
                    if col is None or col >= len(total):
                        continue

                    val = safe_float(total[col])
                    brand_val = safe_float(brand[col]) if brand and col < len(brand) else 0.0
                    nb_val = safe_float(nb[col]) if nb and col < len(nb) else 0.0

                    if month not in actuals[market]:
                        actuals[market][month] = {}

                    actuals[market][month][metric] = val
                    actuals[market][month][f'brand_{metric}'] = brand_val
                    actuals[market][month][f'nb_{metric}'] = nb_val

        # Derive CPA, CPC, CVR, CTR from raw metrics where not directly available
        for market in ALL_MARKETS:
            for month in monthly_months:
                m = actuals[market].get(month, {})
                if not m:
                    continue
                # Skip months with no spend and no regs
                if m.get('spend', 0) == 0 and m.get('regs', 0) == 0:
                    actuals[market].pop(month, None)
                    continue
                # Ensure regs are ints
                for key in ['regs', 'brand_regs', 'nb_regs', 'clicks', 'brand_clicks',
                            'nb_clicks', 'impressions', 'brand_impressions', 'nb_impressions']:
                    if key in m:
                        m[key] = safe_int(m[key])
                # Derive CPA if not in a dedicated block
                if 'cpa' not in blocks:
                    spend = m.get('spend', 0)
                    regs = m.get('regs', 0)
                    m['cpa'] = spend / regs if regs > 0 else None
                    bs = m.get('brand_spend', 0)
                    br = m.get('brand_regs', 0)
                    m['brand_cpa'] = bs / br if br > 0 else None
                    ns = m.get('nb_spend', 0)
                    nr = m.get('nb_regs', 0)
                    m['nb_cpa'] = ns / nr if nr > 0 else None

        return actuals

    def _read_monthly_budget(self):
        """Read 2026 Monthly tab for OP2 budget targets across all metric blocks."""
        if '2026 Monthly' not in self.sheet_names:
            return {}

        # Ensure blocks are detected (may already be cached from _read_monthly_actuals)
        if not hasattr(self, '_monthly_all_rows'):
            blocks = self._detect_monthly_metric_blocks()
        else:
            blocks = self._detect_monthly_metric_blocks()

        col_layout = self._detect_monthly_col_layout()
        month_cols = col_layout['month_cols']
        monthly_months = col_layout.get('months', [])
        rows = self._monthly_all_rows

        if not month_cols or not monthly_months:
            return {}

        budgets = {m: {} for m in ALL_MARKETS}

        for metric, block_info in blocks.items():
            op2_start = block_info.get('op2_start')
            if op2_start is None:
                continue

            for market in ALL_MARKETS:
                total_row_1idx = self._find_monthly_market_row(op2_start, market)
                if total_row_1idx is None or total_row_1idx > len(rows):
                    continue

                total = rows[total_row_1idx - 1]

                op2_key = f'{metric}_op2'
                if op2_key not in budgets[market]:
                    budgets[market][op2_key] = {}

                for month in monthly_months:
                    col = month_cols.get(month)
                    if col is None or col >= len(total):
                        continue
                    val = safe_float(total[col])
                    if metric in ('regs', 'clicks', 'impressions'):
                        val = safe_int(val)
                    budgets[market][op2_key][month] = val

        # Backward compatibility: also provide spend_op2 and regs_op2 at top level
        for market in ALL_MARKETS:
            if 'spend_op2' not in budgets[market] and budgets[market].get('spend_op2'):
                pass  # already there
            # Ensure the old keys exist for projection code
            if 'spend_op2' in budgets[market]:
                pass
            elif budgets[market]:
                # Already keyed as spend_op2 from the loop above
                pass

        return budgets

    def _read_wow_yoy_tab(self):
        """Read the pre-computed WoW & YoY summary tab."""
        if 'WoW & YoY' not in self.sheet_names:
            return {}
        ws = self.wb['WoW & YoY']
        data = {}
        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=47, values_only=True))

        # Row 3 has TW/LW week labels
        tw_label = str(rows[2][0]) if rows[2][0] else ''
        lw_label = str(rows[2][1]) if rows[2][1] else ''
        data['tw'] = tw_label
        data['lw'] = lw_label

        # Row 64 has TY/LY labels
        if len(rows) > 63:
            ty_label = str(rows[63][0]) if rows[63][0] else ''
            ly_label = str(rows[63][1]) if rows[63][1] else ''
            data['ty'] = ty_label
            data['ly'] = ly_label

        return data

    def _compute_multi_week_trend(self, days, target_week, num_weeks=6):
        """Compute trend over the last N weeks for key metrics."""
        weeks = self._get_unique_weeks(days)
        if target_week not in weeks:
            return []
        idx = weeks.index(target_week)
        start_idx = max(0, idx - num_weeks + 1)
        trend_weeks = weeks[start_idx:idx + 1]

        trend = []
        for w in trend_weeks:
            agg = self._aggregate_week(days, w)
            if agg:
                trend.append(agg)
        return trend

    def _detect_anomalies(self, trend, current_week_agg):
        """Detect anomalies in the current week vs recent trend."""
        if len(trend) < 3 or not current_week_agg:
            return []

        anomalies = []
        # Use the weeks before current as baseline
        baseline = trend[:-1]

        for metric in ['regs', 'cost', 'cpa', 'cvr', 'brand_regs', 'nb_regs', 'brand_cvr', 'nb_cvr', 'cpc']:
            baseline_vals = [w.get(metric) for w in baseline if w.get(metric) is not None]
            current_val = current_week_agg.get(metric)
            if not baseline_vals or current_val is None:
                continue

            avg = sum(baseline_vals) / len(baseline_vals)
            if avg == 0:
                continue

            deviation = (current_val - avg) / abs(avg)

            # Flag if >20% deviation from recent average
            # Suppress noise: require minimum absolute values for small-volume metrics
            min_thresholds = {'nb_regs': 20, 'brand_regs': 20, 'nb_cvr': 0.005}
            min_val = min_thresholds.get(metric, 0)
            if max(abs(current_val), abs(avg)) < min_val:
                continue

            if abs(deviation) > 0.20:
                direction = 'above' if deviation > 0 else 'below'
                anomalies.append({
                    'metric': metric,
                    'current': current_val,
                    'avg': avg,
                    'deviation_pct': deviation,
                    'direction': direction,
                    'weeks_in_baseline': len(baseline_vals),
                })

        # Check for data lag: if last 2 days of the week have <30% of the daily avg
        # of the first 5 days, flag as potential incomplete data
        daily = current_week_agg.get('num_days', 7)
        if daily == 7:
            # We need the raw daily data — check via daily_patterns if available
            pass  # handled in generate_callout_md via daily_patterns

        return anomalies

    def _compute_daily_patterns(self, days, target_week):
        """Analyze day-of-week patterns within the target week."""
        week_days = [d for d in days if d['week'] == target_week]
        if not week_days:
            return {}

        dow_names = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
        patterns = []
        for d in week_days:
            dt = datetime.datetime.strptime(d['date'], '%Y-%m-%d')
            dow = dt.weekday()
            patterns.append({
                'date': d['date'],
                'dow': dow_names[dow],
                'regs': d['regs'],
                'cost': d['cost'],
                'brand_regs': d['brand_regs'],
                'nb_regs': d['nb_regs'],
            })

        # Find best/worst days
        if patterns:
            best_day = max(patterns, key=lambda x: x['regs'])
            worst_day = min(patterns, key=lambda x: x['regs'])
            weekday_avg = sum(p['regs'] for p in patterns[:5]) / min(5, len([p for p in patterns if p['dow'] not in ['Sat', 'Sun']]) or 1)
            weekend_avg = sum(p['regs'] for p in patterns if p['dow'] in ['Sat', 'Sun']) / max(1, len([p for p in patterns if p['dow'] in ['Sat', 'Sun']]))

            return {
                'daily': patterns,
                'best_day': best_day,
                'worst_day': worst_day,
                'weekday_avg_regs': weekday_avg,
                'weekend_avg_regs': weekend_avg,
            }
        return {}

    def _project_month(self, days, target_week, market, budgets):
        """Project month-end performance based on daily data and remaining days."""
        week_days = [d for d in days if d['week'] == target_week]
        if not week_days:
            return None

        # Determine the month
        month_label = week_days[0]['month']  # e.g., '2026 Mar'
        if not month_label:
            return None

        # Get all days in this month so far
        month_days = [d for d in days if d['month'] == month_label and d['regs'] > 0]
        if not month_days:
            return None

        # Parse month to get total days
        try:
            parts = month_label.split()
            year = int(parts[0])
            month_abbr = parts[1]
            month_map = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
                         'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12}
            month_num = month_map.get(month_abbr, 3)
            import calendar
            total_days_in_month = calendar.monthrange(year, month_num)[1]
        except:
            return None

        days_elapsed = len(month_days)
        days_remaining = total_days_in_month - days_elapsed

        mtd_spend = sum(d['cost'] for d in month_days)
        mtd_regs = sum(d['regs'] for d in month_days)
        mtd_brand_regs = sum(d['brand_regs'] for d in month_days)
        mtd_nb_regs = sum(d['nb_regs'] for d in month_days)

        # Use recent daily average (last 7 days) for projection
        recent = month_days[-7:] if len(month_days) >= 7 else month_days
        daily_avg_spend = sum(d['cost'] for d in recent) / len(recent)
        daily_avg_regs = sum(d['regs'] for d in recent) / len(recent)

        projected_spend = mtd_spend + (daily_avg_spend * days_remaining)
        projected_regs = mtd_regs + (daily_avg_regs * days_remaining)
        projected_cpa = projected_spend / projected_regs if projected_regs > 0 else None

        # Get OP2 targets
        op2_spend = None
        op2_regs = None
        if market in budgets:
            op2_spend = budgets[market].get('spend_op2', {}).get(month_label)
            op2_regs = budgets[market].get('regs_op2', {}).get(month_label)

        return {
            'month': month_label,
            'days_elapsed': days_elapsed,
            'days_remaining': days_remaining,
            'total_days': total_days_in_month,
            'mtd_spend': mtd_spend,
            'mtd_regs': mtd_regs,
            'mtd_brand_regs': mtd_brand_regs,
            'mtd_nb_regs': mtd_nb_regs,
            'projected_spend': projected_spend,
            'projected_regs': int(round(projected_regs)),
            'projected_cpa': projected_cpa,
            'op2_spend': op2_spend,
            'op2_regs': op2_regs,
            'vs_op2_spend_pct': pct(projected_spend, op2_spend) if op2_spend else None,
            'vs_op2_regs_pct': pct(projected_regs, op2_regs) if op2_regs else None,
        }


    def analyze_market(self, market, target_week=None):
        """Full analysis for a single market."""
        days = self._read_daily_tab(market)
        if not days:
            return {'market': market, 'error': f'No data found for {market}'}

        if target_week is None:
            target_week = self._detect_latest_week(days)
        if not target_week:
            return {'market': market, 'error': 'Could not detect target week'}

        # Current week
        tw = self._aggregate_week(days, target_week)

        # Last week
        weeks = self._get_unique_weeks(days)
        tw_idx = weeks.index(target_week) if target_week in weeks else -1
        lw_label = weeks[tw_idx - 1] if tw_idx > 0 else None
        lw = self._aggregate_week(days, lw_label) if lw_label else None

        # YoY week
        yoy_label = self._find_yoy_week(target_week)
        yoy = self._aggregate_week(days, yoy_label) if yoy_label else None

        # YoY of last week (for WoW YoY comparison)
        yoy_lw_label = self._find_yoy_week(lw_label) if lw_label else None
        yoy_lw = self._aggregate_week(days, yoy_lw_label) if yoy_lw_label else None

        # Multi-week trend (last 8 weeks)
        trend = self._compute_multi_week_trend(days, target_week, num_weeks=8)

        # Anomalies
        anomalies = self._detect_anomalies(trend, tw)

        # Daily patterns
        daily_patterns = self._compute_daily_patterns(days, target_week)

        # Monthly projection
        budgets = self._read_monthly_budget()
        projection = self._project_month(days, target_week, market, budgets)

        # ie%CCP data
        ieccp_data = self._read_ieccp(target_week)
        market_ieccp = ieccp_data.get(market)

        # Weekly tab historical data (longer horizon than daily-derived trend)
        weekly_history = self._read_weekly_tab(market, target_week, num_weeks=52)

        # WoW changes
        wow = {}
        if tw and lw:
            for metric in ['cost', 'regs', 'cpa', 'clicks', 'impressions', 'cpc', 'cvr', 'ctr',
                           'brand_cost', 'brand_regs', 'brand_cpa', 'brand_clicks', 'brand_cvr',
                           'nb_cost', 'nb_regs', 'nb_cpa', 'nb_clicks', 'nb_cvr']:
                tw_val = tw.get(metric)
                lw_val = lw.get(metric)
                if tw_val is not None and lw_val is not None:
                    wow[metric] = {
                        'tw': tw_val, 'lw': lw_val,
                        'change': tw_val - lw_val,
                        'pct': pct(tw_val, lw_val),
                    }

        # YoY changes
        yoy_changes = {}
        if tw and yoy:
            for metric in ['cost', 'regs', 'cpa', 'clicks', 'impressions', 'cpc', 'cvr',
                           'brand_cost', 'brand_regs', 'brand_cpa',
                           'nb_cost', 'nb_regs', 'nb_cpa', 'nb_cvr']:
                tw_val = tw.get(metric)
                yoy_val = yoy.get(metric)
                if tw_val is not None and yoy_val is not None:
                    yoy_changes[metric] = {
                        'ty': tw_val, 'ly': yoy_val,
                        'change': tw_val - yoy_val,
                        'pct': pct(tw_val, yoy_val),
                    }

        # WoW comparison: this year's WoW vs last year's WoW
        wow_yoy = {}
        if tw and lw and yoy and yoy_lw:
            tw_wow_regs = pct(tw['regs'], lw['regs']) if lw['regs'] > 0 else None
            ly_wow_regs = pct(yoy['regs'], yoy_lw['regs']) if yoy_lw and yoy_lw['regs'] > 0 else None
            wow_yoy['regs_wow_ty'] = tw_wow_regs
            wow_yoy['regs_wow_ly'] = ly_wow_regs

        return {
            'market': market,
            'target_week': target_week,
            'current_week': tw,
            'last_week': lw,
            'yoy_week': yoy,
            'yoy_lw': yoy_lw,
            'wow': wow,
            'yoy_changes': yoy_changes,
            'wow_yoy': wow_yoy,
            'trend': [{'week': t['week'], 'regs': t['regs'], 'cost': t['cost'],
                       'cpa': t.get('cpa'), 'cvr': t.get('cvr'),
                       'brand_regs': t['brand_regs'], 'nb_regs': t['nb_regs']}
                      for t in trend],
            'anomalies': anomalies,
            'daily_patterns': daily_patterns,
            'projection': projection,
            'ieccp': market_ieccp,
            'weekly_history': weekly_history,
        }

    def generate_data_brief(self, analysis):
        """Generate a structured markdown data brief for analyst agents.

        Organized around the questions analysts need to answer, not raw metrics.
        """
        a = analysis
        if 'error' in a:
            return f"# {a['market']} Data Brief — Error\n\n{a['error']}\n"

        market = a['market']
        tw = a['current_week']
        lw = a['last_week']
        wow = a.get('wow', {})
        yoy = a.get('yoy_changes', {})
        trend = a.get('trend', [])
        anomalies = a.get('anomalies', [])
        daily = a.get('daily_patterns', {})
        proj = a.get('projection')
        wow_yoy = a.get('wow_yoy', {})
        ieccp = a.get('ieccp')

        def rpct(v):
            if v is None: return 'N/A'
            r = round(v * 100)
            if r == 0: return 'flat'
            return f'{r:+d}%'

        def rdollar(v):
            if v is None: return 'N/A'
            return f'${round(v):,}'

        def rk(v):
            if v >= 1000: return f'{v/1000:.1f}K'
            return str(v)

        def rspend(v):
            if v is None: return 'N/A'
            if v >= 1000000: return f'${v/1000000:.1f}M'
            if v >= 1000: return f'${round(v/1000):,}K'
            return f'${round(v):,}'

        lines = []
        wk = tw['week'].split()[-1] if tw else '?'
        lines.append(f'# {market} {wk} Data Brief')
        lines.append('')

        # ── Section 1: Headline numbers ──
        lines.append('## Headline numbers')
        if tw:
            lines.append(f'- Registrations: {tw["regs"]} ({rpct(wow.get("regs",{}).get("pct"))} WoW)')
            lines.append(f'- Spend: {rspend(tw["cost"])} ({rpct(wow.get("cost",{}).get("pct"))} WoW)')
            lines.append(f'- CPA: {rdollar(tw.get("cpa"))} ({rpct(wow.get("cpa",{}).get("pct"))} WoW)')
            lines.append(f'- Brand regs: {tw["brand_regs"]} ({rpct(wow.get("brand_regs",{}).get("pct"))} WoW)')
            lines.append(f'- NB regs: {tw["nb_regs"]} ({rpct(wow.get("nb_regs",{}).get("pct"))} WoW)')
        lines.append('')

        # ── Section 1.5: ie%CCP (if available) ──
        if ieccp:
            lines.append('## ie%CCP')
            tw_pct = round(ieccp['tw'] * 100)
            lines.append(f'- This week: {tw_pct}%')
            if ieccp.get('lw'):
                lw_pct = round(ieccp['lw'] * 100)
                lines.append(f'- Last week: {lw_pct}%')
            lines.append(f'- Target: 100%')
            lines.append('')

        # ── Section 2: Registration driver decomposition ──
        lines.append('## Registration drivers (what caused the WoW change?)')
        if tw and lw:
            # Overall: was it CVR or volume?
            cvr_wow = wow.get('cvr', {}).get('pct')
            clicks_wow = wow.get('clicks', {}).get('pct')
            if cvr_wow is not None and clicks_wow is not None:
                if abs(cvr_wow) > abs(clicks_wow):
                    lines.append(f'Primary driver: CVR ({rpct(cvr_wow)} WoW) more than clicks ({rpct(clicks_wow)} WoW)')
                else:
                    lines.append(f'Primary driver: clicks ({rpct(clicks_wow)} WoW) more than CVR ({rpct(cvr_wow)} WoW)')

            # Brand decomposition
            lines.append('')
            lines.append('Brand:')
            lines.append(f'  Regs: {tw["brand_regs"]} vs {lw["brand_regs"]} LW ({rpct(wow.get("brand_regs",{}).get("pct"))})')
            b_cvr = wow.get('brand_cvr', {}).get('pct')
            b_clicks = wow.get('brand_clicks', {}).get('pct')
            b_cpc = wow.get('brand_cpc', {}).get('pct') if wow.get('brand_cpc') else None
            b_imp = None
            if b_cvr is not None: lines.append(f'  CVR: {tw.get("brand_cvr",0):.2%} vs {lw.get("brand_cvr",0):.2%} ({rpct(b_cvr)})')
            if b_clicks is not None: lines.append(f'  Clicks: {tw["brand_clicks"]} vs {lw["brand_clicks"]} ({rpct(b_clicks)})')
            if b_cpc is not None: lines.append(f'  CPC: {rdollar(tw.get("brand_cpc"))} ({rpct(b_cpc)})')
            lines.append(f'  CPA: {rdollar(tw.get("brand_cpa"))} vs {rdollar(lw.get("brand_cpa"))} ({rpct(wow.get("brand_cpa",{}).get("pct"))})')

            # NB decomposition
            lines.append('')
            lines.append('Non-Brand:')
            lines.append(f'  Regs: {tw["nb_regs"]} vs {lw["nb_regs"]} LW ({rpct(wow.get("nb_regs",{}).get("pct"))})')
            n_cvr = wow.get('nb_cvr', {}).get('pct')
            n_clicks = wow.get('nb_clicks', {}).get('pct')
            n_cpc = wow.get('nb_cpc', {}).get('pct') if wow.get('nb_cpc') else None
            if n_cvr is not None: lines.append(f'  CVR: {tw.get("nb_cvr",0):.2%} vs {lw.get("nb_cvr",0):.2%} ({rpct(n_cvr)})')
            if n_clicks is not None: lines.append(f'  Clicks: {tw["nb_clicks"]} vs {lw["nb_clicks"]} ({rpct(n_clicks)})')
            if n_cpc is not None: lines.append(f'  CPC: {rdollar(tw.get("nb_cpc"))} ({rpct(n_cpc)})')
            lines.append(f'  CPA: {rdollar(tw.get("nb_cpa"))} vs {rdollar(lw.get("nb_cpa"))} ({rpct(wow.get("nb_cpa",{}).get("pct"))})')
        lines.append('')

        # ── Section 3: 8-week trend (query hint only — data is in DuckDB) ──
        lines.append('## 8-week trend')
        lines.append(f'<!-- Data: market_trend("{market}", weeks=8) -->')
        lines.append('')

        # ── Section 4: YoY comparison ──
        yoy_has_data = yoy and yoy.get('regs', {}).get('ly', 0) > 0
        if yoy_has_data:
            lines.append('## YoY comparison')
            lines.append(f'- Regs: {tw["regs"]} TY vs {yoy["regs"]["ly"]} LY ({rpct(yoy["regs"]["pct"])})')
            lines.append(f'- Spend: {rspend(tw["cost"])} TY vs {rspend(yoy["cost"]["ly"])} LY ({rpct(yoy["cost"]["pct"])})')
            if yoy.get('brand_regs', {}).get('pct') is not None:
                lines.append(f'- Brand regs: {rpct(yoy["brand_regs"]["pct"])} YoY')
            if yoy.get('nb_regs', {}).get('pct') is not None:
                lines.append(f'- NB regs: {rpct(yoy["nb_regs"]["pct"])} YoY')
            if yoy.get('nb_cpa', {}).get('pct') is not None:
                lines.append(f'- NB CPA: {rdollar(yoy["nb_cpa"]["ty"])} vs {rdollar(yoy["nb_cpa"]["ly"])} LY ({rpct(yoy["nb_cpa"]["pct"])})')
            if wow_yoy.get('regs_wow_ty') is not None and wow_yoy.get('regs_wow_ly') is not None:
                lines.append(f'- WoW pattern: TY {rpct(wow_yoy["regs_wow_ty"])} vs LY {rpct(wow_yoy["regs_wow_ly"])} (same week)')
            lines.append('')

        # ── Section 5: Monthly projection inputs ──
        if proj:
            lines.append('## Monthly projection inputs')
            lines.append(f'- Month: {proj["month"]} ({proj["days_elapsed"]}/{proj["total_days"]} days elapsed, {proj["days_remaining"]} remaining)')
            lines.append(f'- MTD actuals: {rspend(proj["mtd_spend"])} spend, {proj["mtd_regs"]} regs ({proj.get("mtd_brand_regs", "?")} Brand, {proj.get("mtd_nb_regs", "?")} NB)')
            if proj.get('op2_spend'):
                lines.append(f'- OP2 targets: {rspend(proj["op2_spend"])} spend, {rk(proj["op2_regs"])} regs')
                pct_through = proj['days_elapsed'] / proj['total_days']
                expected_mtd_regs = proj['op2_regs'] * pct_through
                expected_mtd_spend = proj['op2_spend'] * pct_through
                lines.append(f'- OP2 pace check: at {pct_through:.0%} through the month, linear OP2 pace would be {rk(int(expected_mtd_regs))} regs and {rspend(expected_mtd_spend)} spend')
                lines.append(f'- MTD vs OP2 pace: {rpct(pct(proj["mtd_regs"], expected_mtd_regs))} regs, {rpct(pct(proj["mtd_spend"], expected_mtd_spend))} spend')
            lines.append(f'- Simple linear projection (ingester estimate, not accounting for seasonality/holidays): {rspend(proj["projected_spend"])} spend, {rk(proj["projected_regs"])} regs, {rdollar(proj["projected_cpa"])} CPA')
            lines.append(f'- NOTE: Analyst should produce the actual projection accounting for weekday/weekend mix, holidays, LY patterns, and known upcoming changes.')
            lines.append('')

        # ── Section 6: Anomalies ──
        if anomalies:
            lines.append('## Anomalies (>20% deviation from recent avg)')
            for an in anomalies:
                metric_name = an['metric'].replace('_', ' ')
                lines.append(f'- {metric_name}: {an["direction"]} avg by {abs(an["deviation_pct"]):.0%} (current: {an["current"]:.2f}, avg: {an["avg"]:.2f})')
            lines.append('')

        # ── Section 7: Weekly tab history (query hints — data is in DuckDB) ──
        weekly_history = a.get('weekly_history', [])
        if weekly_history:
            tw_year = self._get_week_year(a['target_week'])
            tw_wnum = self._get_week_number(a['target_week'])

            if tw_year and tw_wnum:
                # Get last 12 weeks of this year
                ty_weeks = [w for w in weekly_history if self._get_week_year(w['week']) == tw_year]
                ty_recent = ty_weeks[-12:] if len(ty_weeks) > 12 else ty_weeks

                if ty_recent:
                    lines.append('## This year weekly trend (last 12 weeks)')
                    lines.append(f'<!-- Data: market_trend("{market}", weeks=12) -->')
                    lines.append('')

                # Get same period last year
                ly_weeks = []
                for w in weekly_history:
                    wy = self._get_week_year(w['week'])
                    wn = self._get_week_number(w['week'])
                    if wy == tw_year - 1 and wn is not None and abs(wn - tw_wnum) <= 4:
                        ly_weeks.append(w)

                if ly_weeks:
                    lines.append(f'## Last year same period (W{tw_wnum-4} to W{tw_wnum+4})')
                    lines.append(f'<!-- Data: db("SELECT * FROM weekly_metrics WHERE market=\'{market}\' AND week LIKE \'{tw_year - 1}%\' ORDER BY week") -->')
                    lines.append('')

                # Streak detection on key metrics (narrative — keep this)
                streaks = []
                if len(ty_recent) >= 3:
                    # NB CPC streak
                    nb_cpc_vals = [(w['week'].split()[-1], w.get('nb_cpc')) for w in ty_recent if w.get('nb_cpc')]
                    if nb_cpc_vals:
                        cpc_streak = 0
                        cpc_dir = None
                        for i in range(len(nb_cpc_vals) - 1, 0, -1):
                            if nb_cpc_vals[i][1] < nb_cpc_vals[i-1][1]:
                                if cpc_dir is None or cpc_dir == 'declining':
                                    cpc_streak += 1
                                    cpc_dir = 'declining'
                                else:
                                    break
                            elif nb_cpc_vals[i][1] > nb_cpc_vals[i-1][1]:
                                if cpc_dir is None or cpc_dir == 'rising':
                                    cpc_streak += 1
                                    cpc_dir = 'rising'
                                else:
                                    break
                            else:
                                break
                        if cpc_streak >= 3:
                            streaks.append(f'NB CPC {cpc_dir} {cpc_streak} consecutive weeks (${nb_cpc_vals[-cpc_streak-1][1]:.2f} in {nb_cpc_vals[-cpc_streak-1][0]} to ${nb_cpc_vals[-1][1]:.2f} in {nb_cpc_vals[-1][0]})')

                    # CPA streak
                    cpa_vals = [(w['week'].split()[-1], w.get('cpa')) for w in ty_recent if w.get('cpa')]
                    if cpa_vals:
                        cpa_streak = 0
                        cpa_dir = None
                        for i in range(len(cpa_vals) - 1, 0, -1):
                            if cpa_vals[i][1] < cpa_vals[i-1][1]:
                                if cpa_dir is None or cpa_dir == 'declining':
                                    cpa_streak += 1
                                    cpa_dir = 'declining'
                                else:
                                    break
                            elif cpa_vals[i][1] > cpa_vals[i-1][1]:
                                if cpa_dir is None or cpa_dir == 'rising':
                                    cpa_streak += 1
                                    cpa_dir = 'rising'
                                else:
                                    break
                            else:
                                break
                        if cpa_streak >= 3:
                            streaks.append(f'CPA {cpa_dir} {cpa_streak} consecutive weeks ({rdollar(cpa_vals[-cpa_streak-1][1])} in {cpa_vals[-cpa_streak-1][0]} to {rdollar(cpa_vals[-1][1])} in {cpa_vals[-1][0]})')

                    # Regs trend (consecutive up or down)
                    reg_vals = [(w['week'].split()[-1], w['regs']) for w in ty_recent if w['regs'] > 0]
                    if reg_vals:
                        reg_streak = 0
                        reg_dir = None
                        for i in range(len(reg_vals) - 1, 0, -1):
                            if reg_vals[i][1] < reg_vals[i-1][1]:
                                if reg_dir is None or reg_dir == 'declining':
                                    reg_streak += 1
                                    reg_dir = 'declining'
                                else:
                                    break
                            elif reg_vals[i][1] > reg_vals[i-1][1]:
                                if reg_dir is None or reg_dir == 'rising':
                                    reg_streak += 1
                                    reg_dir = 'rising'
                                else:
                                    break
                            else:
                                break
                        if reg_streak >= 3:
                            streaks.append(f'Regs {reg_dir} {reg_streak} consecutive weeks ({reg_vals[-reg_streak-1][1]} in {reg_vals[-reg_streak-1][0]} to {reg_vals[-1][1]} in {reg_vals[-1][0]})')

                if streaks:
                    lines.append('## Detected streaks')
                    for s in streaks:
                        lines.append(f'- {s}')
                    lines.append('')

        # ── Section 8: Daily breakdown (query hint — data is in DuckDB) ──
        if daily and daily.get('daily'):
            wk = tw['week'] if tw else '?'
            lines.append('## Daily breakdown')
            lines.append(f'<!-- Data: db("SELECT * FROM daily_metrics WHERE market=\'{market}\' AND week=\'{wk}\' ORDER BY date") -->')

            # Data lag check (narrative — keep this)
            all_days = daily['daily']
            if len(all_days) >= 5:
                weekday_regs = [d['regs'] for d in all_days[:5]]
                weekday_avg = sum(weekday_regs) / len(weekday_regs)
                last_2 = all_days[-2:]
                last_2_avg = sum(d['regs'] for d in last_2) / len(last_2)
                if weekday_avg > 0 and last_2_avg < weekday_avg * 0.30:
                    lines.append(f'⚠️ DATA LAG: last 2 days avg {last_2_avg:.0f} regs vs weekday avg {weekday_avg:.0f}')
            lines.append('')

        return '\n'.join(lines)

    def generate_callout_md(self, analysis, market_context=None):
        """Generate a markdown callout draft matching the WBR callout style.

        Follows callout-principles.md:
        - Headline: regs, WoW%, spend WoW%, CPA. Monthly projection vs OP2.
        - WoW paragraph: prose, Brand and NB together, attribute to CVR/CPC/clicks.
        - YoY paragraph: spend and regs YoY, Brand vs NB drivers.
        - Note: internal PS context, anomalies, observations.
        - Rounding: whole %, whole dollars, "1.1K" for thousands.
        """
        a = analysis
        if 'error' in a:
            return f"# {a['market']} — Error\n\n{a['error']}\n"

        market = a['market']
        tw = a['current_week']
        lw = a['last_week']
        wow = a['wow']
        yoy = a.get('yoy_changes', {})
        proj = a.get('projection')
        anomalies = a.get('anomalies', [])
        trend = a.get('trend', [])
        daily = a.get('daily_patterns', {})
        wow_yoy_data = a.get('wow_yoy', {})

        # ── Helpers for callout-style rounding ──
        def rpct(v):
            """Round percentage to whole number with sign, e.g. '+33%' or '-7%'."""
            if v is None:
                return 'flat'
            rounded = round(v * 100)
            if rounded == 0:
                return 'flat'
            return f'{rounded:+d}%'

        def rdollar(v):
            """Round to whole dollar, e.g. '$132'."""
            if v is None:
                return 'N/A'
            return f'${round(v):,}'

        def rk(v):
            """Format large numbers as '1.1K', '32.9K', or just the number if <1000."""
            if v >= 10000:
                return f'{v/1000:.1f}K'
            elif v >= 1000:
                return f'{v/1000:.1f}K'
            else:
                return str(v)

        def rspend(v):
            """Format spend as '$139K' or '$68K'."""
            if v is None:
                return 'N/A'
            if v >= 1000000:
                return f'${v/1000000:.1f}M'
            elif v >= 1000:
                return f'${round(v/1000):,}K'
            else:
                return f'${round(v):,}'

        # ── Parse week dates ──
        week_num = tw['week'].split()[-1] if tw else '?'
        date_range_raw = tw.get('date_range', '') if tw else ''
        try:
            parts = date_range_raw.split(' to ')
            d1 = datetime.datetime.strptime(parts[0], '%Y-%m-%d')
            d2 = datetime.datetime.strptime(parts[1], '%Y-%m-%d')
            date_range = f"{d1.strftime('%b')} {d1.day}\u2013{d2.day}"
        except:
            date_range = date_range_raw

        lines = []
        lines.append(f"# {market} \u2014 {week_num} ({date_range})")
        lines.append('')

        # ── Headline paragraph ──
        if tw and wow:
            regs = tw['regs']
            regs_wow_pct = rpct(wow.get('regs', {}).get('pct'))
            spend_wow_pct = rpct(wow.get('cost', {}).get('pct'))
            cpa = tw.get('cpa')

            headline = f"{market} drove {rk(regs)} registrations ({regs_wow_pct} WoW), with {spend_wow_pct} spend WoW"

            # CPA framing
            cpa_wow = wow.get('cpa', {}).get('pct')
            if cpa is not None:
                if cpa_wow is not None and abs(cpa_wow) > 0.03:
                    cpa_direction = 'increased' if cpa_wow > 0 else 'decreased'
                    headline += f". CPA {cpa_direction} to {rdollar(cpa)} ({rpct(cpa_wow)} WoW)"
                else:
                    headline += f". CPA {rdollar(cpa)}"

            headline += '.'

            # Monthly projection
            if proj:
                month_name = proj['month'].split()[-1]
                proj_spend_str = rspend(proj['projected_spend'])
                proj_regs_str = rk(proj['projected_regs'])
                proj_cpa_str = rdollar(proj['projected_cpa']) if proj['projected_cpa'] else 'N/A'

                vs_op2_parts = []
                if proj.get('vs_op2_spend_pct') is not None:
                    vs_op2_parts.append(f"{rpct(proj['vs_op2_spend_pct'])} spend")
                if proj.get('vs_op2_regs_pct') is not None:
                    v = proj['vs_op2_regs_pct']
                    if abs(v) < 0.02:
                        vs_op2_parts.append('flat registrations')
                    else:
                        vs_op2_parts.append(f"{rpct(v)} registrations")

                vs_op2_str = ', '.join(vs_op2_parts)
                headline += f" {month_name} is projected to end at {proj_spend_str} spend and {proj_regs_str} registrations, and {proj_cpa_str} CPA."
                if vs_op2_str:
                    headline += f" (vs. OP2: {vs_op2_str})"

            lines.append(headline)

        # ── WoW paragraph (prose) ──
        if wow:
            br = wow.get('brand_regs', {})
            nb = wow.get('nb_regs', {})
            br_tw, br_lw = br.get('tw', 0), br.get('lw', 0)
            nb_tw, nb_lw = nb.get('tw', 0), nb.get('lw', 0)
            br_pct = br.get('pct')
            nb_pct = nb.get('pct')

            both_up = (br_pct and br_pct > 0.01) and (nb_pct and nb_pct > 0.01)
            both_down = (br_pct and br_pct < -0.01) and (nb_pct and nb_pct < -0.01)

            wow_text = ''
            brand_cvr = wow.get('brand_cvr', {}).get('pct')
            nb_cvr = wow.get('nb_cvr', {}).get('pct')
            clicks_wow = wow.get('clicks', {}).get('pct')
            imp_wow = wow.get('impressions', {}).get('pct')
            nb_cpc_wow = wow.get('nb_cpc', {}).get('pct') if wow.get('nb_cpc') else None

            if both_up:
                if br_tw >= nb_tw:
                    wow_text = f"WoW registrations increased on both Brand ({rpct(br_pct)}) and NB ({rpct(nb_pct)})"
                else:
                    wow_text = f"WoW registrations increased on both NB ({rpct(nb_pct)}) and Brand ({rpct(br_pct)})"
                if brand_cvr is not None and nb_cvr is not None:
                    wow_text += f", driven by CVR increases on both sides ({rpct(brand_cvr)} Brand, {rpct(nb_cvr)} NB)"
                wow_text += '.'
            else:
                # Describe each segment with "On the X side" prose
                seg_parts = []
                # Lead with the segment that has more volume
                segs = [('NB', nb_tw, nb_lw, nb_pct, nb_cvr), ('Brand', br_tw, br_lw, br_pct, brand_cvr)]
                if br_tw >= nb_tw:
                    segs = [('Brand', br_tw, br_lw, br_pct, brand_cvr), ('NB', nb_tw, nb_lw, nb_pct, nb_cvr)]

                for seg_name, s_tw, s_lw, s_pct, s_cvr in segs:
                    verb = 'increased' if s_pct and s_pct > 0 else 'fell' if s_pct and s_pct < -0.01 else 'were flat'
                    part = f"On the {seg_name} side, registrations {verb} {rpct(s_pct)} ({s_lw} to {s_tw})"

                    # Add driver detail
                    drivers = []
                    if s_cvr is not None and abs(s_cvr) > 0.03:
                        drivers.append(f"CVR {rpct(s_cvr)}")
                    seg_clicks = wow.get(f'{"brand" if seg_name == "Brand" else "nb"}_clicks', {}).get('pct')
                    if seg_clicks is not None and abs(seg_clicks) > 0.03:
                        drivers.append(f"clicks {rpct(seg_clicks)}")
                    seg_cpc = wow.get(f'{"brand" if seg_name == "Brand" else "nb"}_cpc', {}).get('pct') if wow.get(f'{"brand" if seg_name == "Brand" else "nb"}_cpc') else None
                    if seg_cpc is not None and abs(seg_cpc) > 0.03:
                        drivers.append(f"CPC {rpct(seg_cpc)}")

                    # Add CPA context
                    seg_cpa_tw = tw.get(f'{"brand" if seg_name == "Brand" else "nb"}_cpa')
                    seg_cpa_lw = lw.get(f'{"brand" if seg_name == "Brand" else "nb"}_cpa') if lw else None
                    if seg_cpa_tw and seg_cpa_lw:
                        seg_cpa_wow = pct(seg_cpa_tw, seg_cpa_lw)
                        if seg_cpa_wow is not None and abs(seg_cpa_wow) > 0.05:
                            cpa_verb = 'rose' if seg_cpa_wow > 0 else 'improved'
                            drivers.append(f"CPA {cpa_verb} to {rdollar(seg_cpa_tw)} ({rpct(seg_cpa_wow)})")

                    if drivers:
                        part += ', ' + ', '.join(drivers[:2])  # limit to 2 drivers per segment

                    seg_parts.append(part)

                wow_text = '. '.join(seg_parts) + '.'

            lines.append(wow_text)

        # ── YoY paragraph (suppress if no meaningful data) ──
        yoy_has_data = yoy and yoy.get('regs', {}).get('ly', 0) > 0
        if yoy_has_data:
            cost_yoy = yoy.get('cost', {})
            regs_yoy = yoy.get('regs', {})

            yoy_text = f"YoY we spent {rpct(cost_yoy.get('pct'))} with {rpct(regs_yoy.get('pct'))} registrations."

            brand_regs_yoy = yoy.get('brand_regs', {})
            nb_regs_yoy = yoy.get('nb_regs', {})
            nb_cpa_yoy = yoy.get('nb_cpa', {})

            breakdown_parts = []
            if brand_regs_yoy.get('pct') is not None and nb_regs_yoy.get('pct') is not None:
                br_yoy_abs = abs(brand_regs_yoy.get('change', 0))
                nb_yoy_abs = abs(nb_regs_yoy.get('change', 0))
                if br_yoy_abs > nb_yoy_abs:
                    breakdown_parts.append(f"Brand regs {rpct(brand_regs_yoy['pct'])} YoY")
                    breakdown_parts.append(f"NB regs {rpct(nb_regs_yoy['pct'])} YoY")
                else:
                    breakdown_parts.append(f"NB regs {rpct(nb_regs_yoy['pct'])} YoY")
                    breakdown_parts.append(f"Brand regs {rpct(brand_regs_yoy['pct'])} YoY")

            if nb_cpa_yoy.get('pct') is not None and nb_cpa_yoy.get('ty') and nb_cpa_yoy.get('ly'):
                breakdown_parts.append(f"NB CPA {rpct(nb_cpa_yoy['pct'])} YoY ({rdollar(nb_cpa_yoy['ly'])} vs {rdollar(nb_cpa_yoy['ty'])})")

            if breakdown_parts:
                yoy_text += ' ' + '. '.join(breakdown_parts) + '.'

            lines.append(yoy_text)

        # ── Note paragraph ──
        note_parts = []

        # NB CVR anomaly
        nb_cvr_tw = tw.get('nb_cvr') if tw else None
        if nb_cvr_tw and trend and len(trend) >= 4:
            recent_nb_cvrs = [t.get('nb_regs', 0) / t.get('nb_clicks', 1) if t.get('nb_clicks', 0) > 0 else 0 for t in trend[:-1]]
            if recent_nb_cvrs:
                avg_nb_cvr = sum(recent_nb_cvrs) / len(recent_nb_cvrs)
                if avg_nb_cvr > 0 and nb_cvr_tw > avg_nb_cvr * 1.20:
                    note_parts.append(f"{week_num} NB CVR of {nb_cvr_tw:.2%} is above the recent average of {avg_nb_cvr:.2%}")
                elif avg_nb_cvr > 0 and nb_cvr_tw < avg_nb_cvr * 0.80:
                    note_parts.append(f"{week_num} NB CVR of {nb_cvr_tw:.2%} is below the recent average of {avg_nb_cvr:.2%}")

        # WoW this year vs last year
        if wow_yoy_data.get('regs_wow_ty') is not None and wow_yoy_data.get('regs_wow_ly') is not None:
            note_parts.append(f"{week_num} WoW last year was {rpct(wow_yoy_data['regs_wow_ly'])} regs")

        # Data lag warning
        if daily and daily.get('daily'):
            all_days = daily.get('daily', [])
            if len(all_days) >= 5:
                weekday_regs = [d['regs'] for d in all_days[:5]]
                weekday_avg = sum(weekday_regs) / len(weekday_regs) if weekday_regs else 0
                last_2 = all_days[-2:]
                if weekday_avg > 0:
                    last_2_avg = sum(d['regs'] for d in last_2) / len(last_2)
                    if last_2_avg < weekday_avg * 0.30:
                        note_parts.append(f"Last 2 days show significantly lower volume (avg {last_2_avg:.0f} vs weekday avg {weekday_avg:.0f}), possible data lag")

        if note_parts:
            lines.append('')
            lines.append('Note: ' + '. '.join(note_parts) + '.')

        lines.append('')

        # ── Supplementary analysis (below the callout) ──
        if len(trend) >= 4:
            lines.append('---')
            lines.append('')
            lines.append('Weekly trend (regs): ' + ' | '.join([f"{t['week'].split()[-1]}: {t['regs']}" for t in trend]))

        if anomalies:
            lines.append('')
            lines.append('Flagged anomalies:')
            for an in anomalies:
                metric_name = an['metric'].replace('_', ' ')
                lines.append(f"- {metric_name} {an['direction']} recent avg by {abs(an['deviation_pct']):.0%}")

        if daily and daily.get('daily'):
            lines.append('')
            lines.append('Daily: ' + ', '.join([f"{d['dow']} {d['regs']}" for d in daily['daily']]))

        lines.append('')
        return '\n'.join(lines)


    def run(self, target_week=None, markets=None, output_dir=None, db_path=None):
        """Main entry point. Analyze all requested markets and produce output."""
        if markets is None:
            markets = ALL_MARKETS

        # Auto-detect week from AU (or first available market)
        if target_week is None:
            for m in markets:
                days = self._read_daily_tab(m)
                if days:
                    target_week = self._detect_latest_week(days)
                    if target_week:
                        break

        if not target_week:
            print("ERROR: Could not detect target week from any market data.")
            return

        print(f"Target week: {target_week}")
        print(f"Markets: {', '.join(markets)}")
        print()

        results = {}
        for market in markets:
            print(f"Analyzing {market}...")
            analysis = self.analyze_market(market, target_week)
            results[market] = analysis

        # Determine output directory
        if output_dir is None:
            week_num = target_week.replace(' ', '-').lower()
            output_dir = f'shared/context/active/callouts'

        os.makedirs(output_dir, exist_ok=True)

        # Write per-market callout drafts
        week_slug = target_week.lower().replace(' ', '-').replace('2026 ', '2026-')
        week_num_only = target_week.split()[-1].lower() if target_week else 'unknown'

        for market in markets:
            analysis = results[market]
            callout_md = self.generate_callout_md(analysis)

            # Write to market subfolder
            market_dir = os.path.join(output_dir, market.lower())
            os.makedirs(market_dir, exist_ok=True)
            callout_path = os.path.join(market_dir, f'{market.lower()}-2026-{week_num_only}.md')
            with open(callout_path, 'w') as f:
                f.write(callout_md)
            print(f"  Wrote {callout_path}")

            # Write per-market data brief for analyst agents
            brief_md = self.generate_data_brief(analysis)
            brief_path = os.path.join(market_dir, f'{market.lower()}-data-brief-2026-{week_num_only}.md')
            with open(brief_path, 'w') as f:
                f.write(brief_md)
            print(f"  Wrote {brief_path}")

        # Read monthly actuals and attach to each market result
        monthly_actuals = self._read_monthly_actuals()
        monthly_budgets = self._read_monthly_budget()
        for market in markets:
            if market in results and 'error' not in results[market]:
                results[market]['monthly_actuals'] = monthly_actuals.get(market, {})
                results[market]['monthly_budget'] = monthly_budgets.get(market, {})

        # Write JSON data extract
        json_dir = os.path.join('shared', 'tools', 'dashboard-ingester', 'data')
        os.makedirs(json_dir, exist_ok=True)
        json_path = os.path.join(json_dir, f'{week_slug}.json')

        # Serialize (handle None and float issues)
        def json_safe(obj):
            if isinstance(obj, float):
                if obj != obj:  # NaN
                    return None
                return round(obj, 6)
            return obj

        with open(json_path, 'w') as f:
            json.dump(results, f, indent=2, default=str)
        print(f"  Wrote {json_path}")

        # Write monthly actuals JSON (standalone, for MBR/flash/reporting use)
        monthly_path = os.path.join(json_dir, 'monthly-actuals.json')
        monthly_combined = {}
        for market in ALL_MARKETS:
            monthly_combined[market] = {
                'actuals': monthly_actuals.get(market, {}),
                'op2': monthly_budgets.get(market, {}),
            }
        with open(monthly_path, 'w') as f:
            json.dump(monthly_combined, f, indent=2, default=str)
        print(f"  Wrote {monthly_path}")

        # Write WW summary
        summary_path = os.path.join(output_dir, f'ww-summary-2026-{week_num_only}.md')
        summary = self._generate_ww_summary(results, target_week)
        with open(summary_path, 'w') as f:
            f.write(summary)
        print(f"  Wrote {summary_path}")

        # Write to DuckDB if path provided
        if db_path:
            self._write_to_duckdb(results, target_week, monthly_actuals, monthly_budgets, db_path)

        print("\nDone.")
        return results

    def _detect_anomalies(self, market, week_data, trend):
        """Flag metrics deviating >20% from 8-week average.

        Pure function — does not write to DB or modify inputs.
        Returns a list of anomaly dicts for any metric whose current value
        deviates more than 20% from the historical average.

        Args:
            market: Market code (e.g. 'AU')
            week_data: Dict with current week metrics (must include 'week' key)
            trend: List of prior-week dicts from market_trend() or weekly_history

        Returns:
            List of anomaly dicts, each containing:
                market, week, metric, value, baseline, deviation_pct, direction
        """
        anomalies = []
        metrics_to_check = ['regs', 'cpa', 'cvr', 'spend', 'clicks']

        for metric in metrics_to_check:
            current = week_data.get(metric)
            if current is None:
                continue

            # Gather historical values, skipping None
            historical = [w.get(metric) for w in trend if w.get(metric) is not None]

            # Skip if fewer than 3 data points
            if len(historical) < 3:
                continue

            baseline = sum(historical) / len(historical)

            # Skip zero baseline (can't compute deviation)
            if baseline == 0:
                continue

            deviation = (current - baseline) / baseline

            if abs(deviation) > 0.20:
                anomalies.append({
                    'market': market,
                    'week': week_data.get('week', ''),
                    'metric': metric,
                    'value': current,
                    'baseline': round(baseline, 2),
                    'deviation_pct': round(deviation * 100, 1),
                    'direction': 'above' if deviation > 0 else 'below',
                })

        return anomalies

    def _write_to_duckdb(self, results, target_week, monthly_actuals, monthly_budgets, db_path):
        """Write all ingested data to DuckDB for persistent, queryable storage."""
        import time
        start_time = time.time()

        try:
            import duckdb
        except ImportError:
            print("  WARN: duckdb not installed, skipping DB write. pip install duckdb")
            return

        # Import query helpers for schema export and data events
        sys.path.insert(0, os.path.expanduser('~/shared/tools/data'))
        from query import schema_export, write_data_event, export_parquet

        con = duckdb.connect(db_path)
        source_file = os.path.basename(self.xlsx_path)
        daily_count = 0
        weekly_count = 0
        monthly_count = 0

        for market, analysis in results.items():
            if 'error' in analysis:
                continue

            # ── Daily metrics ──
            days = self._read_daily_tab(market)
            for d in days:
                try:
                    con.execute("""
                        INSERT OR REPLACE INTO daily_metrics
                        (market, date, week, month, cost, clicks, impressions, regs,
                         brand_cost, brand_clicks, brand_imp, brand_regs,
                         nb_cost, nb_clicks, nb_imp, nb_regs, source_file)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, [market, d['date'], d['week'], d.get('month', ''),
                          d['cost'], d['clicks'], d['impressions'], d['regs'],
                          d['brand_cost'], d['brand_clicks'], d['brand_imp'], d['brand_regs'],
                          d['nb_cost'], d['nb_clicks'], d['nb_imp'], d['nb_regs'],
                          source_file])
                    daily_count += 1
                except Exception as e:
                    pass  # Skip rows with issues

            # ── Weekly metrics (from weekly tab history) ──
            weekly_history = analysis.get('weekly_history', [])
            for w in weekly_history:
                try:
                    con.execute("""
                        INSERT OR REPLACE INTO weekly_metrics
                        (market, week, cost, clicks, impressions, regs,
                         cpa, cpc, cvr, ctr,
                         brand_cost, brand_clicks, brand_regs, brand_cpa, brand_cpc, brand_cvr,
                         nb_cost, nb_clicks, nb_regs, nb_cpa, nb_cpc, nb_cvr,
                         source_file)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, [market, w['week'], w['spend'], w['clicks'], w['impressions'], w['regs'],
                          w.get('cpa'), w.get('cpc'), w.get('cvr'), w.get('ctr'),
                          w.get('brand_spend', 0), w.get('brand_clicks', 0), w.get('brand_regs', 0),
                          w.get('brand_cpa'), w.get('brand_cpc'), w.get('brand_cvr'),
                          w.get('nb_spend', 0), w.get('nb_clicks', 0), w.get('nb_regs', 0),
                          w.get('nb_cpa'), w.get('nb_cpc'), w.get('nb_cvr'),
                          source_file])
                    weekly_count += 1
                except Exception:
                    pass

            # ── IECCP ──
            ieccp = analysis.get('ieccp')
            if ieccp:
                try:
                    con.execute("""
                        INSERT OR REPLACE INTO ieccp (market, week, value)
                        VALUES (?, ?, ?)
                    """, [market, target_week, ieccp['tw']])
                except Exception:
                    pass

            # ── Projections ──
            proj = analysis.get('projection')
            if proj:
                try:
                    con.execute("""
                        INSERT OR REPLACE INTO projections
                        (market, week, month, days_elapsed, total_days,
                         projected_regs, projected_spend, projected_cpa,
                         op2_regs, op2_spend, vs_op2_regs_pct, vs_op2_spend_pct,
                         source)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, [market, target_week, proj['month'],
                          proj['days_elapsed'], proj['total_days'],
                          proj['projected_regs'], proj['projected_spend'], proj['projected_cpa'],
                          proj.get('op2_regs'), proj.get('op2_spend'),
                          proj.get('vs_op2_regs_pct'), proj.get('vs_op2_spend_pct'),
                          'ingester'])
                except Exception:
                    pass

            # ── Anomaly detection ──
            # Build week_data from current_week analysis and use weekly_history as trend
            current_week = analysis.get('current_week', {})
            if current_week and weekly_history:
                # Map ingester field names to the metric names _detect_anomalies expects
                week_data = {
                    'week': target_week,
                    'regs': current_week.get('regs'),
                    'cpa': current_week.get('cpa'),
                    'cvr': current_week.get('cvr'),
                    'spend': current_week.get('cost'),
                    'clicks': current_week.get('clicks'),
                }
                # Build trend from weekly_history, mapping 'spend' key from ingester's 'spend' field
                trend_for_anomaly = []
                for w in weekly_history:
                    trend_for_anomaly.append({
                        'week': w.get('week'),
                        'regs': w.get('regs'),
                        'cpa': w.get('cpa'),
                        'cvr': w.get('cvr'),
                        'spend': w.get('spend'),
                        'clicks': w.get('clicks'),
                    })
                detected = self._detect_anomalies(market, week_data, trend_for_anomaly)
                # Clear previous anomalies for this market+week (idempotent re-runs)
                try:
                    con.execute(
                        "DELETE FROM anomalies WHERE market = ? AND week = ?",
                        [market, target_week],
                    )
                except Exception:
                    pass
                for anom in detected:
                    try:
                        con.execute("""
                            INSERT INTO anomalies
                            (id, market, week, metric, value, baseline, deviation_pct, direction)
                            VALUES (nextval('anomalies_seq'), ?, ?, ?, ?, ?, ?, ?)
                        """, [anom['market'], anom['week'], anom['metric'],
                              anom['value'], anom['baseline'], anom['deviation_pct'],
                              anom['direction']])
                    except Exception as e:
                        print(f"  WARN: anomaly insert failed for {market} {anom['metric']}: {e}")
                if detected:
                    print(f"  {market}: flagged {len(detected)} anomalies")

        # ── Monthly metrics ──
        for market in ALL_MARKETS:
            actuals = monthly_actuals.get(market, {})
            budget = monthly_budgets.get(market, {})
            for month, m in actuals.items():
                try:
                    spend_op2 = budget.get('spend_op2', {}).get(month) if isinstance(budget.get('spend_op2'), dict) else None
                    regs_op2 = budget.get('regs_op2', {}).get(month) if isinstance(budget.get('regs_op2'), dict) else None
                    con.execute("""
                        INSERT OR REPLACE INTO monthly_metrics
                        (market, month, spend, regs, cpa, clicks, impressions,
                         brand_spend, brand_regs, brand_cpa,
                         nb_spend, nb_regs, nb_cpa,
                         spend_op2, regs_op2, source_file)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, [market, month,
                          m.get('spend', 0), safe_int(m.get('regs', 0)), m.get('cpa'),
                          safe_int(m.get('clicks', 0)), safe_int(m.get('impressions', 0)),
                          m.get('brand_spend', 0), safe_int(m.get('brand_regs', 0)), m.get('brand_cpa'),
                          m.get('nb_spend', 0), safe_int(m.get('nb_regs', 0)), m.get('nb_cpa'),
                          spend_op2, safe_int(regs_op2) if regs_op2 else None,
                          source_file])
                    monthly_count += 1
                except Exception:
                    pass

        # ── Ingest log ──
        duration = time.time() - start_time
        try:
            con.execute("""
                INSERT INTO ingest_log
                (id, source_file, markets_processed, target_week,
                 rows_daily, rows_weekly, rows_monthly, duration_seconds)
                VALUES (nextval('ingest_log_seq'), ?, ?, ?, ?, ?, ?, ?)
            """, [source_file, ','.join(results.keys()), target_week,
                  daily_count, weekly_count, monthly_count, round(duration, 2)])
        except Exception:
            pass

        con.close()
        print(f"  DuckDB: wrote {daily_count} daily, {weekly_count} weekly, {monthly_count} monthly rows to {db_path} ({duration:.1f}s)")

        # ── Data event notification (for agent polling) ──
        try:
            write_data_event(
                target_week=target_week,
                markets_processed=list(results.keys()),
                row_counts={'daily': daily_count, 'weekly': weekly_count, 'monthly': monthly_count},
                db_path=db_path,
            )
            print("  Wrote last_ingest.json (agent data event)")
        except Exception as e:
            print(f"  WARN: data event write failed: {e}")

        # ── Parquet export (for cross-environment agent access) ──
        try:
            exported = export_parquet(db_path=db_path)
            if exported:
                print(f"  Parquet: exported {len(exported)} tables to ~/shared/tools/data/exports/")
        except Exception as e:
            print(f"  WARN: parquet export failed: {e}")

        # ── Schema export (portability layer) ──
        try:
            schema_path = schema_export(db_path=db_path)
            print(f"  Schema exported to {schema_path}")
        except Exception as e:
            print(f"  WARN: schema export failed: {e}")

    def _generate_ww_summary(self, results, target_week):
        """Generate a WW summary across all markets."""
        lines = []
        week_num = target_week.split()[-1] if target_week else '?'
        lines.append(f"# WW Summary — {week_num}")
        lines.append('')

        # Summary table
        lines.append('| Market | Regs | WoW% | Spend | WoW% | CPA | WoW% | YoY Regs% | Proj Regs | vs OP2 |')
        lines.append('|--------|------|------|-------|------|-----|------|-----------|-----------|--------|')

        ww_regs = 0
        ww_lw_regs = 0
        ww_spend = 0
        ww_lw_spend = 0

        for market in ALL_MARKETS:
            if market not in results:
                continue
            a = results[market]
            if 'error' in a:
                lines.append(f'| {market} | — | — | — | — | — | — | — | — | — |')
                continue

            tw = a.get('current_week', {})
            wow = a.get('wow', {})
            yoy = a.get('yoy_changes', {})
            proj = a.get('projection')

            regs = tw.get('regs', 0)
            spend = tw.get('cost', 0)
            cpa = tw.get('cpa')
            regs_wow = wow.get('regs', {}).get('pct')
            spend_wow = wow.get('cost', {}).get('pct')
            cpa_wow = wow.get('cpa', {}).get('pct')
            regs_yoy = yoy.get('regs', {}).get('pct')
            # Suppress YoY for markets with no meaningful prior year data
            if yoy.get('regs', {}).get('ly', 0) == 0:
                regs_yoy = None
            proj_regs = proj.get('projected_regs') if proj else None
            vs_op2 = proj.get('vs_op2_regs_pct') if proj else None

            ww_regs += regs
            ww_lw_regs += wow.get('regs', {}).get('lw', 0)
            ww_spend += spend
            ww_lw_spend += wow.get('cost', {}).get('lw', 0)

            lines.append(f"| {market} | {fmt_int(regs)} | {fmt_pct(regs_wow)} | {fmt_dollar(spend)} | {fmt_pct(spend_wow)} | "
                         f"{fmt_dollar(cpa) if cpa else 'N/A'} | {fmt_pct(cpa_wow)} | {fmt_pct(regs_yoy)} | "
                         f"{fmt_int(proj_regs) if proj_regs else 'N/A'} | {fmt_pct(vs_op2) if vs_op2 is not None else 'N/A'} |")

        # WW totals
        ww_cpa = ww_spend / ww_regs if ww_regs > 0 else None
        ww_regs_wow = pct(ww_regs, ww_lw_regs)
        ww_spend_wow = pct(ww_spend, ww_lw_spend)
        lines.append(f"| **WW** | **{fmt_int(ww_regs)}** | **{fmt_pct(ww_regs_wow)}** | **{fmt_dollar(ww_spend)}** | "
                     f"**{fmt_pct(ww_spend_wow)}** | **{fmt_dollar(ww_cpa) if ww_cpa else 'N/A'}** | — | — | — | — |")

        lines.append('')

        # Key callouts
        lines.append('## Key Callouts')
        lines.append('')

        # Find biggest movers
        movers = []
        for market in ALL_MARKETS:
            if market not in results or 'error' in results[market]:
                continue
            wow = results[market].get('wow', {})
            regs_pct = wow.get('regs', {}).get('pct')
            if regs_pct is not None:
                movers.append((market, regs_pct))

        movers.sort(key=lambda x: x[1])
        if movers:
            best = movers[-1]
            worst = movers[0]
            lines.append(f"- Biggest gainer: {best[0]} ({fmt_pct(best[1])} regs WoW)")
            lines.append(f"- Biggest decliner: {worst[0]} ({fmt_pct(worst[1])} regs WoW)")

        # Anomalies across markets
        all_anomalies = []
        for market in ALL_MARKETS:
            if market not in results or 'error' in results[market]:
                continue
            for an in results[market].get('anomalies', []):
                all_anomalies.append((market, an))

        if all_anomalies:
            lines.append('')
            lines.append('## Anomalies')
            lines.append('')
            for market, an in all_anomalies:
                metric_name = an['metric'].replace('_', ' ').title()
                lines.append(f"- {market}: {metric_name} {an['direction']} recent avg by {abs(an['deviation_pct']):.0%}")

        lines.append('')
        return '\n'.join(lines)


def main():
    parser = argparse.ArgumentParser(description='WW Dashboard Ingester')
    parser.add_argument('xlsx_path', help='Path to the WW Dashboard xlsx file')
    parser.add_argument('--week', help='Target week (e.g., "2026 W12"). Auto-detects if omitted.')
    parser.add_argument('--markets', help='Comma-separated market list (e.g., AU,MX). All if omitted.')
    parser.add_argument('--output-dir', help='Output directory for callout files.')
    parser.add_argument('--db', default=os.path.expanduser('~/shared/tools/data/ps-analytics.duckdb'),
                        help='DuckDB database path. Set to "none" to skip DB writes.')
    args = parser.parse_args()

    markets = args.markets.split(',') if args.markets else None
    db_path = None if args.db == 'none' else args.db

    ingester = DashboardIngester(args.xlsx_path)
    ingester.run(target_week=args.week, markets=markets, output_dir=args.output_dir,
                 db_path=db_path)


if __name__ == '__main__':
    main()
