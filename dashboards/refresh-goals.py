#!/usr/bin/env python3
"""Refresh goals data from Brandon's Outbound Marketing Goals sheet.

Downloads the SharePoint xlsx each run via the sharepoint-mcp helper, parses
the 2026 Goals tab into a structured JSON artifact consumed by the dashboards.

No manual local file — source of truth is Brandon's sheet. Runs as part of
refresh-all.py.
"""
from __future__ import annotations

import json
import re
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path.home()
OUT = ROOT / "shared" / "dashboards" / "data" / "goals-data.json"
DOWNLOAD = Path("/tmp/outbound-marketing-goals.xlsx")

SITE_URL = "https://amazon-my.sharepoint.com/personal/brandoxy_amazon_com"
SERVER_RELATIVE_URL = "/personal/brandoxy_amazon_com/Documents/Outbound Marketing Goals.xlsx"

RICHARD = "Richard Williams"

# Map Richard's 3 goals to market state-file keys. Used by the dashboard to
# surface the right goal card at the top of each market page.
GOAL_MARKET_MAP = {
    "Globalized Cross-Market Testing": ["ww"],
    "MX/AU Paid Search Registrations": ["mx", "au"],
    "MX + AU Market Testing": ["mx", "au"],
}


def _fetch_op2_targets_from_duckdb(markets: list[str], fiscal_year: int = 2026) -> tuple[dict, str]:
    """Pull annual registration targets per market.

    Strategy:
      1. Primary — read from projection-data.json `markets[m].op2_targets.annual_regs_target`.
         That file is already refreshed by export-projection-data.py (which
         does the MotherDuck query with proper credentials) so we don't
         duplicate the MotherDuck connection here.
      2. Fallback — try a local DuckDB read if somehow projection-data.json
         is missing. Usually won't work (local DB lacks ps schema) but
         worth trying for operator-run standalone cases.

    Returns ({market: annual_target_int}, source_label).
    """
    proj_path = Path.home() / "shared" / "dashboards" / "data" / "projection-data.json"
    if proj_path.exists():
        try:
            pd = json.loads(proj_path.read_text())
            out = {}
            for m in markets:
                t = pd.get("markets", {}).get(m, {}).get("op2_targets", {})
                if t and t.get("annual_regs_target"):
                    out[m] = int(round(t["annual_regs_target"]))
            if out:
                return out, "ps.targets via projection-data.json"
        except Exception as e:
            print(f"warn: failed reading projection-data.json op2_targets: {e}", file=sys.stderr)

    # Fallback — try local DuckDB directly. Usually fails (local has no
    # ps schema) but sometimes works for standalone operator runs.
    try:
        import duckdb
        db = Path.home() / "shared" / "data" / "duckdb" / "ps-analytics.duckdb"
        if not db.exists():
            return {}, "no source available"
        conn = duckdb.connect(str(db), read_only=True)
        markets_csv = ",".join(f"'{m}'" for m in markets)
        sql = f"""
            SELECT market, SUM(target_value) AS annual
            FROM ps.targets
            WHERE fiscal_year = {fiscal_year}
              AND metric_name = 'registrations'
              AND period_type = 'monthly'
              AND channel = 'ps'
              AND market IN ({markets_csv})
            GROUP BY market
        """
        rows = conn.execute(sql).fetchall()
        return {r[0]: int(round(r[1])) for r in rows}, "ps.targets (local duckdb)"
    except Exception:
        return {}, "no source available"


def compute_progress(goals):
    """Compute YTD progress per goal by reading live data sources.

    Goal 1 (Globalized Cross-Market Testing, 3 tests):
      Count distinct cross-team tests in WW testing state file Appendix D
      that match the 'non-standard integration' criterion — i.e. tests
      where Richard is driving a cross-team partnership (Polaris, Email
      Overlay, Baloo, Enhanced Match, BFCM promo, AI Max).

    Goal 2 (MX/AU Paid Search Registrations):
      YTD + projected regs come from forecast-data.json (same source the
      Weekly Review performance page consumes — single source of truth).
      Targets come from ps.targets in MotherDuck — same table the pacing
      dashboards read, so when OP2 gets revised the goal target tracks
      automatically. Falls back to the hardcoded sheet values (11,100 MX /
      12,900 AU) when ps.targets is unreachable so the page never shows
      blank numbers.

    Goal 3 (MX + AU Market Testing, 4 experiments):
      Count tests where market is MX or AU in WW testing state file +
      ps-testing-dashboard (tracked tests).
    """
    progress = {}
    forecast_path = Path.home() / "shared" / "dashboards" / "data" / "forecast-data.json"
    ww_state_path = Path.home() / "shared" / "wiki" / "state-files" / "ww-testing-state.md"

    # Goal 2: registrations — forecast-data.json (same as Weekly Review) +
    # ps.targets (canonical OP2 source, same as pacing dashboards).
    if forecast_path.exists():
        try:
            fd = json.loads(forecast_path.read_text())
            forecast_generated = fd.get("generated", "unknown")
            mx_yr = fd.get("yearly", {}).get("MX", {})
            au_yr = fd.get("yearly", {}).get("AU", {})
            mx_ytd = mx_yr.get("actual_regs_ytd") or 0
            au_ytd = au_yr.get("actual_regs_ytd") or 0
            mx_eoy_pred = mx_yr.get("pred_regs_annual") or 0
            au_eoy_pred = au_yr.get("pred_regs_annual") or 0
            mx_ci_lo = mx_yr.get("pred_regs_annual_ci_lo") or None
            mx_ci_hi = mx_yr.get("pred_regs_annual_ci_hi") or None
            au_ci_lo = au_yr.get("pred_regs_annual_ci_lo") or None
            au_ci_hi = au_yr.get("pred_regs_annual_ci_hi") or None

            # Pull targets from ps.targets via projection-data.json (primary)
            # or local duckdb (fallback). Surface which source won so the UI
            # can show target provenance at a glance.
            db_targets, target_source_label = _fetch_op2_targets_from_duckdb(["MX", "AU"], fiscal_year=2026)
            if db_targets.get("MX") and db_targets.get("AU"):
                mx_target = db_targets["MX"]
                au_target = db_targets["AU"]
                target_source = target_source_label
            else:
                mx_target = 11100
                au_target = 12900
                target_source = f"hardcoded fallback ({target_source_label})"

            combined_target = mx_target + au_target
            combined_ytd = mx_ytd + au_ytd
            combined_pred = mx_eoy_pred + au_eoy_pred

            progress["MX/AU Paid Search Registrations"] = {
                "unit": "registrations",
                "target": combined_target,
                "ytd": combined_ytd,
                "projected_eoy": combined_pred,
                "projected_eoy_ci_lo": (mx_ci_lo + au_ci_lo) if (mx_ci_lo and au_ci_lo) else None,
                "projected_eoy_ci_hi": (mx_ci_hi + au_ci_hi) if (mx_ci_hi and au_ci_hi) else None,
                "ytd_pct": round(combined_ytd / combined_target * 100, 1) if combined_target else 0,
                "projected_pct": round(combined_pred / combined_target * 100, 1) if combined_target else 0,
                "breakdown": [
                    {"market": "MX", "target": mx_target, "ytd": mx_ytd,
                     "projected_eoy": mx_eoy_pred,
                     "projected_eoy_ci_lo": mx_ci_lo, "projected_eoy_ci_hi": mx_ci_hi,
                     "ytd_pct": round(mx_ytd / mx_target * 100, 1) if mx_target else 0,
                     "projected_pct": round(mx_eoy_pred / mx_target * 100, 1) if mx_target else 0},
                    {"market": "AU", "target": au_target, "ytd": au_ytd,
                     "projected_eoy": au_eoy_pred,
                     "projected_eoy_ci_lo": au_ci_lo, "projected_eoy_ci_hi": au_ci_hi,
                     "ytd_pct": round(au_ytd / au_target * 100, 1) if au_target else 0,
                     "projected_pct": round(au_eoy_pred / au_target * 100, 1) if au_target else 0},
                ],
                "forecast_source": "forecast-data.json (same as Weekly Review)",
                "forecast_generated": forecast_generated,
                "target_source": target_source,
            }
        except Exception as e:
            print(f"warn: failed to compute reg progress: {e}", file=sys.stderr)

    # Goals 1 and 3: test counts — parse WW testing state file Appendix D
    # and classify each test by (a) cross-team partnership, (b) MX/AU market.
    goal1_tests = []
    goal3_tests = []
    if ww_state_path.exists():
        raw = ww_state_path.read_text()
        # Find Appendix D section
        m = re.search(r"## Appendix D.*?\n(.*?)(?=^## |\Z)", raw, re.DOTALL | re.MULTILINE)
        if m:
            block = m.group(1)
            for row in re.finditer(r"^\|\s*([^|]+?)\s*\|\s*([^|]+?)\s*\|\s*([^|]+?)\s*\|\s*([^|]+?)\s*\|\s*([^|]+?)\s*\|\s*([^|]+?)\s*\|\s*$", block, re.MULTILINE):
                workstream, test, market, result, confidence, status = [x.strip() for x in row.groups()]
                # Skip the header row
                if workstream.lower() in {"workstream", ":---"}:
                    continue
                if "---" in workstream:
                    continue
                entry = {
                    "workstream": workstream,
                    "test": test,
                    "market": market,
                    "result": result,
                    "confidence": confidence,
                    "status": status,
                }
                status_l = status.lower()
                result_l = result.lower()
                is_complete = any(k in status_l for k in ["complete", "scaled", "validated", "killed"])
                # Goal 1: cross-team / non-standard integration test
                # Heuristic: UX workstream Polaris/Baloo/Email Overlay, or Audiences
                # cross-team tests. Anything with a WW market scope.
                is_cross_team = (
                    "polaris" in test.lower()
                    or "baloo" in test.lower()
                    or "email overlay" in test.lower()
                    or "in-context registration" in test.lower()
                    or "gated guest" in test.lower()
                    or "enhanced match" in test.lower()
                    or "engagement channel" in test.lower()
                )
                # Goal 3: MX or AU market
                is_market_test = market.upper() in {"MX", "AU"}
                if is_cross_team:
                    entry["is_complete"] = is_complete
                    goal1_tests.append(entry)
                if is_market_test:
                    entry["is_complete"] = is_complete
                    goal3_tests.append(entry)

    progress["Globalized Cross-Market Testing"] = {
        "unit": "cross-team tests",
        "target": 3,
        "ytd": sum(1 for t in goal1_tests if t.get("is_complete")),
        "tests_in_flight": len(goal1_tests),
        "tests": goal1_tests,
        "ytd_pct": round(sum(1 for t in goal1_tests if t.get("is_complete")) / 3 * 100, 1),
        "source": "ww-testing-state.md Appendix D (cross-team tests)",
    }
    progress["MX + AU Market Testing"] = {
        "unit": "MX/AU experiments",
        "target": 4,
        "ytd": sum(1 for t in goal3_tests if t.get("is_complete")),
        "tests_in_flight": len(goal3_tests),
        "tests": goal3_tests,
        "ytd_pct": round(sum(1 for t in goal3_tests if t.get("is_complete")) / 4 * 100, 1),
        "source": "ww-testing-state.md Appendix D (MX/AU-scoped tests)",
    }
    return progress


def download_sheet() -> bool:
    """Check cached copy exists. If missing, skip gracefully instead of failing
    the pipeline — the agent or hook that runs refresh-all.py is responsible
    for refreshing /tmp/outbound-marketing-goals.xlsx via SharePoint MCP.
    """
    if not DOWNLOAD.exists():
        print(
            "WARN: /tmp/outbound-marketing-goals.xlsx missing. "
            "Skipping goals refresh — agent should download via sharepoint_read_file "
            "before the next pipeline run.",
            file=sys.stderr,
        )
        return False
    return True


def normalize_status(raw: str) -> dict:
    """Extract color + narrative from a status cell like 'Green - text here'."""
    if not raw:
        return {"color": None, "text": ""}
    text = str(raw).strip()
    m = re.match(r"^(Green|Yellow|Red|Blue)\s*[-\u2013\u2014:]\s*(.*)$", text, re.IGNORECASE)
    if m:
        return {"color": m.group(1).title(), "text": m.group(2).strip()}
    return {"color": None, "text": text}


def parse_goals(xlsx_path: Path) -> dict:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["2026 Goals"]

    # Header row (row 1) has: Link, Status, Owner, Level, Title, Description,
    # End Date, How It's Measured, Why Is This Goal Important?, Primary Benefit,
    # Risks, Dependencies, March, February, January
    header = [str(c.value).strip() if c.value else "" for c in ws[1]]

    # Locate month columns dynamically (they may expand each month).
    month_cols = {}
    for idx, name in enumerate(header):
        if name in {
            "January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December",
        }:
            month_cols[name] = idx

    goals = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        owner = row[2]
        title = row[4]
        if not owner or not title:
            continue

        # Monthly status updates — keep as ordered list, newest first.
        monthly = []
        # Order columns by their actual appearance (most recent usually first in header).
        month_order = sorted(
            month_cols.items(),
            key=lambda kv: list(header).index(kv[0]),
        )
        for month, idx in month_order:
            raw = row[idx] if idx < len(row) else None
            status = normalize_status(raw)
            if status["text"] or status["color"]:
                monthly.append({
                    "month": month,
                    "color": status["color"],
                    "text": status["text"],
                })

        end_date = row[6]
        if isinstance(end_date, datetime):
            end_date = end_date.date().isoformat()
        elif end_date:
            end_date = str(end_date)

        goals.append({
            "owner": str(owner).strip(),
            "level": str(row[3]).strip() if row[3] else "",
            "title": str(title).strip(),
            "description": str(row[5]).strip() if row[5] else "",
            "end_date": end_date,
            "how_measured": str(row[7]).strip() if row[7] else "",
            "why_important": str(row[8]).strip() if row[8] else "",
            "submission_status": str(row[1]).strip() if row[1] else "",
            "monthly_status": monthly,
            "latest_color": monthly[0]["color"] if monthly else None,
            "latest_update": monthly[0]["text"] if monthly else "",
            "latest_month": monthly[0]["month"] if monthly else None,
            "markets": GOAL_MARKET_MAP.get(str(title).strip(), []),
        })

    # Bucket by owner for easy lookup.
    by_owner = {}
    for g in goals:
        by_owner.setdefault(g["owner"], []).append(g)

    # Trend summary per goal — count green/yellow/red across monthly updates.
    for g in goals:
        counts = {"Green": 0, "Yellow": 0, "Red": 0}
        for m in g["monthly_status"]:
            c = m.get("color")
            if c in counts:
                counts[c] += 1
        g["trend"] = counts

    return {
        "generated": datetime.now(timezone.utc).isoformat(),
        "source": "Outbound Marketing Goals.xlsx",
        "owner_count": len(by_owner),
        "goal_count": len(goals),
        "richard_goals": by_owner.get(RICHARD, []),
        "goals": goals,
        "owners": sorted(by_owner.keys()),
        "by_owner": by_owner,
        "progress": compute_progress(by_owner.get(RICHARD, [])),
    }


def main():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    if not download_sheet():
        # Non-fatal — emit empty payload so consumer dashboards degrade gracefully.
        OUT.write_text(json.dumps({
            "generated": datetime.now(timezone.utc).isoformat(),
            "source": "Outbound Marketing Goals.xlsx",
            "error": "sheet_not_downloaded",
            "owner_count": 0,
            "goal_count": 0,
            "richard_goals": [],
            "goals": [],
            "owners": [],
            "by_owner": {},
        }, indent=2))
        print(f"wrote empty {OUT} (source missing)")
        return 0
    data = parse_goals(DOWNLOAD)
    OUT.write_text(json.dumps(data, indent=2, ensure_ascii=False))
    print(f"wrote {OUT}")
    print(f"  owners: {data['owner_count']}  goals: {data['goal_count']}")
    print(f"  Richard's goals: {len(data['richard_goals'])}")
    for g in data["richard_goals"]:
        trend = "/".join(
            m["color"][0] if m.get("color") else "-" for m in g["monthly_status"]
        )
        print(f"    [{trend}] {g['title']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
