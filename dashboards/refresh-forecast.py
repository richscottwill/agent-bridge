#!/usr/bin/env python3
"""
Refresh forecast-data.json from ps-forecast-tracker.xlsx on SharePoint.

Usage:
  1. Drop updated xlsx into shared/dashboards/ (or let this script pull from SharePoint)
  2. Run: python3 refresh-forecast.py
  3. Dashboard at localhost:8080/forecast-tracker.html auto-loads fresh data.

Source: OneDrive > Kiro-Drive/ps-forecast-tracker.xlsx (_Data sheet)
"""
import json, os, sys
from datetime import datetime, timezone

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, "ps-forecast-tracker.xlsx")
JSON_PATH = os.path.join(SCRIPT_DIR, "data", "forecast-data.json")

# Also check /tmp in case SharePoint MCP just downloaded it
FALLBACK_PATH = "/tmp/ps-forecast-tracker.xlsx"

def find_xlsx():
    if os.path.exists(XLSX_PATH):
        return XLSX_PATH
    if os.path.exists(FALLBACK_PATH):
        return FALLBACK_PATH
    print(f"ERROR: Cannot find {XLSX_PATH} or {FALLBACK_PATH}")
    print("Download from SharePoint first, or copy the file to shared/dashboards/")
    sys.exit(1)

def main():
    path = find_xlsx()
    print(f"Reading: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)

    if "_Data" not in wb.sheetnames:
        print("ERROR: _Data sheet not found. Available sheets:", wb.sheetnames)
        sys.exit(1)

    ws = wb["_Data"]
    # Parse header row
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}

    required = ["Market", "Week", "Wk#", "Actual_Regs", "Actual_Cost", "Actual_CPA"]
    for r in required:
        if r not in col:
            print(f"ERROR: Missing column '{r}'. Found: {headers}")
            sys.exit(1)

    # Parse all data rows into weekly, monthly, quarterly buckets
    weekly = {}
    monthly = {}
    quarterly = {}
    markets_set = set()
    max_wk = 0

    MONTHS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    QUARTERS = ['Q1','Q2','Q3','Q4']

    for row in ws.iter_rows(min_row=2, values_only=True):
        market = row[col["Market"]]
        if not market:
            continue

        wk_raw = row[col["Wk#"]]
        week_label = str(row[col["Week"]] or "")
        regs = row[col["Actual_Regs"]]
        cost = row[col["Actual_Cost"]]
        cpa = row[col["Actual_CPA"]]

        # Skip header-like rows
        if regs is not None and isinstance(regs, str) and not regs.replace('.','').replace('-','').isdigit():
            continue

        # Check for prediction data
        pred_regs_val = row[col.get("Pred_Regs", -1)] if "Pred_Regs" in col else None
        ci_lo_val = row[col.get("CI_Lo", -1)] if "CI_Lo" in col else None
        op2_val = row[col.get("OP2_Regs", -1)] if "OP2_Regs" in col else None

        # Skip rows with no actuals AND no predictions
        has_actuals = regs is not None or cost is not None
        has_predictions = pred_regs_val is not None or op2_val is not None
        if not has_actuals and not has_predictions:
            continue

        markets_set.add(market)

        def safe_int(v):
            try: return int(v) if v is not None else None
            except (ValueError, TypeError): return None

        def safe_float(v):
            try: return round(float(v), 2) if v is not None else None
            except (ValueError, TypeError): return None

        # Route to monthly bucket
        if week_label in MONTHS:
            if market not in monthly:
                monthly[market] = []
            monthly[market].append({
                "period": week_label,
                "actual_regs": safe_int(regs),
                "actual_cost": safe_float(cost),
                "actual_cpa": safe_float(cpa),
                "pred_regs": safe_int(pred_regs_val),
                "ci_lo": safe_int(ci_lo_val),
            })
            continue

        # Route to quarterly bucket
        if week_label in QUARTERS:
            if market not in quarterly:
                quarterly[market] = []
            quarterly[market].append({
                "period": week_label,
                "actual_regs": safe_int(regs),
                "actual_cost": safe_float(cost),
                "actual_cpa": safe_float(cpa),
                "pred_regs": safe_int(pred_regs_val),
                "ci_lo": safe_int(ci_lo_val),
            })
            continue

        # Skip other summary rows (YTD, Total, etc.)
        if week_label and any(x in week_label for x in ['YTD','Total']):
            continue

        try:
            wk = int(wk_raw) if wk_raw is not None else 0
        except (ValueError, TypeError):
            wk = 0

        # Skip summary/total rows (Excel date serials or wk > 52)
        if wk > 52 or wk < 1:
            continue

        if wk > max_wk and (regs is not None or cost is not None):
            max_wk = wk

        if market not in weekly:
            weekly[market] = []

        entry = {
            "wk": wk,
            "week": f"2026 {week_label}" if week_label and not week_label.startswith("2026") else (week_label or f"W{wk}"),
            "regs": safe_int(regs) or 0,
            "cost": round(float(cost), 0) if cost else 0,
            "cpa": safe_float(cpa) or 0,
        }

        # Add prediction data if available
        pred_regs = row[col.get("Pred_Regs", -1)] if "Pred_Regs" in col else None
        ci_lo = row[col.get("CI_Lo", -1)] if "CI_Lo" in col else None
        ci_hi = row[col.get("CI_Hi", -1)] if "CI_Hi" in col else None
        op2 = row[col.get("OP2_Regs", -1)] if "OP2_Regs" in col else None
        brand = row[col.get("Brand_Regs", -1)] if "Brand_Regs" in col else None
        nb = row[col.get("NB_Regs", -1)] if "NB_Regs" in col else None

        if pred_regs is not None:
            entry["pred_regs"] = int(pred_regs)
        if ci_lo is not None:
            entry["ci_lo"] = int(ci_lo)
        if ci_hi is not None:
            entry["ci_hi"] = int(ci_hi)
        if op2 is not None:
            entry["op2_regs"] = int(op2)
        if brand is not None:
            entry["brand_regs"] = int(brand)
        if nb is not None:
            entry["nb_regs"] = int(nb)

        weekly[market].append(entry)

    # Sort each market by week number
    for m in weekly:
        weekly[m].sort(key=lambda x: x["wk"])

    # Note: TY click enrichment happens after _Daily_Data loading below

    # Derive predicted spend/CPA for future weeks using trailing 4-week avg CPA
    for m in weekly:
        rows = weekly[m]
        # Collect actual CPAs for trailing average
        actual_cpas = [r["cpa"] for r in rows if r["cpa"] > 0 and r["regs"] > 0]
        if not actual_cpas:
            continue
        # Use last 4 weeks' CPA as baseline for predictions
        trailing_cpa = round(sum(actual_cpas[-4:]) / len(actual_cpas[-4:]), 2)

        for r in rows:
            has_actual = r["regs"] > 0 or r["cost"] > 0
            pred_regs = r.get("pred_regs")
            op2_regs = r.get("op2_regs")

            if not has_actual and pred_regs:
                r["pred_cost"] = round(pred_regs * trailing_cpa, 0)
                r["pred_cpa"] = trailing_cpa
            if op2_regs:
                r["op2_cost"] = round(op2_regs * trailing_cpa, 0)
                r["op2_cpa"] = trailing_cpa

        # Store trailing CPA for reference
        weekly[m] = rows  # already sorted

    markets = sorted(markets_set)

    # Week-to-month mapping (approximate: W1-4=Jan, W5-8=Feb, W9-13=Mar, W14-17=Apr, etc.)
    # More precise: use week_start dates if available, else approximate
    WEEK_TO_MONTH = {}
    for w in range(1, 53):
        if w <= 4: WEEK_TO_MONTH[w] = 'Jan'
        elif w <= 8: WEEK_TO_MONTH[w] = 'Feb'
        elif w <= 13: WEEK_TO_MONTH[w] = 'Mar'
        elif w <= 17: WEEK_TO_MONTH[w] = 'Apr'
        elif w <= 21: WEEK_TO_MONTH[w] = 'May'
        elif w <= 26: WEEK_TO_MONTH[w] = 'Jun'
        elif w <= 30: WEEK_TO_MONTH[w] = 'Jul'
        elif w <= 34: WEEK_TO_MONTH[w] = 'Aug'
        elif w <= 39: WEEK_TO_MONTH[w] = 'Sep'
        elif w <= 43: WEEK_TO_MONTH[w] = 'Oct'
        elif w <= 47: WEEK_TO_MONTH[w] = 'Nov'
        else: WEEK_TO_MONTH[w] = 'Dec'

    MONTH_TO_Q = {'Jan':'Q1','Feb':'Q1','Mar':'Q1','Apr':'Q2','May':'Q2','Jun':'Q2',
                  'Jul':'Q3','Aug':'Q3','Sep':'Q3','Oct':'Q4','Nov':'Q4','Dec':'Q4'}

    # Enrich monthly/quarterly with cost+CPA aggregated from weekly data
    for m in markets:
        weeks = weekly.get(m, [])
        # Aggregate weekly cost by month
        month_cost = {}
        month_regs_actual = {}
        for w in weeks:
            mo = WEEK_TO_MONTH.get(w['wk'])
            if not mo:
                continue
            if mo not in month_cost:
                month_cost[mo] = 0
                month_regs_actual[mo] = 0
            if w['cost'] > 0:
                month_cost[mo] += w['cost']
            if w['regs'] > 0:
                month_regs_actual[mo] += w['regs']

        # Update monthly entries with aggregated cost/CPA
        for entry in monthly.get(m, []):
            mo = entry['period']
            if mo in month_cost and month_cost[mo] > 0:
                entry['actual_cost'] = round(month_cost[mo], 0)
                if month_regs_actual.get(mo, 0) > 0:
                    entry['actual_cpa'] = round(month_cost[mo] / month_regs_actual[mo], 2)
                else:
                    entry['actual_cpa'] = None
            else:
                entry['actual_cost'] = None
                entry['actual_cpa'] = None

        # Aggregate quarterly cost from monthly
        for entry in quarterly.get(m, []):
            q = entry['period']
            q_months = [mo for mo, qq in MONTH_TO_Q.items() if qq == q]
            q_cost = sum(month_cost.get(mo, 0) for mo in q_months)
            q_regs = sum(month_regs_actual.get(mo, 0) for mo in q_months)
            if q_cost > 0:
                entry['actual_cost'] = round(q_cost, 0)
                entry['actual_cpa'] = round(q_cost / q_regs, 2) if q_regs > 0 else None
            else:
                entry['actual_cost'] = None
                entry['actual_cpa'] = None

    # Build year-end projections by summing monthly predictions
    yearly = {}
    for m in markets:
        m_months = monthly.get(m, [])
        m_weeks = weekly.get(m, [])
        if m_months:
            total_actual = sum(r["actual_regs"] or 0 for r in m_months if r["actual_regs"])
            total_pred = sum(r["pred_regs"] or 0 for r in m_months if r["pred_regs"])
            total_cost = sum(r["actual_cost"] or 0 for r in m_months if r["actual_cost"])
            yearly[m] = {
                "actual_regs_ytd": total_actual,
                "pred_regs_annual": total_pred,
                "actual_cost_ytd": round(total_cost, 0) if total_cost else None,
                "actual_cpa_ytd": round(total_cost / total_actual, 2) if total_actual and total_cost else None,
                "months_with_actuals": sum(1 for r in m_months if r["actual_regs"]),
            }

    # Determine last data date from the highest week with actuals
    # W15 of 2026 starts 2026-04-07, ends 2026-04-13
    # Approximate: week_start = 2025-12-29 + (wk-1)*7, week_end = week_start + 6
    from datetime import timedelta
    w1_start = datetime(2025, 12, 29)  # 2026 W1 starts Dec 29 2025
    last_data_end = w1_start + timedelta(days=(max_wk - 1) * 7 + 6)
    last_data_date = last_data_end.strftime("%Y-%m-%d")

    # ── Load daily data from _Daily_Data sheet (2025+2026) ──
    ly_weekly = {}
    daily_agg = {}
    if "_Daily_Data" in wb.sheetnames:
        ws_dd = wb["_Daily_Data"]
        dd_headers = [c.value for c in next(ws_dd.iter_rows(min_row=1, max_row=1))]
        dd_col = {h: i for i, h in enumerate(dd_headers)}
        from collections import defaultdict
        daily_agg = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
        for row in ws_dd.iter_rows(min_row=2, values_only=True):
            m = row[dd_col.get("Market", 0)]
            yr = row[dd_col.get("Year", 2)]
            wk = row[dd_col.get("Week", 4)]
            channel = row[dd_col.get("Channel", 6)]
            if not m or not wk: continue
            try: wk = int(wk)
            except: continue
            yr_val = None
            try: yr_val = int(yr)
            except: continue
            ch = str(channel or "Total").strip()
            for field in ["Cost","Clicks","Impressions","Regs"]:
                idx = dd_col.get(field)
                if idx is not None and row[idx] is not None:
                    try:
                        val = float(row[idx])
                        if yr_val == 2025:
                            if ch == "Total":
                                daily_agg[m][wk][field.lower()] += val
                            elif ch == "Brand":
                                daily_agg[m][wk]["brand_" + field.lower()] += val
                            elif ch == "Non-Brand":
                                daily_agg[m][wk]["nb_" + field.lower()] += val
                        elif yr_val == 2026:
                            # Store TY clicks for enriching weekly data
                            if field == "Clicks":
                                if ch == "Brand":
                                    daily_agg[m].setdefault(-wk, {})["ty_brand_clicks"] = daily_agg[m].get(-wk, {}).get("ty_brand_clicks", 0) + val
                                elif ch == "Non-Brand":
                                    daily_agg[m].setdefault(-wk, {})["ty_nb_clicks"] = daily_agg[m].get(-wk, {}).get("ty_nb_clicks", 0) + val
                    except: pass
        for m in daily_agg:
            ly_weekly[m] = []
            for wk in sorted(daily_agg[m].keys()):
                d = daily_agg[m][wk]
                regs = d.get("regs", 0)
                cost = d.get("cost", 0)
                ly_weekly[m].append({
                    "wk": wk,
                    "regs": int(regs) if regs else 0,
                    "cost": round(cost, 0) if cost else 0,
                    "cpa": round(cost / regs, 2) if regs > 0 else 0,
                    "clicks": int(d.get("clicks", 0)),
                    "brand_regs": int(d.get("brand_regs", 0)),
                    "nb_regs": int(d.get("nb_regs", 0)),
                    "brand_cost": round(d.get("brand_cost", 0), 0),
                    "nb_cost": round(d.get("nb_cost", 0), 0),
                    "brand_clicks": int(d.get("brand_clicks", 0)),
                    "nb_clicks": int(d.get("nb_clicks", 0)),
                    "brand_cpa": round(d.get("brand_cost", 0) / d.get("brand_regs", 1), 2) if d.get("brand_regs", 0) > 0 else 0,
                    "nb_cpa": round(d.get("nb_cost", 0) / d.get("nb_regs", 1), 2) if d.get("nb_regs", 0) > 0 else 0,
                })
        print(f"LY data: {len(ly_weekly)} markets, {sum(len(v) for v in ly_weekly.values())} weekly rows")
    elif "_LY_Data" in wb.sheetnames:
        # Backward compat with old sheet name
        ws_ly = wb["_LY_Data"]
        ly_headers = [c.value for c in next(ws_ly.iter_rows(min_row=1, max_row=1))]
        ly_col = {h: i for i, h in enumerate(ly_headers)}
        from collections import defaultdict
        ly_daily = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
        for row in ws_ly.iter_rows(min_row=2, values_only=True):
            m = row[ly_col.get("Market", 0)]
            wk = row[ly_col.get("Week", 2)]
            if not m or not wk: continue
            try: wk = int(wk)
            except: continue
            for field in ["Cost","Clicks","Impressions","Regs","Brand_Cost","Brand_Clicks","Brand_Imp","Brand_Regs","NB_Cost","NB_Clicks","NB_Imp","NB_Regs"]:
                idx = ly_col.get(field)
                if idx is not None and row[idx] is not None:
                    try: ly_daily[m][wk][field.lower()] += float(row[idx])
                    except: pass
        for m in ly_daily:
            ly_weekly[m] = []
            for wk in sorted(ly_daily[m].keys()):
                d = ly_daily[m][wk]
                regs = d.get("regs", 0)
                cost = d.get("cost", 0)
                ly_weekly[m].append({
                    "wk": wk, "regs": int(regs), "cost": round(cost, 0),
                    "cpa": round(cost / regs, 2) if regs > 0 else 0,
                    "brand_regs": int(d.get("brand_regs", 0)), "nb_regs": int(d.get("nb_regs", 0)),
                    "brand_cost": round(d.get("brand_cost", 0), 0), "nb_cost": round(d.get("nb_cost", 0), 0),
                    "brand_clicks": int(d.get("brand_clicks", 0)), "nb_clicks": int(d.get("nb_clicks", 0)),
                    "brand_cpa": round(d.get("brand_cost", 0) / d.get("brand_regs", 1), 2) if d.get("brand_regs", 0) > 0 else 0,
                    "nb_cpa": round(d.get("nb_cost", 0) / d.get("nb_regs", 1), 2) if d.get("nb_regs", 0) > 0 else 0,
                })
        print(f"LY data (legacy): {len(ly_weekly)} markets, {sum(len(v) for v in ly_weekly.values())} weekly rows")
    else:
        print("No _Daily_Data or _LY_Data sheet — run extract-ly-data.py first")

    # Enrich weekly entries with TY click data from _Daily_Data
    if daily_agg:
        for m in weekly:
            for entry in weekly[m]:
                wk_num = entry["wk"]
                ty_data = daily_agg.get(m, {}).get(-wk_num, {})
                if ty_data.get("ty_brand_clicks"):
                    entry["brand_clicks"] = int(ty_data["ty_brand_clicks"])
                if ty_data.get("ty_nb_clicks"):
                    entry["nb_clicks"] = int(ty_data["ty_nb_clicks"])

    # ── Derived-market aggregation (WW = all 10, EU5 = UK+DE+FR+IT+ES) ──
    # The xlsx _Data sheet only stores raw markets. WW and EU5 are sums.
    # Aggregate weekly, ly_weekly, monthly, quarterly for both.
    DERIVED = {
        "WW":  ["US", "CA", "UK", "DE", "FR", "IT", "ES", "JP", "AU", "MX"],
        "EU5": ["UK", "DE", "FR", "IT", "ES"],
    }

    # ── Prediction history from ps.forecasts (DuckDB) ─────────────────────
    # The xlsx only stores the LATEST prediction per week (it overwrites).
    # The authoritative prediction history — every prediction ever made for
    # every target week, with CI bounds and actuals when scored — lives in
    # ps.forecasts. We pull the full history here so the chart can show:
    #   • first prediction (the model's initial call)
    #   • latest prediction (how it updated as data arrived)
    #   • CI band from the latest prediction
    #   • actual when scored
    # This is Richard's calibration feedback loop. Do not overwrite.
    predictions_history = {}
    try:
        import duckdb as _ddb
        import sys as _sys
        _sys.path.insert(0, os.path.expanduser('~/shared/tools'))
        from prediction.config import MOTHERDUCK_TOKEN as _TOKEN
        _con = _ddb.connect(f'md:ps_analytics?motherduck_token={_TOKEN}', read_only=True)
        # Pull ALL predictions for metric_name in {registrations, cost} for 2026.
        # Returns first and latest per (market, target_period, metric_name).
        hist_rows = _con.execute("""
            WITH ranked AS (
              SELECT market, target_period, metric_name,
                     predicted_value, confidence_low, confidence_high,
                     actual_value, error_pct, score, forecast_date, created_at,
                     ROW_NUMBER() OVER (PARTITION BY market, target_period, metric_name
                                        ORDER BY created_at ASC)  AS rn_asc,
                     ROW_NUMBER() OVER (PARTITION BY market, target_period, metric_name
                                        ORDER BY created_at DESC) AS rn_desc
              FROM ps.forecasts
              WHERE metric_name IN ('registrations','cost')
                AND target_period LIKE '2026-W%'
                AND period_type = 'weekly'
            )
            SELECT market, target_period, metric_name,
                   MAX(CASE WHEN rn_asc  = 1 THEN predicted_value END)  AS first_pred,
                   MAX(CASE WHEN rn_asc  = 1 THEN confidence_low END)   AS first_ci_lo,
                   MAX(CASE WHEN rn_asc  = 1 THEN confidence_high END)  AS first_ci_hi,
                   MAX(CASE WHEN rn_asc  = 1 THEN created_at END)       AS first_date,
                   MAX(CASE WHEN rn_desc = 1 THEN predicted_value END)  AS latest_pred,
                   MAX(CASE WHEN rn_desc = 1 THEN confidence_low END)   AS latest_ci_lo,
                   MAX(CASE WHEN rn_desc = 1 THEN confidence_high END)  AS latest_ci_hi,
                   MAX(CASE WHEN rn_desc = 1 THEN created_at END)       AS latest_date,
                   MAX(CASE WHEN rn_desc = 1 THEN actual_value END)     AS actual_value,
                   MAX(CASE WHEN rn_desc = 1 THEN error_pct END)        AS error_pct,
                   MAX(CASE WHEN rn_desc = 1 THEN score END)            AS score,
                   COUNT(*)                                              AS n_preds
            FROM ranked
            GROUP BY market, target_period, metric_name
            ORDER BY market, target_period, metric_name
        """).fetchall()
        # Structure: predictions_history[market][wk] = {"regs": {...}, "cost": {...}}
        METRIC_KEY = {"registrations": "regs", "cost": "cost"}
        for r in hist_rows:
            market_ = r[0]
            target  = r[1]  # '2026-W16'
            metric_ = r[2]
            try:
                wk_ = int(target.split("W")[-1])
            except ValueError:
                continue
            key = METRIC_KEY.get(metric_)
            if not key:
                continue
            predictions_history.setdefault(market_, {}).setdefault(wk_, {})[key] = {
                "first_pred":    r[3],
                "first_ci_lo":   r[4],
                "first_ci_hi":   r[5],
                "first_date":    str(r[6]) if r[6] else None,
                "latest_pred":   r[7],
                "latest_ci_lo":  r[8],
                "latest_ci_hi":  r[9],
                "latest_date":   str(r[10]) if r[10] else None,
                "actual":        r[11],
                "error_pct":     r[12],
                "score":         r[13],
                "n_preds":       r[14],
            }
        _con.close()
        print(f"Prediction history: {len(predictions_history)} markets, "
              f"{sum(len(v) for v in predictions_history.values())} (market,week) records")
    except Exception as _e:
        print(f"WARNING: could not load prediction history from ps.forecasts: {_e}")
        predictions_history = {}

    # Aggregate prediction history for derived markets (WW/EU5).
    # Approach: for each target week, sum first/latest predictions and CI bounds
    # across constituent markets. This is conservative — each constituent's
    # "first" may have been at a different lead time; the aggregate "first"
    # represents the sum of each market's earliest call.
    for derived_name, constituents in DERIVED.items():
        agg = {}
        for c in constituents:
            mkt_hist = predictions_history.get(c, {})
            for wk_, metrics_dict in mkt_hist.items():
                agg_wk = agg.setdefault(wk_, {})
                for metric_key, pred in metrics_dict.items():
                    bucket = agg_wk.setdefault(metric_key, {
                        "first_pred": 0, "first_ci_lo": 0, "first_ci_hi": 0,
                        "latest_pred": 0, "latest_ci_lo": 0, "latest_ci_hi": 0,
                        "actual": 0, "n_preds": 0,
                        "first_date": None, "latest_date": None,
                        "_actual_complete": True, "_first_complete": True, "_latest_complete": True,
                    })
                    for fld in ("first_pred","first_ci_lo","first_ci_hi",
                                "latest_pred","latest_ci_lo","latest_ci_hi","actual"):
                        v = pred.get(fld)
                        if v is None:
                            # Track whether we had a full set of values for the aggregate
                            if fld == "actual":
                                bucket["_actual_complete"] = False
                            elif fld.startswith("first"):
                                bucket["_first_complete"] = False
                            elif fld.startswith("latest"):
                                bucket["_latest_complete"] = False
                        else:
                            bucket[fld] = (bucket[fld] or 0) + float(v)
                    bucket["n_preds"] = max(bucket["n_preds"], pred.get("n_preds") or 0)
                    # Keep the earliest first_date and latest latest_date across constituents
                    if pred.get("first_date") and (not bucket["first_date"] or pred["first_date"] < bucket["first_date"]):
                        bucket["first_date"] = pred["first_date"]
                    if pred.get("latest_date") and (not bucket["latest_date"] or pred["latest_date"] > bucket["latest_date"]):
                        bucket["latest_date"] = pred["latest_date"]
        # Finalize — clear incomplete sums to None so the chart doesn't show bogus partials
        for wk_, metrics_dict in agg.items():
            for metric_key, b in metrics_dict.items():
                if not b.pop("_actual_complete"):
                    b["actual"] = None
                if not b.pop("_first_complete"):
                    for f in ("first_pred","first_ci_lo","first_ci_hi"):
                        b[f] = None
                if not b.pop("_latest_complete"):
                    for f in ("latest_pred","latest_ci_lo","latest_ci_hi"):
                        b[f] = None
                # Error% for derived = latest_pred vs actual (when both present)
                if b.get("actual") and b.get("latest_pred"):
                    try:
                        b["error_pct"] = round(abs(b["latest_pred"] - b["actual"]) / b["actual"] * 100, 1)
                    except (TypeError, ZeroDivisionError):
                        b["error_pct"] = None
                else:
                    b["error_pct"] = None
                b["score"] = None  # derived score is ambiguous; leave blank
        if agg:
            predictions_history[derived_name] = agg
            print(f"Aggregated prediction history for {derived_name}: {len(agg)} weeks")

    # ── Make ps.forecasts authoritative for prediction fields ─────────────
    # Overwrite xlsx-derived pred_regs / ci_lo / ci_hi / pred_cost in weekly[]
    # with predictions_history.latest_pred where it exists. The xlsx values
    # remain as a COLD-START FALLBACK only — if ps.forecasts has no prediction
    # for a given (market, week), the xlsx value stays put. This keeps the
    # dashboard working when the DuckDB connection fails, but means that as
    # soon as the pipeline writes a prediction to ps.forecasts, every
    # downstream consumer (chart, weekly table, projections, refresh-goals)
    # sees the authoritative value automatically — no per-consumer change
    # needed. This also guarantees the derived-market aggregation below
    # (SUM_FIELDS_WEEKLY includes pred_regs, ci_lo, ci_hi, pred_cost) sums
    # authoritative values for WW/EU5 instead of stale xlsx ones.
    overwrites_applied = 0
    for market_name, weeks_history in predictions_history.items():
        weeks_rows = weekly.get(market_name, [])
        if not weeks_rows:
            continue
        # Index the weekly rows by wk for O(1) lookups
        rows_by_wk = {r.get("wk"): r for r in weeks_rows if r.get("wk") is not None}
        for wk_key_str, per_metric in weeks_history.items():
            try:
                wk_num = int(wk_key_str)
            except (TypeError, ValueError):
                continue
            row = rows_by_wk.get(wk_num)
            if row is None:
                continue
            regs_hist = per_metric.get("regs") or {}
            cost_hist = per_metric.get("cost") or {}
            if regs_hist.get("latest_pred") is not None:
                row["pred_regs"] = int(round(regs_hist["latest_pred"]))
                overwrites_applied += 1
            if regs_hist.get("latest_ci_lo") is not None:
                row["ci_lo"] = int(round(regs_hist["latest_ci_lo"]))
            if regs_hist.get("latest_ci_hi") is not None:
                row["ci_hi"] = int(round(regs_hist["latest_ci_hi"]))
            if cost_hist.get("latest_pred") is not None:
                row["pred_cost"] = round(cost_hist["latest_pred"], 0)
    print(f"Overwrote xlsx predictions with ps.forecasts values: {overwrites_applied} weekly cells")

    # ── Same authoritative-overwrite for monthly/quarterly/yearly ────────
    # Pull latest monthly + quarterly + year_end predictions from ps.forecasts
    # and inject into the monthly / quarterly / yearly dicts. ps.forecasts
    # target_period uses '2026-M04' / '2026-Q2' / '2026-YE'; the xlsx-derived
    # dicts use 'Apr' / 'Q2' / single yearly entry. Map between them.
    MONTH_NUM_TO_NAME = {
        "01": "Jan", "02": "Feb", "03": "Mar", "04": "Apr", "05": "May", "06": "Jun",
        "07": "Jul", "08": "Aug", "09": "Sep", "10": "Oct", "11": "Nov", "12": "Dec",
    }
    mq_overwrites = 0
    try:
        import duckdb as _ddb
        _con2 = _ddb.connect(f'md:ps_analytics?motherduck_token={_TOKEN}', read_only=True)
        mq_rows = _con2.execute("""
            WITH ranked AS (
              SELECT market, target_period, period_type, metric_name,
                     predicted_value, confidence_low, confidence_high,
                     created_at,
                     ROW_NUMBER() OVER (PARTITION BY market, target_period, period_type, metric_name
                                        ORDER BY created_at DESC) AS rn_desc
              FROM ps.forecasts
              WHERE period_type IN ('monthly','quarterly','year_end')
                AND metric_name IN ('registrations','cost')
                AND target_period LIKE '2026-%'
            )
            SELECT market, target_period, period_type, metric_name,
                   predicted_value, confidence_low, confidence_high
            FROM ranked WHERE rn_desc = 1
        """).fetchall()
        _con2.close()

        # Index for quick lookups: by_key[(market, period_type, period_key)][metric] = (pred, ci_lo, ci_hi)
        mq_by_key = {}
        for r in mq_rows:
            market_name, target, period_type_, metric_, pred_, ci_l_, ci_h_ = r
            # Map ps.forecasts target_period → xlsx-style key
            period_key = None
            if period_type_ == "monthly" and "-M" in target:
                month_num = target.split("-M")[-1]
                period_key = MONTH_NUM_TO_NAME.get(month_num)
            elif period_type_ == "quarterly" and "-Q" in target:
                period_key = "Q" + target.split("-Q")[-1]
            elif period_type_ == "year_end":
                period_key = "__YE__"  # sentinel for yearly
            if period_key is None:
                continue
            mq_by_key.setdefault((market_name, period_type_, period_key), {})[metric_] = (pred_, ci_l_, ci_h_)

        # Apply to monthly{}
        for market_name, entries in monthly.items():
            for entry in entries:
                p = entry.get("period")
                data = mq_by_key.get((market_name, "monthly", p))
                if not data:
                    continue
                regs_data = data.get("registrations")
                cost_data = data.get("cost")
                if regs_data and regs_data[0] is not None:
                    entry["pred_regs"] = int(round(regs_data[0]))
                    mq_overwrites += 1
                if regs_data and regs_data[1] is not None:
                    entry["ci_lo"] = int(round(regs_data[1]))
                if cost_data and cost_data[0] is not None:
                    entry["pred_cost"] = round(cost_data[0], 0)

        # Apply to quarterly{}
        for market_name, entries in quarterly.items():
            for entry in entries:
                p = entry.get("period")
                data = mq_by_key.get((market_name, "quarterly", p))
                if not data:
                    continue
                regs_data = data.get("registrations")
                cost_data = data.get("cost")
                if regs_data and regs_data[0] is not None:
                    entry["pred_regs"] = int(round(regs_data[0]))
                    mq_overwrites += 1
                if regs_data and regs_data[1] is not None:
                    entry["ci_lo"] = int(round(regs_data[1]))
                if cost_data and cost_data[0] is not None:
                    entry["pred_cost"] = round(cost_data[0], 0)

        # Apply to yearly{} — overwrite pred_regs_annual with year_end registrations
        for market_name in list(yearly.keys()):
            data = mq_by_key.get((market_name, "year_end", "__YE__"))
            if not data:
                continue
            regs_data = data.get("registrations")
            cost_data = data.get("cost")
            if regs_data and regs_data[0] is not None:
                yearly[market_name]["pred_regs_annual"] = int(round(regs_data[0]))
                mq_overwrites += 1
            if regs_data and regs_data[1] is not None:
                yearly[market_name]["pred_regs_annual_ci_lo"] = int(round(regs_data[1]))
            if regs_data and regs_data[2] is not None:
                yearly[market_name]["pred_regs_annual_ci_hi"] = int(round(regs_data[2]))
            if cost_data and cost_data[0] is not None:
                yearly[market_name]["pred_cost_annual"] = round(cost_data[0], 0)

        print(f"Overwrote monthly/quarterly/yearly predictions: {mq_overwrites} cells")
    except Exception as _e:
        print(f"WARNING: could not overwrite monthly/quarterly from ps.forecasts: {_e}")

    SUM_FIELDS_WEEKLY = [
        "regs", "cost", "brand_regs", "nb_regs", "brand_clicks", "nb_clicks",
        "pred_regs", "ci_lo", "ci_hi", "op2_regs", "pred_cost", "op2_cost",
    ]
    SUM_FIELDS_LY = [
        "regs", "cost", "clicks", "brand_regs", "nb_regs",
        "brand_cost", "nb_cost", "brand_clicks", "nb_clicks",
    ]

    def aggregate_weekly(constituents, source_dict, sum_fields, derive_cpa=True):
        """Sum per-week rows across constituent markets; derive CPA fields last."""
        per_wk = {}
        for c in constituents:
            for row in source_dict.get(c, []):
                wk = row.get("wk")
                if wk is None:
                    continue
                bucket = per_wk.setdefault(wk, {"wk": wk, "week": row.get("week", f"W{wk}")})
                for f in sum_fields:
                    v = row.get(f)
                    if v is None:
                        continue
                    try:
                        bucket[f] = bucket.get(f, 0) + float(v)
                    except (TypeError, ValueError):
                        pass
        out = []
        for wk in sorted(per_wk):
            r = per_wk[wk]
            # Cast integer-ish fields back to int, round cost
            for f in ("regs", "brand_regs", "nb_regs", "brand_clicks", "nb_clicks",
                      "pred_regs", "ci_lo", "ci_hi", "op2_regs", "clicks"):
                if f in r and r[f] is not None:
                    try:
                        r[f] = int(round(r[f]))
                    except (TypeError, ValueError):
                        pass
            for f in ("cost", "pred_cost", "op2_cost", "brand_cost", "nb_cost"):
                if f in r and r[f] is not None:
                    r[f] = round(r[f], 0)
            if derive_cpa:
                if r.get("regs", 0) > 0 and r.get("cost", 0) > 0:
                    r["cpa"] = round(r["cost"] / r["regs"], 2)
                else:
                    r["cpa"] = 0
                if r.get("brand_regs", 0) > 0 and r.get("brand_cost", 0) > 0:
                    r["brand_cpa"] = round(r["brand_cost"] / r["brand_regs"], 2)
                if r.get("nb_regs", 0) > 0 and r.get("nb_cost", 0) > 0:
                    r["nb_cpa"] = round(r["nb_cost"] / r["nb_regs"], 2)
                if r.get("pred_regs") and r.get("pred_cost"):
                    r["pred_cpa"] = round(r["pred_cost"] / r["pred_regs"], 2)
                if r.get("op2_regs") and r.get("op2_cost"):
                    r["op2_cpa"] = round(r["op2_cost"] / r["op2_regs"], 2)
            out.append(r)
        return out

    def aggregate_periodic(constituents, source_dict):
        """Sum monthly or quarterly rows across constituents; derive CPA."""
        per_period = {}
        for c in constituents:
            for row in source_dict.get(c, []):
                p = row.get("period")
                if not p:
                    continue
                b = per_period.setdefault(p, {"period": p})
                for f in ("actual_regs", "pred_regs", "ci_lo", "actual_cost"):
                    v = row.get(f)
                    if v is None:
                        continue
                    try:
                        b[f] = b.get(f, 0) + float(v)
                    except (TypeError, ValueError):
                        pass
        out = []
        for p in per_period:
            r = per_period[p]
            for f in ("actual_regs", "pred_regs", "ci_lo"):
                if f in r and r[f] is not None:
                    r[f] = int(round(r[f]))
            if "actual_cost" in r and r["actual_cost"]:
                r["actual_cost"] = round(r["actual_cost"], 0)
                if r.get("actual_regs"):
                    r["actual_cpa"] = round(r["actual_cost"] / r["actual_regs"], 2)
            else:
                r["actual_cost"] = None
                r["actual_cpa"] = None
            out.append(r)
        return out

    for derived, constituents in DERIVED.items():
        have = [c for c in constituents if c in weekly and weekly[c]]
        if not have:
            continue
        weekly[derived] = aggregate_weekly(have, weekly, SUM_FIELDS_WEEKLY)
        ly_have = [c for c in constituents if c in ly_weekly and ly_weekly[c]]
        if ly_have:
            ly_weekly[derived] = aggregate_weekly(ly_have, ly_weekly, SUM_FIELDS_LY)
        m_have = [c for c in constituents if c in monthly and monthly[c]]
        if m_have:
            monthly[derived] = aggregate_periodic(m_have, monthly)
        q_have = [c for c in constituents if c in quarterly and quarterly[c]]
        if q_have:
            quarterly[derived] = aggregate_periodic(q_have, quarterly)
        if derived not in markets:
            markets.append(derived)
        print(f"Aggregated {derived}: {len(weekly[derived])} weekly rows from {have}")

    markets = sorted(set(markets))

    output = {
        "generated": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "last_data_date": last_data_date,
        "source": "ps-forecast-tracker.xlsx (SharePoint) + ps.forecasts (MotherDuck)",
        "max_week": max_wk,
        "weekly": weekly,
        "ly_weekly": ly_weekly,
        "monthly": monthly,
        "quarterly": quarterly,
        "yearly": yearly,
        "markets": markets,
        "predictions_history": predictions_history,
    }

    os.makedirs(os.path.dirname(JSON_PATH), exist_ok=True)
    with open(JSON_PATH, "w") as f:
        json.dump(output, f, indent=2)

    total_rows = sum(len(v) for v in weekly.values())
    total_monthly = sum(len(v) for v in monthly.values())
    total_quarterly = sum(len(v) for v in quarterly.values())
    print(f"Done: {len(markets)} markets, {total_rows} weekly + {total_monthly} monthly + {total_quarterly} quarterly rows, through W{max_wk}")
    print(f"Written: {JSON_PATH}")

if __name__ == "__main__":
    main()
