#!/usr/bin/env python3
"""
PS Forecast Tracker v3 — Excel Dashboard + Prediction Database
Layout: Year-End → Quarterly → Monthly → Weekly (top-down)
Charts: right side of tables (col P+)
Predictions: derived from OP2 monthly targets where available
"""

import duckdb
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule, IconSetRule
from openpyxl.utils import get_column_letter
import os, math

TOKEN = os.environ.get('MOTHERDUCK_TOKEN',
    'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJlbWFpbCI6InJpY2hzY290dHdpbGxAZ21haWwuY29tIiwibWRSZWdpb24iOiJhd3MtdXMtZWFzdC0xIiwic2Vzc2lvbiI6InJpY2hzY290dHdpbGwuZ21haWwuY29tIiwicGF0IjoiVDNIYzFVQWYzT3o1bjVkLS03ckdHNlBjMlpUdVNNbFItT3RXMS1qNzVPUSIsInVzZXJJZCI6ImU2MDhlNDZiLTE4YzctNGE5Ny04M2I2LWE0N2ZhOThmNjBhYyIsImlzcyI6Im1kX3BhdCIsInJlYWRPbmx5IjpmYWxzZSwidG9rZW5UeXBlIjoicmVhZF93cml0ZSIsImlhdCI6MTc3NTQ0MzY0N30.tS0Cab3FQ8_CDZ1PqOo9z09KYHEUFHwuLVXRQrxcHig')
# Use MotherDuck cloud DB (always current) with local fallback
_LOCAL_DB = os.path.expanduser('~/shared/data/duckdb/ps-analytics.duckdb')
try:
    DB = f'md:ps_analytics?motherduck_token={TOKEN}'
    _test = duckdb.connect(DB, read_only=True)
    _test.execute("SELECT 1")
    _test.close()
    print("Connected to MotherDuck (cloud)")
except Exception:
    DB = _LOCAL_DB
    print(f"MotherDuck unavailable, falling back to local: {DB}")
OUT = os.path.expanduser('~/shared/dashboards/ps-forecast-tracker.xlsx')

# Palette
NAVY="1F4E79"; BLUE="4A9EFF"; GREEN="27AE60"; RED="E74C3C"
ORANGE="F39C12"; GRAY="95A5A6"; WHITE="FFFFFF"
HEADER_BG="2C3E6B"; ALT_ROW="EDF2F9"; FUTURE_BG="FFF9E6"; KPI_BG="E8F0FE"
MKT_C = {'US':'3498DB','UK':'2ECC71','DE':'E67E22','FR':'E74C3C',
         'IT':'1ABC9C','ES':'C0392B','CA':'F1C40F','JP':'9B59B6','MX':'D35400','AU':'16A085'}

hdr_font=Font(bold=True,color=WHITE,size=10,name='Calibri')
hdr_align=Alignment(horizontal='left',vertical='center',wrap_text=True)
title_font=Font(bold=True,size=16,color=NAVY,name='Calibri')
sub_font=Font(size=11,color=GRAY,name='Calibri')
body=Font(size=10,name='Calibri')
bold=Font(bold=True,size=10,name='Calibri')
kpi_lbl=Font(size=9,color=GRAY,name='Calibri')
kpi_val=Font(bold=True,size=20,color=NAVY,name='Calibri')
kpi_sub=Font(size=9,color=GRAY,name='Calibri')
gf=Font(color=GREEN,bold=True,name='Calibri')
bf=Font(color=RED,bold=True,name='Calibri')
wf=Font(color=ORANGE,bold=True,name='Calibri')
future_font=Font(color=ORANGE,size=10,italic=True,name='Calibri')
thin_bdr=Border(bottom=Side(style='thin',color='DEE2E6'))
alt_fill=PatternFill(start_color=ALT_ROW,end_color=ALT_ROW,fill_type="solid")
future_fill=PatternFill(start_color=FUTURE_BG,end_color=FUTURE_BG,fill_type="solid")
kpi_fill=PatternFill(start_color=KPI_BG,end_color=KPI_BG,fill_type="solid")
kpi_bdr=Border(top=Side(style='thin',color='B8CCE4'),bottom=Side(style='thin',color='B8CCE4'),
               left=Side(style='thin',color='B8CCE4'),right=Side(style='thin',color='B8CCE4'))
# Richard's preferred alignment: left-aligned, vertically centered everywhere
LEFT_MID=Alignment(horizontal='left',vertical='center')

MONTHS=["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
# Weeks per month (approximate)
MONTH_WEEKS = {0:(1,4),1:(5,8),2:(9,13),3:(14,17),4:(18,22),5:(23,26),
               6:(27,31),7:(32,35),8:(36,39),9:(40,44),10:(45,48),11:(49,52)}

def hdr_row(ws,row,headers,fill=None):
    for j,h in enumerate(headers):
        c=ws.cell(row=row,column=j+1,value=h)
        c.font=hdr_font;c.fill=PatternFill(start_color=fill or HEADER_BG,end_color=fill or HEADER_BG,fill_type="solid")
        c.border=Border(bottom=Side(style='medium',color=NAVY))

def style_row(ws,row,ncols,idx,is_future=False):
    for col in range(1,ncols+1):
        c=ws.cell(row=row,column=col)
        c.font=future_font if is_future else body
        c.border=thin_bdr
        if is_future: c.fill=future_fill
        elif idx%2==1: c.fill=alt_fill

def col_w(ws,widths):
    for i,w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width=w

def kpi(ws,row,col,label,value,sub="",span=2):
    for r in range(row,row+3):
        for sc in range(col,col+span):
            ws.cell(row=r,column=sc).fill=kpi_fill;ws.cell(row=r,column=sc).border=kpi_bdr
    ws.merge_cells(start_row=row,start_column=col,end_row=row,end_column=col+span-1)
    ws.cell(row=row,column=col,value=label).font=kpi_lbl
    ws.merge_cells(start_row=row+1,start_column=col,end_row=row+1,end_column=col+span-1)
    ws.cell(row=row+1,column=col,value=value).font=kpi_val
    if sub:
        ws.merge_cells(start_row=row+2,start_column=col,end_row=row+2,end_column=col+span-1)
        ws.cell(row=row+2,column=col,value=sub).font=kpi_sub

# ── Query ──
print("Querying DuckDB...")
con=duckdb.connect(DB,read_only=True)

# Query ps.performance (pivoted weekly data) — the current MotherDuck schema
weekly_raw=con.execute("""
    SELECT market, period_key as week,
        CAST(REPLACE(period_key, '2026-W', '') AS INTEGER) as wk,
        registrations as regs, clicks, impressions, cost, cpa, cpc, cvr, ctr,
        brand_registrations as brand_regs, brand_cost, brand_cpa, brand_clicks,
        nb_registrations as nb_regs, nb_cost, nb_cpa, nb_clicks
    FROM ps.performance
    WHERE period_type='weekly' AND period_key LIKE '2026%'
    ORDER BY market, CAST(REPLACE(period_key, '2026-W', '') AS INTEGER)
""").fetchall()

# Monthly OP2 targets from ps.targets (long format — pivot regs + cost)
monthly_targets=con.execute("""
    SELECT r.market,
        CASE CAST(REPLACE(r.period_key, '2026-M', '') AS INTEGER)
            WHEN 1 THEN '2026 Jan' WHEN 2 THEN '2026 Feb' WHEN 3 THEN '2026 Mar'
            WHEN 4 THEN '2026 Apr' WHEN 5 THEN '2026 May' WHEN 6 THEN '2026 Jun'
            WHEN 7 THEN '2026 Jul' WHEN 8 THEN '2026 Aug' WHEN 9 THEN '2026 Sep'
            WHEN 10 THEN '2026 Oct' WHEN 11 THEN '2026 Nov' WHEN 12 THEN '2026 Dec'
        END as month,
        r.target_value as regs_op2,
        c.target_value as spend_op2
    FROM ps.targets r
    LEFT JOIN ps.targets c ON r.market=c.market AND r.period_key=c.period_key
        AND c.metric_name='cost' AND c.period_type='monthly'
    WHERE r.metric_name='registrations' AND r.period_type='monthly'
        AND r.period_key LIKE '2026%'
    ORDER BY r.market, r.period_key
""").fetchall()

# Projections — use forecast_tracker year-end as summary (empty projections table)
projections=[]
try:
    for m in ['US','UK','DE','FR','IT','ES','CA','JP','MX','AU']:
        ye = con.execute(f"SELECT predicted, actual FROM ps.forecast_tracker WHERE market='{m}' AND metric_name='registrations' AND horizon='year_end'").fetchone()
        if ye and ye[0]:
            # Find monthly OP2 total for this market
            op2_row = con.execute(f"SELECT SUM(target_value) FROM ps.targets WHERE market='{m}' AND metric_name='registrations' AND period_type='monthly'").fetchone()
            op2_total = op2_row[0] if op2_row and op2_row[0] else 0
            vs_op2 = ((ye[0] - op2_total) / op2_total * 100) if op2_total else None
            projections.append((m, 'YE', ye[0], None, None, op2_total, None, vs_op2, None, None, 'bayesian'))
except Exception:
    pass

# NOTE: con stays open — needed for per-market forecast_tracker queries below


# Group
markets_order=['US','UK','DE','FR','IT','ES','CA','JP','MX','AU']
weekly_by_mkt={}
for r in weekly_raw:
    weekly_by_mkt.setdefault(r[0],{})[r[2]]=r

proj_by_mkt={r[0]:r for r in projections}

monthly_tgt_by_mkt={}
for r in monthly_targets:
    monthly_tgt_by_mkt.setdefault(r[0],{})[r[1]]=r

markets=[m for m in markets_order if m in weekly_by_mkt]
print(f"  {len(weekly_raw)} rows, {len(markets)} markets")

wb=Workbook()


# ════════════════════════════════════════════
# DASHBOARD SHEET
# ════════════════════════════════════════════
print("Building Dashboard...")
ws=wb.active;ws.title="Dashboard";ws.sheet_properties.tabColor=NAVY;ws.sheet_view.showGridLines=False

ws.merge_cells('A1:L1')
ws['A1']="📊 PS Forecast Tracker — 2026";ws['A1'].font=Font(bold=True,size=18,color=NAVY,name='Calibri')
ws.merge_cells('A2:L2')
ws['A2']="Year-End → Quarterly → Monthly → Weekly • Predictions + Actuals + Error Tracking";ws['A2'].font=sub_font

total_regs=sum(r[3] for r in weekly_raw)
total_cost=sum(r[6] for r in weekly_raw)
avg_cpa=total_cost/total_regs if total_regs else 0
max_wk=max(r[2] for r in weekly_raw)

kpi(ws,4,1,"WW YTD REGS",f"{total_regs:,}",f"W1–W{max_wk}",span=3)
kpi(ws,4,4,"WW YTD SPEND",f"${total_cost:,.0f}","",span=3)
kpi(ws,4,7,"WW AVG CPA",f"${avg_cpa:.2f}","",span=2)
kpi(ws,4,9,"MARKETS",str(len(markets)),f"{len(weekly_raw)} data points",span=2)

r0=9
headers=['Market','YTD Regs','YTD Spend','Avg CPA','Avg Wk Regs','Latest Wk','WoW Δ%','Proj Mo Regs','vs OP2 %','Source']
hdr_row(ws,r0,headers)
for j,h in enumerate(headers): ws.cell(row=r0,column=j+1,value=h)

for i,m in enumerate(markets):
    row=r0+1+i
    wk_data=weekly_by_mkt[m];wk_nums=sorted(wk_data.keys())
    ytd_r=sum(wk_data[w][3] for w in wk_nums);ytd_c=sum(wk_data[w][6] for w in wk_nums)
    avg_c=ytd_c/ytd_r if ytd_r else 0;avg_wk=ytd_r/len(wk_nums)
    last=wk_data[wk_nums[-1]];prev=wk_data.get(wk_nums[-2]) if len(wk_nums)>1 else None
    wow=((last[3]-prev[3])/prev[3]*100) if prev and prev[3] else None
    proj=proj_by_mkt.get(m)

    ws.cell(row=row,column=1,value=m).font=bold
    ws.cell(row=row,column=2,value=ytd_r).number_format='#,##0'
    ws.cell(row=row,column=3,value=round(ytd_c)).number_format='$#,##0'
    ws.cell(row=row,column=4,value=round(avg_c,2)).number_format='$#,##0.00'
    ws.cell(row=row,column=5,value=round(avg_wk)).number_format='#,##0'
    ws.cell(row=row,column=6,value=last[3]).number_format='#,##0'
    if wow is not None:
        c=ws.cell(row=row,column=7,value=round(wow,1));c.number_format='+0.0"%";-0.0"%"'
        c.font=gf if wow>0 else bf if wow<-5 else wf
    if proj:
        ws.cell(row=row,column=8,value=proj[2]).number_format='#,##0'
        if proj[7] is not None:
            c=ws.cell(row=row,column=9,value=proj[7]);c.number_format='+0.0"%";-0.0"%"'
            c.font=gf if proj[7]>=0 else bf
        ws.cell(row=row,column=10,value=proj[10] or '')
    style_row(ws,row,len(headers),i)

de=r0+len(markets)
ws.conditional_formatting.add(f'B{r0+1}:B{de}',DataBarRule(start_type='min',end_type='max',color=BLUE,showValue=True))
ws.conditional_formatting.add(f'D{r0+1}:D{de}',ColorScaleRule(start_type='min',start_color=GREEN,mid_type='percentile',mid_value=50,mid_color='FFFFFF',end_type='max',end_color=RED))
ws.conditional_formatting.add(f'G{r0+1}:G{de}',IconSetRule('3Arrows','num',[-5,0,5],showValue=True))

chart=BarChart();chart.type="col";chart.title="YTD Registrations by Market";chart.style=2
chart.width=24;chart.height=14;chart.legend=None
chart.y_axis.title="Registrations";chart.y_axis.numFmt='#,##0';chart.x_axis.title="Market"
d=Reference(ws,min_col=2,min_row=r0,max_row=de);c=Reference(ws,min_col=1,min_row=r0+1,max_row=de)
chart.add_data(d,titles_from_data=True);chart.set_categories(c)
chart.series[0].graphicalProperties.solidFill=BLUE
ws.add_chart(chart,f"L{r0}")  # chart to the right


# ════════════════════════════════════════════
# PER-MARKET SHEETS
# Order: Year-End → Quarterly → Monthly → Weekly
# Charts: right side (col P+)
# ════════════════════════════════════════════
print("Building per-market sheets...")

PROJ_HEADERS = ['Period','Actual Regs','Pred Regs','CI Low','CI High','OP2 Target',
                'vs OP2 %','Error %','Confidence','Score','Notes']
NCOLS_PROJ = len(PROJ_HEADERS)

WEEKLY_HEADERS = ['Week','Wk#','Actual Regs','Actual Cost','Actual CPA',
                  'Pred Regs','CI Low','CI High','Confidence',
                  'Error %','Within CI','WoW Δ%','Brand Regs','NB Regs']
NCOLS_WK = len(WEEKLY_HEADERS)

for m in markets:
    ws_m=wb.create_sheet(title=m)
    ws_m.sheet_properties.tabColor=MKT_C.get(m,BLUE)
    ws_m.sheet_view.showGridLines=False
    mc=MKT_C.get(m,HEADER_BG)

    wk_data=weekly_by_mkt[m];wk_nums=sorted(wk_data.keys())
    last_actual_wk=max(wk_nums) if wk_nums else 0
    proj=proj_by_mkt.get(m)
    mo_tgts=monthly_tgt_by_mkt.get(m,{})
    ytd_r=sum(wk_data[w][3] for w in wk_nums)
    ytd_c=sum(wk_data[w][6] for w in wk_nums)
    avg_cpa_m=ytd_c/ytd_r if ytd_r else 0

    # Pull weekly predictions from forecast_tracker (weighted Bayesian predictions)
    weekly_preds = {}
    weekly_ci = {}
    try:
        fc_rows = con.execute(f"""SELECT period_label, predicted, ci_low, ci_high 
            FROM ps.forecast_tracker 
            WHERE market='{m}' AND metric_name='registrations' AND horizon='weekly' AND predicted IS NOT NULL
            ORDER BY period_order""").fetchall()
        for r in fc_rows:
            wk_num = int(r[0].replace('W', ''))
            weekly_preds[wk_num] = round(r[1])
            weekly_ci[wk_num] = (round(r[2]) if r[2] else None, round(r[3]) if r[3] else None)
    except Exception:
        pass
    # Fallback: derive from OP2 monthly targets for any weeks without predictions
    for mi_idx, mn in enumerate(MONTHS):
        month_key = f"2026 {mn}"
        tgt = mo_tgts.get(month_key)
        if tgt and tgt[2]:
            w_start, w_end = MONTH_WEEKS[mi_idx]
            n_weeks = w_end - w_start + 1
            wk_pred = round(tgt[2] / n_weeks)
            for ww in range(w_start, w_end + 1):
                if ww not in weekly_preds:
                    weekly_preds[ww] = wk_pred
                    weekly_ci[ww] = (round(wk_pred * 0.8), round(wk_pred * 1.2))

    # Title
    ws_m.merge_cells('A1:N1')
    ws_m['A1']=f"{m} — Forecast Tracker 2026"
    ws_m['A1'].font=Font(bold=True,size=16,color=mc,name='Calibri')

    # KPIs
    kpi(ws_m,3,1,"YTD REGS",f"{ytd_r:,}",f"W1–W{last_actual_wk}",span=2)
    kpi(ws_m,3,3,"YTD SPEND",f"${ytd_c:,.0f}","",span=2)
    kpi(ws_m,3,5,"AVG CPA",f"${avg_cpa_m:.2f}","",span=2)
    # Pull predictions from forecast_tracker for all horizons
    ft_preds = {}  # (horizon, period_label, metric) → (predicted, ci_low, ci_high)
    try:
        ft_rows = con.execute(f"""SELECT horizon, period_label, metric_name, predicted, ci_low, ci_high 
            FROM ps.forecast_tracker WHERE market='{m}' AND predicted IS NOT NULL""").fetchall()
        for r in ft_rows:
            ft_preds[(r[0], r[1], r[2])] = (r[3], r[4], r[5])
    except Exception:
        pass

    # Compute vs OP2 from year-end prediction
    ye_pred_regs = ft_preds.get(('year_end', '2026', 'registrations'), (None,None,None))[0]
    annual_op2_regs = sum(t[2] for t in mo_tgts.values() if t[2])
    if ye_pred_regs and annual_op2_regs:
        vs_op2_pct = (ye_pred_regs - annual_op2_regs) / annual_op2_regs * 100
        kpi(ws_m,3,7,"vs OP2",f"{vs_op2_pct:+.1f}%",'bayesian weighted',span=2)
    elif proj and proj[7] is not None:
        kpi(ws_m,3,7,"vs OP2",f"{proj[7]:+.1f}%",proj[10] or '',span=2)

    cur_row = 8

    # ═══ YEAR-END ═══
    ws_m.merge_cells(f'A{cur_row}:K{cur_row}')
    ws_m.cell(row=cur_row,column=1,value="YEAR-END PROJECTION").font=Font(bold=True,size=12,color=NAVY,name='Calibri')
    cur_row += 1
    hdr_row(ws_m,cur_row,PROJ_HEADERS,mc)
    for j,h in enumerate(PROJ_HEADERS): ws_m.cell(row=cur_row,column=j+1,value=h)
    cur_row += 1
    ye_row = cur_row
    ws_m.cell(row=ye_row,column=1,value="2026 Full Year").font=bold
    # Sum all OP2 monthly targets for annual target
    annual_op2 = sum(t[2] for t in mo_tgts.values() if t[2])
    if annual_op2:
        ws_m.cell(row=ye_row,column=6,value=annual_op2).number_format='#,##0'
    # Year-end prediction from forecast_tracker
    ye_ft = ft_preds.get(('year_end', '2026', 'registrations'))
    if ye_ft and ye_ft[0]:
        ws_m.cell(row=ye_row,column=3,value=ye_ft[0]).number_format='#,##0'
        if ye_ft[1]: ws_m.cell(row=ye_row,column=4,value=ye_ft[1]).number_format='#,##0'
        if ye_ft[2]: ws_m.cell(row=ye_row,column=5,value=ye_ft[2]).number_format='#,##0'
    for col in [2]: ws_m.cell(row=ye_row,column=col).number_format='#,##0'
    ws_m.cell(row=ye_row,column=7).value=f'=IF(AND(B{ye_row}>0,F{ye_row}>0),(B{ye_row}-F{ye_row})/F{ye_row},"")';ws_m.cell(row=ye_row,column=7).number_format='+0.0%;-0.0%'
    ws_m.cell(row=ye_row,column=8).value=f'=IF(AND(B{ye_row}>0,C{ye_row}>0),ABS(C{ye_row}-B{ye_row})/B{ye_row},"")';ws_m.cell(row=ye_row,column=8).number_format='0.0%'
    ws_m.cell(row=ye_row,column=9).number_format='0%'
    style_row(ws_m,ye_row,NCOLS_PROJ,0,is_future=True)
    cur_row += 2

    # ═══ QUARTERLY ═══
    ws_m.merge_cells(f'A{cur_row}:K{cur_row}')
    ws_m.cell(row=cur_row,column=1,value="QUARTERLY PROJECTIONS").font=Font(bold=True,size=12,color=NAVY,name='Calibri')
    cur_row += 1
    hdr_row(ws_m,cur_row,PROJ_HEADERS,mc)
    for j,h in enumerate(PROJ_HEADERS): ws_m.cell(row=cur_row,column=j+1,value=h)
    cur_row += 1
    q_start = cur_row

    for qi,ql in enumerate(['Q1','Q2','Q3','Q4']):
        row=cur_row+qi
        ws_m.cell(row=row,column=1,value=ql).font=bold
        # Q OP2 target = sum of 3 months
        q_months_idx = [qi*3, qi*3+1, qi*3+2]
        q_op2 = 0
        for mi2 in q_months_idx:
            mk = f"2026 {MONTHS[mi2]}"
            t = mo_tgts.get(mk)
            if t and t[2]: q_op2 += t[2]
        if q_op2:
            ws_m.cell(row=row,column=6,value=q_op2).number_format='#,##0'
        # Quarterly prediction from forecast_tracker
        q_ft = ft_preds.get(('quarterly', ql, 'registrations'))
        if q_ft and q_ft[0]:
            ws_m.cell(row=row,column=3,value=q_ft[0]).number_format='#,##0'
            if q_ft[1]: ws_m.cell(row=row,column=4,value=q_ft[1]).number_format='#,##0'
            if q_ft[2]: ws_m.cell(row=row,column=5,value=q_ft[2]).number_format='#,##0'
        # Q actual = sum of weekly actuals in that quarter
        q_wk_start, q_wk_end = MONTH_WEEKS[qi*3][0], MONTH_WEEKS[qi*3+2][1]
        q_actual = sum(wk_data[w][3] for w in range(q_wk_start, q_wk_end+1) if w in wk_data)
        if q_actual > 0 and all(w in wk_data for w in range(q_wk_start, q_wk_end+1)):
            ws_m.cell(row=row,column=2,value=q_actual).number_format='#,##0'

        for col in [3,4,5]: ws_m.cell(row=row,column=col).number_format='#,##0'
        ws_m.cell(row=row,column=7).value=f'=IF(AND(B{row}>0,F{row}>0),(B{row}-F{row})/F{row},"")';ws_m.cell(row=row,column=7).number_format='+0.0%;-0.0%'
        ws_m.cell(row=row,column=8).value=f'=IF(AND(B{row}>0,C{row}>0),ABS(C{row}-B{row})/B{row},"")';ws_m.cell(row=row,column=8).number_format='0.0%'
        ws_m.cell(row=row,column=9).number_format='0%'
        style_row(ws_m,row,NCOLS_PROJ,qi,is_future=(qi>=1))
    cur_row += 5

    # ═══ MONTHLY ═══
    ws_m.merge_cells(f'A{cur_row}:K{cur_row}')
    ws_m.cell(row=cur_row,column=1,value="MONTHLY PROJECTIONS").font=Font(bold=True,size=12,color=NAVY,name='Calibri')
    cur_row += 1
    hdr_row(ws_m,cur_row,PROJ_HEADERS,mc)
    for j,h in enumerate(PROJ_HEADERS): ws_m.cell(row=cur_row,column=j+1,value=h)
    cur_row += 1
    mo_start = cur_row

    for mi,mn in enumerate(MONTHS):
        row=cur_row+mi
        month_key=f"2026 {mn}"
        tgt=mo_tgts.get(month_key)
        ws_m.cell(row=row,column=1,value=mn).font=bold

        # Monthly actual = sum of weekly actuals in that month
        w_start, w_end = MONTH_WEEKS[mi]
        mo_actual = sum(wk_data[w][3] for w in range(w_start, w_end+1) if w in wk_data)
        all_weeks_in = all(w in wk_data for w in range(w_start, w_end+1))
        if mo_actual > 0 and all_weeks_in:
            ws_m.cell(row=row,column=2,value=mo_actual).number_format='#,##0'

        if tgt and tgt[2]:
            ws_m.cell(row=row,column=6,value=tgt[2]).number_format='#,##0'
        # Monthly prediction from forecast_tracker (weighted Bayesian)
        mo_ft = ft_preds.get(('monthly', mn, 'registrations'))
        if mo_ft and mo_ft[0]:
            ws_m.cell(row=row,column=3,value=mo_ft[0]).number_format='#,##0'
            if mo_ft[1]: ws_m.cell(row=row,column=4,value=mo_ft[1]).number_format='#,##0'
            if mo_ft[2]: ws_m.cell(row=row,column=5,value=mo_ft[2]).number_format='#,##0'
            ws_m.cell(row=row,column=9,value=0.7).number_format='0%'
        elif tgt and tgt[2] and not all_weeks_in:
            # Fallback: OP2 as prediction for months without Bayesian forecast
            ws_m.cell(row=row,column=3,value=tgt[2]).number_format='#,##0'
            ws_m.cell(row=row,column=4,value=round(tgt[2]*0.85)).number_format='#,##0'
            ws_m.cell(row=row,column=5,value=round(tgt[2]*1.15)).number_format='#,##0'
            ws_m.cell(row=row,column=9,value=0.4).number_format='0%'
            ws_m.cell(row=row,column=11,value="OP2 baseline").font=Font(color=GRAY,size=9,italic=True,name='Calibri')

        for col in [2,3,4,5]: ws_m.cell(row=row,column=col).number_format='#,##0'
        ws_m.cell(row=row,column=7).value=f'=IF(AND(B{row}>0,F{row}>0),(B{row}-F{row})/F{row},"")';ws_m.cell(row=row,column=7).number_format='+0.0%;-0.0%'
        ws_m.cell(row=row,column=8).value=f'=IF(AND(B{row}>0,C{row}>0),ABS(C{row}-B{row})/B{row},"")';ws_m.cell(row=row,column=8).number_format='0.0%'
        ws_m.cell(row=row,column=9).number_format='0%'
        style_row(ws_m,row,NCOLS_PROJ,mi,is_future=(mi>=3))
    cur_row += 13

    # ═══ WEEKLY (W1-W52) ═══
    ws_m.merge_cells(f'A{cur_row}:N{cur_row}')
    ws_m.cell(row=cur_row,column=1,value="WEEKLY (W1–W52)").font=Font(bold=True,size=12,color=NAVY,name='Calibri')
    cur_row += 1
    wk_hdr_row = cur_row
    hdr_row(ws_m,cur_row,WEEKLY_HEADERS,mc)
    for j,h in enumerate(WEEKLY_HEADERS): ws_m.cell(row=cur_row,column=j+1,value=h)
    cur_row += 1
    wk_data_start = cur_row

    for w in range(1,53):
        row=cur_row+w-1
        is_future=w>last_actual_wk
        actual=wk_data.get(w)
        pred_val=weekly_preds.get(w)

        ws_m.cell(row=row,column=1,value=f"2026 W{w}")
        ws_m.cell(row=row,column=2,value=w)

        if actual:
            ws_m.cell(row=row,column=3,value=actual[3]).number_format='#,##0'
            ws_m.cell(row=row,column=4,value=round(actual[6])).number_format='$#,##0'
            ws_m.cell(row=row,column=5,value=round(actual[7],2)).number_format='$#,##0.00'
            ws_m.cell(row=row,column=13,value=actual[11] or 0).number_format='#,##0'
            ws_m.cell(row=row,column=14,value=actual[15] or 0).number_format='#,##0'

        # Predictions — always show, regardless of whether actuals exist
        if pred_val:
            ws_m.cell(row=row,column=6,value=pred_val).number_format='#,##0'
            ci = weekly_ci.get(w, (None, None))
            ws_m.cell(row=row,column=7,value=ci[0] if ci[0] else round(pred_val*0.8)).number_format='#,##0'
            ws_m.cell(row=row,column=8,value=ci[1] if ci[1] else round(pred_val*1.2)).number_format='#,##0'
            ws_m.cell(row=row,column=9,value=0.7 if w in weekly_ci else 0.4).number_format='0%'
        else:
            for col in [6,7,8]: ws_m.cell(row=row,column=col).number_format='#,##0'
            ws_m.cell(row=row,column=9).number_format='0%'

        # Formulas
        cr=row
        ws_m.cell(row=cr,column=10).value=f'=IF(AND(C{cr}>0,F{cr}>0),ABS(F{cr}-C{cr})/C{cr},"")';ws_m.cell(row=cr,column=10).number_format='0.0%'
        ws_m.cell(row=cr,column=11).value=f'=IF(AND(C{cr}>0,G{cr}>0),IF(AND(C{cr}>=G{cr},C{cr}<=H{cr}),"✓","✗"),"")' 
        if w>1:
            pr=cr-1
            ws_m.cell(row=cr,column=12).value=f'=IF(AND(C{cr}>0,C{pr}>0),(C{cr}-C{pr})/C{pr},"")';ws_m.cell(row=cr,column=12).number_format='+0.0%;-0.0%'

        style_row(ws_m,row,NCOLS_WK,w-1,is_future=is_future)

    wk_data_end = cur_row + 51

    # Conditional formatting
    ws_m.conditional_formatting.add(f'C{wk_data_start}:C{wk_data_end}',
        DataBarRule(start_type='num',start_value=0,end_type='max',color=mc,showValue=True))
    ws_m.conditional_formatting.add(f'J{wk_data_start}:J{wk_data_end}',
        ColorScaleRule(start_type='num',start_value=0,start_color=GREEN,
                       mid_type='num',mid_value=0.1,mid_color='FFFFFF',
                       end_type='num',end_value=0.3,end_color=RED))

    # ═══ CHARTS — right side (col P = 16) ═══
    cats=Reference(ws_m,min_col=2,min_row=wk_data_start,max_row=wk_data_end)  # Wk# for x-axis

    # ── Chart 1: Actual Regs vs Predicted Regs (combo, both lines) ──
    ch=LineChart()
    ch.title=f"{m} — Actual vs Predicted Registrations (W1–W52)"
    ch.style=2;ch.width=32;ch.height=18

    # Y-axis (left) — Actual
    ch.y_axis.title="Actual Regs"
    ch.y_axis.numFmt='#,##0'
    ch.y_axis.majorGridlines=None

    # X-axis
    ch.x_axis.title="Week Number"
    ch.x_axis.tickLblPos="low"
    ch.x_axis.numFmt='0'
    ch.x_axis.scaling.min=1
    ch.x_axis.scaling.max=52

    # Actual regs (col C = 3)
    actual_ref=Reference(ws_m,min_col=3,min_row=wk_hdr_row,max_row=wk_data_end)
    ch.add_data(actual_ref,titles_from_data=True)
    ch.set_categories(cats)
    s1=ch.series[0]
    s1.graphicalProperties.line.solidFill=mc
    s1.graphicalProperties.line.width=28000
    s1.smooth=False

    # Predicted regs (col F = 6) — second Y-axis
    from openpyxl.chart import LineChart as LC2
    ch2_overlay=LC2()
    pred_ref=Reference(ws_m,min_col=6,min_row=wk_hdr_row,max_row=wk_data_end)
    ch2_overlay.add_data(pred_ref,titles_from_data=True)
    ch2_overlay.set_categories(cats)
    ch2_overlay.y_axis.title="Predicted Regs"
    ch2_overlay.y_axis.numFmt='#,##0'
    ch2_overlay.y_axis.axId=200
    ch2_overlay.y_axis.crosses="max"  # right side
    s2=ch2_overlay.series[0]
    s2.graphicalProperties.line.solidFill=ORANGE
    s2.graphicalProperties.line.width=22000
    s2.graphicalProperties.line.dashStyle="dash"
    s2.smooth=False

    # Combine
    ch.y_axis.crosses="min"
    ch += ch2_overlay

    # Legend
    ch.legend.position='b'

    ws_m.add_chart(ch,"P1")

    # ── Chart 2: Actual CPA vs Predicted CPA (combo, both lines) ──
    ch3=LineChart()
    ch3.title=f"{m} — Actual vs Predicted CPA (W1–W52)"
    ch3.style=2;ch3.width=32;ch3.height=16

    ch3.y_axis.title="CPA ($)"
    ch3.y_axis.numFmt='$#,##0.00'
    ch3.x_axis.title="Week Number"
    ch3.x_axis.tickLblPos="low"
    ch3.x_axis.numFmt='0'
    ch3.x_axis.scaling.min=1
    ch3.x_axis.scaling.max=52

    # Actual CPA (col E = 5)
    cpa_ref=Reference(ws_m,min_col=5,min_row=wk_hdr_row,max_row=wk_data_end)
    ch3.add_data(cpa_ref,titles_from_data=True)
    ch3.set_categories(cats)
    s3=ch3.series[0]
    s3.graphicalProperties.line.solidFill=mc
    s3.graphicalProperties.line.width=25000
    s3.smooth=False

    # Legend
    ch3.legend.position='b'

    ws_m.add_chart(ch3,"P35")

    # Rationale
    if proj and proj[9]:
        rat_row=wk_data_end+2
        ws_m.cell(row=rat_row,column=1,value="Latest Projection Rationale").font=Font(bold=True,size=11,color=NAVY,name='Calibri')
        ws_m.merge_cells(f'A{rat_row+1}:N{rat_row+2}')
        ws_m.cell(row=rat_row+1,column=1,value=proj[9]).font=Font(color=GRAY,size=10,italic=True,name='Calibri')

    # Layout NOT controlled by script — Richard owns column widths, row heights, freeze panes, alignment


# ════════════════════════════════════════════
# SAVE & UPLOAD
# ════════════════════════════════════════════
print(f"Saving to {OUT}...")
wb.save(OUT)
con.close()
print(f"Done! {os.path.getsize(OUT):,} bytes")
print(f"Sheets: {[ws.title for ws in wb.worksheets]}")
