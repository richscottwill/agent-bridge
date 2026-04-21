#!/usr/bin/env python3
"""Update ps-forecast-tracker.xlsx — writes ONLY to hidden _Data sheet.

Visible market sheets reference _Data via formulas. This script never touches
visible sheets, preserving all formatting (column widths, row heights, alignments,
freeze panes, chart positions, fonts, fills, borders).

Usage: python3 update-forecast-tracker.py
"""
import duckdb, os
import sys; sys.path.insert(0, os.path.expanduser('~/shared/tools'))
from prediction.config import MOTHERDUCK_TOKEN as TOKEN, MARKETS, ML
XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'ps-forecast-tracker.xlsx')
MONTH_WEEKS = {1:(1,4),2:(5,8),3:(9,13),4:(14,17),5:(18,22),6:(23,26),
               7:(27,31),8:(32,35),9:(36,39),10:(40,44),11:(45,48),12:(49,52)}

# _Data sheet row offsets (must match the template)
WK_START = 2      # Weekly data starts row 2 (52 rows per market)
MO_START = 524    # Monthly data starts row 524 (12 rows per market)
Q_START = 648     # Quarterly starts row 648 (4 rows per market)
YE_START = 692    # Year-end starts row 692 (1 row per market)


def run():
    from openpyxl import load_workbook
    
    print("Connecting to MotherDuck...")
    con = duckdb.connect(f'md:ps_analytics?motherduck_token={TOKEN}', read_only=True)
    
    print(f"Opening {XLSX}...")
    wb = load_workbook(XLSX)
    
    if '_Data' not in wb.sheetnames:
        print("ERROR: _Data sheet not found. Run the template builder first.")
        return
    
    ds = wb['_Data']
    
    for mi, m in enumerate(MARKETS):
        # Pull latest predictions from ps.forecasts directly (authoritative source).
        # ps.forecast_tracker is a snapshot table that can go stale if not
        # refreshed by every pipeline run; ps.forecasts is always current.
        # target_period in ps.forecasts is '2026-W16'/'2026-M04'/'2026-Q2'/'2026-YE';
        # map to the xlsx-shaped horizon/period_label tuples.
        ft = {}
        MONTH_NUM_TO_NAME = {'01':'Jan','02':'Feb','03':'Mar','04':'Apr','05':'May','06':'Jun',
                             '07':'Jul','08':'Aug','09':'Sep','10':'Oct','11':'Nov','12':'Dec'}
        rows = con.execute(f"""
            WITH ranked AS (
              SELECT market, target_period, period_type, metric_name,
                     predicted_value, confidence_low, confidence_high, actual_value,
                     ROW_NUMBER() OVER (
                       PARTITION BY market, target_period, period_type, metric_name
                       ORDER BY created_at DESC
                     ) AS rn
              FROM ps.forecasts
              WHERE market = '{m}' AND target_period LIKE '2026-%'
            )
            SELECT target_period, period_type, metric_name,
                   predicted_value, confidence_low, confidence_high, actual_value
            FROM ranked WHERE rn = 1
        """).fetchall()
        for r in rows:
            target, ptype, metric, pred, ci_lo, ci_hi, actual = r
            # Map target_period → xlsx period_label
            horizon = None; label = None
            if ptype == 'weekly' and '-W' in target:
                horizon = 'weekly'
                label = 'W' + target.split('-W')[-1]
            elif ptype == 'monthly' and '-M' in target:
                horizon = 'monthly'
                label = MONTH_NUM_TO_NAME.get(target.split('-M')[-1])
            elif ptype == 'quarterly' and '-Q' in target:
                horizon = 'quarterly'
                label = 'Q' + target.split('-Q')[-1]
            elif ptype == 'year_end':
                horizon = 'year_end'
                label = '2026'
            if horizon is None or label is None:
                continue
            ft[(horizon, label, metric)] = {
                'pred': pred, 'ci_lo': ci_lo, 'ci_hi': ci_hi, 'actual': actual
            }
        
        # Pull performance (actuals with brand/NB split)
        perf = {}
        try:
            prows = con.execute(f"""SELECT week_num, registrations, cost, cpa, brand_registrations, nb_registrations
                FROM ps.dive_weekly WHERE market='{m}' AND registrations IS NOT NULL""").fetchall()
            for r in prows:
                perf[r[0]] = {'regs': r[1], 'cost': r[2], 'cpa': r[3], 'brand': r[4], 'nb': r[5]}
        except: pass
        
        # Pull OP2 targets
        targets = {}
        try:
            trows = con.execute(f"""SELECT period_key, target_value FROM ps.targets 
                WHERE market='{m}' AND metric_name='registrations' AND period_type='monthly'""").fetchall()
            for r in trows:
                mo = int(r[0].replace('2026-M', ''))
                targets[mo] = r[1]
        except: pass
        
        # ── Write weekly rows ──
        for wk in range(1, 53):
            row = WK_START + mi * 52 + (wk - 1)
            
            p = perf.get(wk)
            if p:
                ds.cell(row=row, column=4).value = p['regs']
                ds.cell(row=row, column=5).value = round(p['cost']) if p['cost'] else None
                ds.cell(row=row, column=6).value = round(p['cpa'], 2) if p['cpa'] else None
                ds.cell(row=row, column=11).value = p.get('brand')
                ds.cell(row=row, column=12).value = p.get('nb')
            
            ft_wk = ft.get(('weekly', f'W{wk}', 'registrations'))
            if ft_wk:
                ds.cell(row=row, column=7).value = ft_wk['pred']
                ds.cell(row=row, column=8).value = ft_wk['ci_lo']
                ds.cell(row=row, column=9).value = ft_wk['ci_hi']
                ds.cell(row=row, column=10).value = 0.7 if ft_wk['pred'] else None
            
            # OP2 weekly
            for mo_num, (ws_s, ws_e) in MONTH_WEEKS.items():
                if ws_s <= wk <= ws_e and mo_num in targets:
                    ds.cell(row=row, column=13).value = round(targets[mo_num] / (ws_e - ws_s + 1))
                    break
        
        # ── Write monthly rows ──
        for mo in range(1, 13):
            row = MO_START + mi * 12 + (mo - 1)
            mo_label = ML[mo]
            ft_mo = ft.get(('monthly', mo_label, 'registrations'))
            if ft_mo:
                ds.cell(row=row, column=4).value = ft_mo['actual']
                ds.cell(row=row, column=5).value = ft_mo['pred']
                ds.cell(row=row, column=6).value = ft_mo['ci_lo']
                ds.cell(row=row, column=7).value = ft_mo['ci_hi']
            if mo in targets:
                ds.cell(row=row, column=8).value = round(targets[mo])
        
        # ── Write quarterly rows ──
        for qi, ql in enumerate(['Q1','Q2','Q3','Q4']):
            row = Q_START + mi * 4 + qi
            ft_q = ft.get(('quarterly', ql, 'registrations'))
            if ft_q:
                ds.cell(row=row, column=4).value = ft_q['actual']
                ds.cell(row=row, column=5).value = ft_q['pred']
                ds.cell(row=row, column=6).value = ft_q['ci_lo']
                ds.cell(row=row, column=7).value = ft_q['ci_hi']
            q_months = [qi*3+1, qi*3+2, qi*3+3]
            q_op2 = sum(targets.get(m2, 0) for m2 in q_months)
            if q_op2: ds.cell(row=row, column=8).value = round(q_op2)
        
        # ── Write year-end row ──
        row = YE_START + mi
        ft_ye = ft.get(('year_end', '2026', 'registrations'))
        if ft_ye:
            ds.cell(row=row, column=2).value = ft_ye['actual']
            ds.cell(row=row, column=3).value = ft_ye['pred']
            ds.cell(row=row, column=4).value = ft_ye['ci_lo']
            ds.cell(row=row, column=5).value = ft_ye['ci_hi']
        annual_op2 = sum(targets.values())
        if annual_op2: ds.cell(row=row, column=6).value = round(annual_op2)
        
        print(f"  {m}: {len(perf)} weeks, {sum(1 for k in ft if ft[k]['pred']) } predictions")
    
    # ── Write _Constraints sheet (single source of truth projection) ──
    write_constraints_sheet(wb, con)
    
    con.close()
    
    print(f"Saving {XLSX}...")
    wb.save(XLSX)
    print(f"Done: {os.path.getsize(XLSX):,} bytes")


CONSTRAINTS_COLUMNS = [
    'market', 'governing_constraint', 'handoff_status', 'oci_status',
    'ccp_availability', 'next_milestone', 'manual_notes',
    'latest_week', 'last_week_regs', 'last_week_cost', 'last_week_cpa',
    'next_week_predicted_regs', 'next_week_ci_low_regs', 'next_week_ci_high_regs',
    'next_week_predicted_cost', 'next_week_ci_low_cost', 'next_week_ci_high_cost',
    'month_op2_regs', 'month_op2_cost', 'month_op2_cpa',
    'hit_rate_regs', 'avg_error_regs',
    'structural_baseline_count', 'structural_baselines',
    'active_impact_count', 'active_impact_regimes',
    'recent_past_count', 'recent_past_regimes',
    'manual_updated_at', 'manual_updated_by',
]


def write_constraints_sheet(wb, con):
    """Write or refresh _Constraints hidden sheet from ps.market_constraints view.
    
    Visible market tabs can reference this sheet via formulas like
    ='_Constraints'!B2 to show governing_constraint for the first market.
    Schema is stable (column order preserved across runs).
    """
    print("\nWriting _Constraints sheet from ps.market_constraints view...")
    col_list = ', '.join(CONSTRAINTS_COLUMNS)
    rows = con.execute(f"""
        SELECT {col_list} FROM ps.market_constraints 
        ORDER BY CASE market WHEN 'WW' THEN 1 WHEN 'US' THEN 2 WHEN 'AU' THEN 3 WHEN 'MX' THEN 4 ELSE 5 END, market
    """).fetchall()
    
    if '_Constraints' in wb.sheetnames:
        del wb['_Constraints']  # rebuild from scratch — schema may evolve
    ws = wb.create_sheet('_Constraints')
    ws.sheet_state = 'hidden'
    
    # Header row
    for ci, col in enumerate(CONSTRAINTS_COLUMNS, start=1):
        ws.cell(row=1, column=ci).value = col
    
    # Data rows
    for ri, row in enumerate(rows, start=2):
        for ci, val in enumerate(row, start=1):
            # openpyxl handles None, strings, numbers, datetimes natively
            ws.cell(row=ri, column=ci).value = val
    
    print(f"  ✓ _Constraints sheet: {len(rows)} markets × {len(CONSTRAINTS_COLUMNS)} cols (hidden)")


if __name__ == '__main__':
    run()
