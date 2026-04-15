#!/usr/bin/env python3
"""
extract-ly-data.py — Extract 2025+2026 daily data from WW Dashboard Excel
and add as _Daily_Data hidden sheet to ps-forecast-tracker.xlsx.

Flat schema — one row per market per day:
  Market, Date, Year, Week, Cost, Clicks, Impressions, Regs,
  Brand_Cost, Brand_Clicks, Brand_Imp, Brand_Regs,
  NB_Cost, NB_Clicks, NB_Imp, NB_Regs

Source: WW Dashboard Excel (market tabs)
Target: ps-forecast-tracker.xlsx → _Daily_Data sheet (hidden)
"""
import openpyxl, re, sys
from pathlib import Path

SCRIPT_DIR = Path(__file__).parent
FORECAST_XLSX = SCRIPT_DIR / "ps-forecast-tracker.xlsx"
MARKETS = ["US", "CA", "UK", "DE", "FR", "IT", "ES", "JP", "MX", "AU"]
YEARS = [2025, 2026]

# Column indices in market sheets (0-indexed)
COL = {
    "week_label": 0, "date": 1, "month": 2,
    "cost": 3, "clicks": 4, "impressions": 5, "regs": 6,
    "brand_cost": 8, "brand_clicks": 9, "brand_imp": 10, "brand_regs": 11,
    "nb_cost": 13, "nb_clicks": 14, "nb_imp": 15, "nb_regs": 16,
}

HEADERS = [
    "Market", "Date", "Year", "Quarter", "Month", "Week", "Week_Label", "Channel",
    "Cost", "Clicks", "Impressions", "Regs",
]

def safe_num(v):
    if v is None: return None
    if isinstance(v, str):
        if "#" in v or v.strip() == "": return None
        try: return float(v)
        except: return None
    try: return float(v)
    except: return None

def find_ww_xlsx():
    for p in [Path.home() / "shared/uploads/sheets", SCRIPT_DIR]:
        if p.exists():
            for f in sorted(p.glob("*AB SEM WW Dashboard*.xlsx"), reverse=True):
                return f
    return None

def main():
    ww_xlsx = find_ww_xlsx()
    if not ww_xlsx:
        print("ERROR: WW Dashboard Excel not found")
        sys.exit(1)
    if not FORECAST_XLSX.exists():
        print(f"ERROR: {FORECAST_XLSX} not found")
        sys.exit(1)

    print(f"Source: {ww_xlsx.name}")
    wb_src = openpyxl.load_workbook(ww_xlsx, data_only=True)

    all_rows = []
    for market in MARKETS:
        sheet_name = next((s for s in wb_src.sheetnames if s.upper() == market.upper()), None)
        if not sheet_name:
            print(f"  WARNING: {market} sheet not found")
            continue
        ws = wb_src[sheet_name]
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            dt = row[COL["date"]]
            if not dt or not hasattr(dt, "year") or dt.year not in YEARS:
                continue
            wk_label = str(row[COL["week_label"]] or "")
            wm = re.search(r"W(\d+)", wk_label)
            wk = int(wm.group(1)) if wm else 0
            month_label = str(row[COL["month"]] or "")
            date_str = dt.strftime("%Y-%m-%d")
            wk_str = f"W{wk}" if wk else ""
            quarter = f"Q{(dt.month - 1) // 3 + 1}"

            base = [market, date_str, dt.year, quarter, month_label, wk, wk_str]

            # Total row
            all_rows.append(base + ["Total",
                safe_num(row[COL["cost"]]),
                safe_num(row[COL["clicks"]]),
                safe_num(row[COL["impressions"]]),
                safe_num(row[COL["regs"]]),
            ])
            # Brand row
            all_rows.append(base + ["Brand",
                safe_num(row[COL["brand_cost"]]),
                safe_num(row[COL["brand_clicks"]]),
                safe_num(row[COL["brand_imp"]]),
                safe_num(row[COL["brand_regs"]]),
            ])
            # Non-Brand row
            all_rows.append(base + ["Non-Brand",
                safe_num(row[COL["nb_cost"]]),
                safe_num(row[COL["nb_clicks"]]),
                safe_num(row[COL["nb_imp"]]),
                safe_num(row[COL["nb_regs"]]),
            ])
            count += 1
        print(f"  {market}: {count} daily rows")

    print(f"\nTotal: {len(all_rows)} rows ({len(YEARS)} years × {len(MARKETS)} markets)")

    # Write to forecast tracker
    print(f"Writing to: {FORECAST_XLSX.name}")
    wb = openpyxl.load_workbook(FORECAST_XLSX)

    # Remove old sheets
    for old_name in ["_LY_Data", "_Daily_Data"]:
        if old_name in wb.sheetnames:
            del wb[old_name]
            print(f"  Removed {old_name}")

    ws = wb.create_sheet("_Daily_Data")
    ws.append(HEADERS)
    for r in all_rows:
        ws.append(r)
    ws.sheet_state = "hidden"

    wb.save(FORECAST_XLSX)
    print(f"  Written {len(all_rows)} rows to _Daily_Data (hidden)")

    # Summary
    from collections import Counter
    year_counts = Counter(r[2] for r in all_rows)
    for y in sorted(year_counts):
        print(f"  {y}: {year_counts[y]} daily rows")

if __name__ == "__main__":
    main()
