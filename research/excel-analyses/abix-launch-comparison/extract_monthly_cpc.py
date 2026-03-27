import openpyxl
from collections import defaultdict
import datetime

wb = openpyxl.load_workbook('shared/research/2026.01_ABIX-Markets_Launch-Comparison.xlsx', data_only=True)
ws = wb['amazon_business_-_Jan_21,_2026_']

# Row 17 has headers
# Col A=Campaign, B=Portfolio, C=Account, D=Search Engine, E=Start Date, F=End Date
# G=Impressions, H=Clicks, I=Cost, J=EFID RefMarker Reg(CT+VT)

# Aggregate by Account + Month: sum clicks, sum cost, compute CPC
data = defaultdict(lambda: {'clicks': 0, 'cost': 0.0, 'impressions': 0, 'regs': 0})

for i, row in enumerate(ws.iter_rows(min_row=18, max_row=ws.max_row, values_only=True), start=18):
    account = row[2]  # Account (col C)
    start_date = row[4]  # Start Date (col E)
    impressions = row[6]  # Impressions (col G)
    clicks = row[7]  # Clicks (col H)
    cost = row[8]  # Cost (col I)
    regs = row[9]  # EFID RefMarker Reg (col J)
    portfolio = row[1]  # Portfolio (col B)

    if account is None or start_date is None:
        continue
    if not isinstance(start_date, datetime.datetime):
        continue
    if clicks is None or cost is None:
        continue

    # Determine market
    if 'AU' in str(account):
        market = 'AU'
    elif 'MX' in str(account):
        market = 'MX'
    elif 'CA' in str(account):
        market = 'CA'
    else:
        continue

    # Determine segment from portfolio
    port = str(portfolio) if portfolio else ''
    if 'Brand' in port or 'brand' in port:
        segment = 'Brand'
    elif 'NB' in port or 'Generic' in port or 'Non' in port:
        segment = 'NB'
    else:
        segment = 'Other'

    month_key = start_date.strftime('%Y-%m')
    key = (market, segment, month_key)

    data[key]['clicks'] += int(clicks) if clicks else 0
    data[key]['cost'] += float(cost) if cost else 0.0
    data[key]['impressions'] += int(impressions) if impressions else 0
    data[key]['regs'] += int(regs) if regs else 0

out = open('shared/research/monthly_cpc_data.txt', 'w')

# Sort and output
out.write('Market\tSegment\tMonth\tClicks\tCost\tCPC\tImpressions\tCTR\tRegs\tCPA\n')
for key in sorted(data.keys()):
    market, segment, month = key
    d = data[key]
    cpc = d['cost'] / d['clicks'] if d['clicks'] > 0 else 0
    ctr = d['clicks'] / d['impressions'] if d['impressions'] > 0 else 0
    cpa = d['cost'] / d['regs'] if d['regs'] > 0 else 0
    out.write('{}\t{}\t{}\t{}\t{:.2f}\t{:.2f}\t{}\t{:.4f}\t{}\t{:.2f}\n'.format(
        market, segment, month, d['clicks'], d['cost'], cpc, d['impressions'], ctr, d['regs'], cpa))

# Also output combined (Brand+NB+Other) by market
out.write('\n\n=== COMBINED (ALL SEGMENTS) BY MARKET ===\n')
combined = defaultdict(lambda: {'clicks': 0, 'cost': 0.0, 'impressions': 0, 'regs': 0})
for key in data:
    market, segment, month = key
    ckey = (market, month)
    combined[ckey]['clicks'] += data[key]['clicks']
    combined[ckey]['cost'] += data[key]['cost']
    combined[ckey]['impressions'] += data[key]['impressions']
    combined[ckey]['regs'] += data[key]['regs']

out.write('Market\tMonth\tClicks\tCost\tCPC\tImpressions\tCTR\tRegs\tCPA\n')
for key in sorted(combined.keys()):
    market, month = key
    d = combined[key]
    cpc = d['cost'] / d['clicks'] if d['clicks'] > 0 else 0
    ctr = d['clicks'] / d['impressions'] if d['impressions'] > 0 else 0
    cpa = d['cost'] / d['regs'] if d['regs'] > 0 else 0
    out.write('{}\t{}\t{}\t{:.2f}\t{:.2f}\t{}\t{:.4f}\t{}\t{:.2f}\n'.format(
        market, month, d['clicks'], d['cost'], cpc, d['impressions'], ctr, d['regs'], cpa))

# NB-only by market (most relevant for the CPC debate)
out.write('\n\n=== NB ONLY BY MARKET ===\n')
out.write('Market\tMonth\tClicks\tCost\tCPC\tRegs\tCPA\n')
for key in sorted(data.keys()):
    market, segment, month = key
    if segment != 'NB':
        continue
    d = data[key]
    cpc = d['cost'] / d['clicks'] if d['clicks'] > 0 else 0
    cpa = d['cost'] / d['regs'] if d['regs'] > 0 else 0
    out.write('{}\t{}\t{}\t{:.2f}\t{:.2f}\t{}\t{:.2f}\n'.format(
        market, month, d['clicks'], d['cost'], cpc, d['regs'], cpa))

# Brand-only by market
out.write('\n\n=== BRAND ONLY BY MARKET ===\n')
out.write('Market\tMonth\tClicks\tCost\tCPC\tRegs\tCPA\n')
for key in sorted(data.keys()):
    market, segment, month = key
    if segment != 'Brand':
        continue
    d = data[key]
    cpc = d['cost'] / d['clicks'] if d['clicks'] > 0 else 0
    cpa = d['cost'] / d['regs'] if d['regs'] > 0 else 0
    out.write('{}\t{}\t{}\t{:.2f}\t{:.2f}\t{}\t{:.2f}\n'.format(
        market, month, d['clicks'], d['cost'], cpc, d['regs'], cpa))

out.close()
print('DONE')
