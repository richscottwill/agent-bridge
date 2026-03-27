import openpyxl
import datetime

wb = openpyxl.load_workbook('shared/context/intake/WW Dashboard - IECCP, WoW & YoY, 2026 Monthly, AU, MX/AB SEM WW Dashboard_Y26 W11.xlsx', data_only=True)

out = open('shared/research/excel-analyses/abix-launch-comparison/wbr_parsed.txt', 'w')

out.write('Sheet names: ' + str(wb.sheetnames) + '\n\n')

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    out.write('=== ' + sheet_name + ' ===\n')
    out.write('Max row: ' + str(ws.max_row) + ', Max col: ' + str(ws.max_column) + '\n')
    # Print all rows for smaller sheets, first 50 for larger ones
    max_rows = min(80, ws.max_row)
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=1):
        vals = []
        for v in row:
            if v is None:
                vals.append('')
            elif isinstance(v, float):
                vals.append(str(round(v, 4)))
            elif isinstance(v, datetime.datetime):
                vals.append(v.strftime('%Y-%m-%d'))
            else:
                vals.append(str(v))
        if any(v != '' for v in vals):
            out.write(str(i) + ': ' + '\t'.join(vals) + '\n')
    if ws.max_row > max_rows:
        out.write('... (' + str(ws.max_row - max_rows) + ' more rows)\n')
    out.write('\n')

out.close()
print('DONE')
