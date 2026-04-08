#!/usr/bin/env python3
"""One-time historical load: 2025 Monthly data → MotherDuck ps.targets + ps.performance.

Reads the '2025 Monthly' tab from the WW Dashboard xlsx and loads:
- OP2 targets (rows 3-12) → ps.targets with period_keys 2025-M01..2025-M12
- Monthly actuals (rows 19-48) → ps.performance with period_type='monthly'
"""

import openpyxl
import duckdb
import os
import sys

XLSX_PATH = os.path.expanduser('~/shared/uploads/sheets/AB SEM WW Dashboard_Y26 W14  vEU5.xlsx')
SHEET_NAME = '2025 Monthly'

# Market order in the sheet (rows 3-12 for OP2, rows 19-48 for actuals in 3-row groups)
MARKET_ORDER_OP2 = ['US', 'JP', 'UK', 'DE', 'FR', 'IT', 'ES', 'CA', 'MX', 'AU']

# Actuals start at row 19 (US SEM), 3 rows per market (total, Brand, NB)
ACTUALS_START_ROW = 19
ACTUALS_MARKETS = ['US', 'JP', 'UK', 'DE', 'FR', 'IT', 'ES', 'CA', 'MX', 'AU']

# Column layout (0-indexed):
# Spend block: col 0=label, cols 1-12=Jan-Dec
# Regs block:  col 19=label, cols 20-31=Jan-Dec
# CPA block:   col 38=label, cols 39-50=Jan-Dec
SPEND_COLS = list(range(1, 13))   # cols 1-12
REGS_COLS = list(range(20, 32))   # cols 20-31
CPA_COLS = list(range(39, 51))    # cols 39-50

MONTH_LABELS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


def safe_float(v):
    if v is None:
        return None
    try:
        f = float(v)
        return f if f != 0 else None
    except (ValueError, TypeError):
        return None


def safe_int(v):
    if v is None:
        return None
    try:
        return int(round(float(v)))
    except (ValueError, TypeError):
        return None


def connect_motherduck():
    token = os.environ.get('MOTHERDUCK_TOKEN',
        'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJlbWFpbCI6InJpY2hzY290dHdpbGxAZ21haWwuY29tIiwibWRSZWdpb24iOiJhd3MtdXMtZWFzdC0xIiwic2Vzc2lvbiI6InJpY2hzY290dHdpbGwuZ21haWwuY29tIiwicGF0IjoiVDNIYzFVQWYzT3o1bjVkLS03ckdHNlBjMlpUdVNNbFItT3RXMS1qNzVPUSIsInVzZXJJZCI6ImU2MDhlNDZiLTE4YzctNGE5Ny04M2I2LWE0N2ZhOThmNjBhYyIsImlzcyI6Im1kX3BhdCIsInJlYWRPbmx5IjpmYWxzZSwidG9rZW5UeXBlIjoicmVhZF93cml0ZSIsImlhdCI6MTc3NTQ0MzY0N30.tS0Cab3FQ8_CDZ1PqOo9z09KYHEUFHwuLVXRQrxcHig')
    return duckdb.connect(f'md:ps_analytics?motherduck_token={token}')


def load_op2_targets(ws, con):
    """Load OP2 targets from rows 3-12 into ps.targets."""
    targets_sql = """
        INSERT OR REPLACE INTO ps.targets
        (market, channel, metric_name, fiscal_year, period_type,
         period_key, target_value, source)
        VALUES (?, 'ps', ?, 2025, 'monthly', ?, ?, '2025_monthly_historical')
    """
    count = 0
    for idx, market in enumerate(MARKET_ORDER_OP2):
        row_num = 3 + idx  # rows 3-12 (1-indexed)
        row = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]

        for m_idx in range(12):
            pk = f"2025-M{m_idx+1:02d}"

            # Spend OP2
            spend_val = safe_float(row[SPEND_COLS[m_idx]] if SPEND_COLS[m_idx] < len(row) else None)
            if spend_val:
                con.execute(targets_sql, [market, 'cost', pk, spend_val])
                count += 1

            # Regs OP2
            regs_val = safe_float(row[REGS_COLS[m_idx]] if REGS_COLS[m_idx] < len(row) else None)
            if regs_val:
                con.execute(targets_sql, [market, 'registrations', pk, regs_val])
                count += 1

            # CPA OP2
            cpa_val = safe_float(row[CPA_COLS[m_idx]] if CPA_COLS[m_idx] < len(row) else None)
            if cpa_val and not isinstance(cpa_val, str):
                con.execute(targets_sql, [market, 'cpa', pk, cpa_val])
                count += 1

    print(f"  Loaded {count} OP2 target rows for 2025")
    return count


def load_monthly_actuals(ws, con):
    """Load monthly actuals from rows 19-48 into ps.performance."""
    perf_sql = """INSERT OR REPLACE INTO ps.performance
        (market, period_type, period_key, period_start,
         registrations, cost, cpa, clicks, impressions, cpc, cvr, ctr,
         brand_registrations, brand_cost, brand_cpa, brand_clicks, brand_cpc, brand_cvr,
         nb_registrations, nb_cost, nb_cpa, nb_clicks, nb_cpc, nb_cvr,
         ieccp, source)
        VALUES (?,?,?,?, ?,?,?,?,?,?,?,?, ?,?,?,?,?,?, ?,?,?,?,?,?, ?,?)"""

    count = 0
    for idx, market in enumerate(ACTUALS_MARKETS):
        total_row_num = ACTUALS_START_ROW + (idx * 3)      # SEM total
        brand_row_num = ACTUALS_START_ROW + (idx * 3) + 1   # Brand
        nb_row_num = ACTUALS_START_ROW + (idx * 3) + 2      # Non-Brand

        total_row = list(ws.iter_rows(min_row=total_row_num, max_row=total_row_num, values_only=True))[0]
        brand_row = list(ws.iter_rows(min_row=brand_row_num, max_row=brand_row_num, values_only=True))[0]
        nb_row = list(ws.iter_rows(min_row=nb_row_num, max_row=nb_row_num, values_only=True))[0]

        for m_idx in range(12):
            pk = f"2025-M{m_idx+1:02d}"
            ps_date = f"2025-{m_idx+1:02d}-01"

            # Spend block (cols 1-12)
            sc = SPEND_COLS[m_idx]
            spend = safe_float(total_row[sc] if sc < len(total_row) else None)
            brand_spend = safe_float(brand_row[sc] if sc < len(brand_row) else None)
            nb_spend = safe_float(nb_row[sc] if sc < len(nb_row) else None)

            # Regs block (cols 20-31)
            rc = REGS_COLS[m_idx]
            regs = safe_int(total_row[rc] if rc < len(total_row) else None)
            brand_regs = safe_int(brand_row[rc] if rc < len(brand_row) else None)
            nb_regs = safe_int(nb_row[rc] if rc < len(nb_row) else None)

            # CPA block (cols 39-50)
            cc = CPA_COLS[m_idx]
            cpa = safe_float(total_row[cc] if cc < len(total_row) else None)
            brand_cpa = safe_float(brand_row[cc] if cc < len(brand_row) else None)
            nb_cpa = safe_float(nb_row[cc] if cc < len(nb_row) else None)

            # Skip months with no data
            if not spend and not regs:
                continue

            con.execute(perf_sql, [
                market, 'monthly', pk, ps_date,
                regs, spend, cpa,
                None, None, None, None, None,  # clicks, impressions, cpc, cvr, ctr
                brand_regs, brand_spend, brand_cpa,
                None, None, None,  # brand_clicks, brand_cpc, brand_cvr
                nb_regs, nb_spend, nb_cpa,
                None, None, None,  # nb_clicks, nb_cpc, nb_cvr
                None,  # ieccp
                '2025_monthly_historical',
            ])
            count += 1

    print(f"  Loaded {count} monthly actuals rows for 2025")
    return count


def main():
    print("Loading 2025 Monthly data into MotherDuck...")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[SHEET_NAME]
    con = connect_motherduck()

    targets_count = load_op2_targets(ws, con)
    actuals_count = load_monthly_actuals(ws, con)

    # Verify
    print("\n=== Verification ===")
    r = con.execute("SELECT COUNT(*) FROM ps.targets WHERE period_key LIKE '2025%'").fetchone()
    print(f"  ps.targets 2025 rows: {r[0]}")

    r = con.execute("SELECT COUNT(*) FROM ps.performance WHERE period_type='monthly' AND period_key LIKE '2025%'").fetchone()
    print(f"  ps.performance 2025 monthly rows: {r[0]}")

    # Sample
    print("\n=== Sample targets (AU, MX) ===")
    rows = con.execute("""
        SELECT market, metric_name, period_key, target_value
        FROM ps.targets WHERE period_key LIKE '2025%' AND market IN ('AU','MX')
        ORDER BY market, metric_name, period_key
    """).fetchall()
    for row in rows:
        print(f"  {row}")

    print("\n=== Sample actuals (AU, MX) ===")
    rows = con.execute("""
        SELECT market, period_key, registrations, cost, cpa,
               brand_registrations, nb_registrations
        FROM ps.performance
        WHERE period_type='monthly' AND period_key LIKE '2025%'
        AND market IN ('AU','MX')
        ORDER BY market, period_key
    """).fetchall()
    for row in rows:
        print(f"  {row}")

    con.close()
    print(f"\nDone. {targets_count} targets + {actuals_count} actuals loaded.")


if __name__ == '__main__':
    main()
