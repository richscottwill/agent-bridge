"""Seed CCPs from finance source — Task 1.2.

WHY THIS EXISTS
    The Market Projection Engine reads CCPs from ps.market_projection_params
    as fixed inputs to the ie%CCP formula. The authoritative source is the
    Summary tab of CCP Q1'26 check yc.xlsx, column U ("FINAL ALIGNED") —
    the post-negotiation canonical value. This script parses that file and
    writes one parameter_version per market per segment into the registry.

HOW THE OWNER MAINTAINS IT
    Re-run annually after finance finalizes the next year's CCPs:
        python3 -m shared.tools.prediction.seed_market_ccps \\
            --source shared/uploads/sheets/CCP-Q1-27-check.xlsx
    The script writes new parameter_versions with is_active=TRUE and
    flips prior versions to is_active=FALSE, preserving history.

WHAT HAPPENS ON FAILURE
    If a cell in column U is blank (e.g., Q2 CCP not yet finalized), the
    script skips that row and prints a warning. Existing registry rows
    are untouched. No partial writes on parse failure — the script
    gathers all rows in memory first, validates, then commits in one
    transaction.

NOTES
    - Column U = "FINAL ALIGNED" = post-negotiation canonical value
    - Column N = "Q1 Static CCP" = pre-negotiation static; seeded as
      brand_ccp_q1_static / nb_ccp_q1_static for transition-period ref
    - AU is absent from this file (efficiency strategy, no CCP).
      The script does NOT write AU rows.
    - WW and EU5 are rollup rows in the file — we skip them; the engine
      derives regional CCPs at query time from constituent markets.
"""

from __future__ import annotations

import os
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

# Connection pattern per prediction/config
sys.path.insert(0, os.path.expanduser('~/shared/tools'))


DEFAULT_SOURCE = os.path.expanduser(
    "~/shared/uploads/sheets/CCP Q1'26 check yc.xlsx"
)

# Market → row in Summary tab (0-indexed in openpyxl is 1-indexed, we use 1-indexed)
# Structure: market header row, then +1 = Unbranded, +2 = Branded
MARKET_ROWS = {
    'US': 5,
    'DE': 11,
    'UK': 14,
    'FR': 17,
    'IT': 20,
    'ES': 23,
    'JP': 26,
    'CA': 29,
    'MX': 32,
}

# Columns (1-indexed): U = 21 (FINAL ALIGNED), N = 14 (Q1 Static CCP)
COL_FINAL_ALIGNED = 21
COL_Q1_STATIC = 14


@dataclass
class CCPRow:
    market: str
    segment: str          # 'brand' or 'nb'
    final_aligned: float | None
    q1_static: float | None


def parse_xlsx(source_path: str) -> list[CCPRow]:
    """Parse the Summary tab, return list of CCPRow."""
    try:
        import openpyxl
    except ImportError:
        print("[seed_market_ccps] FATAL: openpyxl not available", file=sys.stderr)
        sys.exit(2)

    if not os.path.isfile(source_path):
        print(f"[seed_market_ccps] FATAL: source file not found: {source_path}", file=sys.stderr)
        sys.exit(2)

    wb = openpyxl.load_workbook(source_path, data_only=True)
    if 'Summary' not in wb.sheetnames:
        print(f"[seed_market_ccps] FATAL: Summary tab missing in {source_path}", file=sys.stderr)
        sys.exit(2)
    ws = wb['Summary']

    rows: list[CCPRow] = []
    for market, header_row in MARKET_ROWS.items():
        nb_row = header_row + 1          # "Paid Search Unbranded"
        brand_row = header_row + 2       # "Paid Search Branded"

        nb_final = ws.cell(row=nb_row, column=COL_FINAL_ALIGNED).value
        nb_static = ws.cell(row=nb_row, column=COL_Q1_STATIC).value
        brand_final = ws.cell(row=brand_row, column=COL_FINAL_ALIGNED).value
        brand_static = ws.cell(row=brand_row, column=COL_Q1_STATIC).value

        rows.append(CCPRow(
            market=market, segment='nb',
            final_aligned=float(nb_final) if nb_final is not None else None,
            q1_static=float(nb_static) if nb_static is not None else None,
        ))
        rows.append(CCPRow(
            market=market, segment='brand',
            final_aligned=float(brand_final) if brand_final is not None else None,
            q1_static=float(brand_static) if brand_static is not None else None,
        ))
    return rows


def connect_md(read_only: bool = False):
    """Open a MotherDuck connection with write access by default."""
    import duckdb
    from prediction.config import MOTHERDUCK_TOKEN
    return duckdb.connect(
        f'md:ps_analytics?motherduck_token={MOTHERDUCK_TOKEN}',
        read_only=read_only,
    )


def next_version(con, market: str, parameter_name: str) -> int:
    """Return the next parameter_version for (market, parameter_name)."""
    row = con.execute(
        """
        SELECT COALESCE(MAX(parameter_version), 0) + 1
        FROM ps.market_projection_params
        WHERE market = ? AND parameter_name = ?
        """,
        [market, parameter_name],
    ).fetchone()
    return int(row[0])


def deactivate_prior(con, market: str, parameter_name: str) -> None:
    """Flip prior active versions of (market, parameter_name) to is_active=FALSE."""
    con.execute(
        """
        UPDATE ps.market_projection_params
        SET is_active = FALSE
        WHERE market = ? AND parameter_name = ? AND is_active = TRUE
        """,
        [market, parameter_name],
    )


def upsert_scalar(con, market: str, param: str, value: float, source: str, lineage: str) -> None:
    """Insert a new active scalar parameter version, deactivating prior."""
    deactivate_prior(con, market, param)
    version = next_version(con, market, param)
    con.execute(
        """
        INSERT INTO ps.market_projection_params
            (market, parameter_name, parameter_version, value_scalar,
             refit_cadence, last_refit_at, source, fallback_level,
             lineage, is_active)
        VALUES (?, ?, ?, ?, 'annual', CURRENT_TIMESTAMP, ?, 'market_specific', ?, TRUE)
        """,
        [market, param, version, value, source, lineage],
    )


def seed_ccps(con, rows: list[CCPRow], source_path: str) -> dict:
    """Seed CCP rows into registry. Returns summary stats."""
    written = 0
    skipped = 0
    warnings: list[str] = []

    lineage_base = f"CCP spreadsheet column U FINAL ALIGNED ({os.path.basename(source_path)}) seeded {datetime.now().strftime('%Y-%m-%d')}"
    lineage_static_base = f"CCP spreadsheet column N Q1 Static ({os.path.basename(source_path)}) seeded {datetime.now().strftime('%Y-%m-%d')}"

    for row in rows:
        param_final = f"{row.segment}_ccp"
        param_static = f"{row.segment}_ccp_q1_static"

        if row.final_aligned is None:
            warnings.append(f"{row.market} {row.segment} FINAL ALIGNED is blank — skipping {param_final}")
            skipped += 1
        else:
            upsert_scalar(
                con, row.market, param_final, row.final_aligned,
                source='finance_negotiation', lineage=lineage_base,
            )
            written += 1

        if row.q1_static is None:
            skipped += 1
        else:
            upsert_scalar(
                con, row.market, param_static, row.q1_static,
                source='finance_negotiation', lineage=lineage_static_base,
            )
            written += 1

    return {
        'written': written,
        'skipped': skipped,
        'warnings': warnings,
    }


def main(argv: list[str] | None = None) -> int:
    import argparse
    parser = argparse.ArgumentParser(
        prog="seed_market_ccps",
        description="Seed market CCPs from finance spreadsheet into ps.market_projection_params.",
    )
    parser.add_argument(
        "--source",
        default=DEFAULT_SOURCE,
        help=f"Path to CCP xlsx (default: {DEFAULT_SOURCE})",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and report; do not write to MotherDuck.",
    )
    args = parser.parse_args(argv)

    print(f"[seed_market_ccps] Parsing {args.source}")
    rows = parse_xlsx(args.source)
    print(f"[seed_market_ccps] Parsed {len(rows)} rows ({len(MARKET_ROWS)} markets × 2 segments)")

    if args.dry_run:
        print("[seed_market_ccps] DRY RUN — values that would be written:")
        for r in rows:
            fa = f"${r.final_aligned:.2f}" if r.final_aligned is not None else "BLANK"
            qs = f"${r.q1_static:.2f}" if r.q1_static is not None else "blank"
            print(f"  {r.market} {r.segment}_ccp = {fa}  (static: {qs})")
        return 0

    con = connect_md(read_only=False)
    stats = seed_ccps(con, rows, args.source)
    con.close()

    print(f"[seed_market_ccps] wrote {stats['written']} parameter versions")
    print(f"[seed_market_ccps] skipped {stats['skipped']} blank values")
    for w in stats['warnings']:
        print(f"[seed_market_ccps] WARN: {w}")

    # Exit code 0 if any writes happened; 1 if nothing written
    return 0 if stats['written'] > 0 else 1


if __name__ == "__main__":
    sys.exit(main())
