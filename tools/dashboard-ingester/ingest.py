#!/usr/bin/env python3
"""
WW Dashboard Ingester — Weekly Performance Analysis Tool

Reads the AB SEM WW Dashboard xlsx file and produces:
1. Structured JSON data extract (for programmatic use)
2. Markdown callout drafts per market (AU, MX, and all WW markets)
3. WW summary with trends, anomalies, and projections

Usage:
    python3 ingest.py <path_to_xlsx> [--week YYYY_WNN] [--markets AU,MX] [--output-dir DIR]

If --week is omitted, auto-detects the latest week with data.
If --markets is omitted, processes all markets.
"""

import openpyxl
import datetime
import json
import sys
import os
import argparse
from collections import defaultdict, OrderedDict
from pathlib import Path

# ── Column layout for daily market tabs ──
# A=Weeks, B=Date, C=Month, D=Cost, E=Clicks, F=Impressions, G=Reg
# (H=spacer)
# I=Brand Cost, J=Brand Clicks, K=Brand Imp, L=Brand Reg
# (M=spacer)
# N=NB Cost, O=NB Clicks, P=NB Imp, Q=NB Reg

DAILY_COLS = {
    'week': 0, 'date': 1, 'month': 2,
    'cost': 3, 'clicks': 4, 'impressions': 5, 'regs': 6,
    'brand_cost': 8, 'brand_clicks': 9, 'brand_imp': 10, 'brand_regs': 11,
    'nb_cost': 13, 'nb_clicks': 14, 'nb_imp': 15, 'nb_regs': 16,
}

ALL_MARKETS = ['US', 'CA', 'UK', 'DE', 'FR', 'IT', 'ES', 'JP', 'AU', 'MX']

# Weekly tab metric blocks (row offsets from block start)
# Each block: row 1=header, rows 2-31 = markets (US,JP,UK,DE,FR,IT,ES,CA,MX,AU + Brand/NB per market)
WEEKLY_METRICS = {
    'spend': 1, 'regs': 40, 'cpa': 79, 'clicks': 118,
    'impressions': 157, 'cpc': 196, 'cvr': 235, 'ctr': 274,
}

# Market row offsets within each Weekly metric block (0-indexed from block start)
WEEKLY_MARKET_ROWS = {
    'US': (1, 2, 3),    # total, brand, nb
    'JP': (4, 5, 6),
    'UK': (7, 8, 9),
    'DE': (10, 11, 12),
    'FR': (13, 14, 15),
    'IT': (16, 17, 18),
    'ES': (19, 20, 21),
    'CA': (22, 23, 24),
    'MX': (25, 26, 27),
    'AU': (28, 29, 30),
}


def safe_float(v):
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def safe_int(v):
    if v is None:
        return 0
    try:
        return int(round(float(v)))
    except (ValueError, TypeError):
        return 0


def pct(a, b):
    """Percentage change from b to a."""
    if b == 0:
        return None
    return (a - b) / b


def fmt_pct(v):
    if v is None:
        return 'N/A'
    return f'{v:+.1%}'


def fmt_dollar(v):
    if abs(v) >= 1000:
        return f'${v:,.0f}'
    return f'${v:,.2f}'


def fmt_int(v):
    return f'{v:,}'



class DashboardIngester:
    def __init__(self, xlsx_path):
        self.xlsx_path = xlsx_path
        self.wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        self.sheet_names = self.wb.sheetnames

    def _read_daily_tab(self, market):
        """Read daily data from a market tab. Returns list of day dicts."""
        if market not in self.sheet_names:
            return []
        ws = self.wb[market]
        days = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            week_label = str(row[DAILY_COLS['week']]) if row[DAILY_COLS['week']] else ''
            date_val = row[DAILY_COLS['date']]
            if not week_label or not date_val:
                continue
            if isinstance(date_val, datetime.datetime):
                date_str = date_val.strftime('%Y-%m-%d')
            else:
                date_str = str(date_val)

            cost = safe_float(row[DAILY_COLS['cost']])
            regs = safe_int(row[DAILY_COLS['regs']])
            # Skip rows with no data (future dates)
            if cost == 0 and regs == 0 and safe_int(row[DAILY_COLS['clicks']]) == 0:
                continue

            day = {
                'week': week_label,
                'date': date_str,
                'month': str(row[DAILY_COLS['month']]) if row[DAILY_COLS['month']] else '',
                'cost': cost,
                'clicks': safe_int(row[DAILY_COLS['clicks']]),
                'impressions': safe_int(row[DAILY_COLS['impressions']]),
                'regs': regs,
                'brand_cost': safe_float(row[DAILY_COLS['brand_cost']]),
                'brand_clicks': safe_int(row[DAILY_COLS['brand_clicks']]),
                'brand_imp': safe_int(row[DAILY_COLS['brand_imp']]),
                'brand_regs': safe_int(row[DAILY_COLS['brand_regs']]),
                'nb_cost': safe_float(row[DAILY_COLS['nb_cost']]),
                'nb_clicks': safe_int(row[DAILY_COLS['nb_clicks']]),
                'nb_imp': safe_int(row[DAILY_COLS['nb_imp']]),
                'nb_regs': safe_int(row[DAILY_COLS['nb_regs']]),
            }
            days.append(day)
        return days

    def _aggregate_week(self, days, week_label):
        """Aggregate daily data for a given week label."""
        week_days = [d for d in days if d['week'] == week_label]
        if not week_days:
            return None
        agg = {
            'week': week_label,
            'date_range': f"{week_days[0]['date']} to {week_days[-1]['date']}",
            'num_days': len(week_days),
            'cost': sum(d['cost'] for d in week_days),
            'clicks': sum(d['clicks'] for d in week_days),
            'impressions': sum(d['impressions'] for d in week_days),
            'regs': sum(d['regs'] for d in week_days),
            'brand_cost': sum(d['brand_cost'] for d in week_days),
            'brand_clicks': sum(d['brand_clicks'] for d in week_days),
            'brand_imp': sum(d['brand_imp'] for d in week_days),
            'brand_regs': sum(d['brand_regs'] for d in week_days),
            'nb_cost': sum(d['nb_cost'] for d in week_days),
            'nb_clicks': sum(d['nb_clicks'] for d in week_days),
            'nb_imp': sum(d['nb_imp'] for d in week_days),
            'nb_regs': sum(d['nb_regs'] for d in week_days),
        }
        # Derived metrics
        agg['cpa'] = agg['cost'] / agg['regs'] if agg['regs'] > 0 else None
        agg['cpc'] = agg['cost'] / agg['clicks'] if agg['clicks'] > 0 else None
        agg['cvr'] = agg['regs'] / agg['clicks'] if agg['clicks'] > 0 else None
        agg['ctr'] = agg['clicks'] / agg['impressions'] if agg['impressions'] > 0 else None
        agg['brand_cpa'] = agg['brand_cost'] / agg['brand_regs'] if agg['brand_regs'] > 0 else None
        agg['brand_cpc'] = agg['brand_cost'] / agg['brand_clicks'] if agg['brand_clicks'] > 0 else None
        agg['brand_cvr'] = agg['brand_regs'] / agg['brand_clicks'] if agg['brand_clicks'] > 0 else None
        agg['nb_cpa'] = agg['nb_cost'] / agg['nb_regs'] if agg['nb_regs'] > 0 else None
        agg['nb_cpc'] = agg['nb_cost'] / agg['nb_clicks'] if agg['nb_clicks'] > 0 else None
        agg['nb_cvr'] = agg['nb_regs'] / agg['nb_clicks'] if agg['nb_clicks'] > 0 else None
        return agg

    def _get_unique_weeks(self, days):
        """Get ordered list of unique week labels from daily data."""
        seen = set()
        weeks = []
        for d in days:
            if d['week'] not in seen:
                seen.add(d['week'])
                weeks.append(d['week'])
        return weeks

    def _detect_latest_week(self, days):
        """Find the latest week that has a full 7 days of data."""
        weeks = self._get_unique_weeks(days)
        # Walk backwards to find the latest complete week
        for w in reversed(weeks):
            week_days = [d for d in days if d['week'] == w]
            if len(week_days) == 7 and sum(d['regs'] for d in week_days) > 0:
                return w
        # Fallback: latest week with any data
        for w in reversed(weeks):
            week_days = [d for d in days if d['week'] == w]
            if sum(d['regs'] for d in week_days) > 0:
                return w
        return weeks[-1] if weeks else None

    def _find_yoy_week(self, week_label):
        """Given '2026 W12', return '2025 W12'."""
        parts = week_label.split()
        if len(parts) == 2:
            year = int(parts[0])
            return f'{year - 1} {parts[1]}'
        return None

    def _get_week_number(self, week_label):
        """Extract week number from '2026 W12' -> 12."""
        parts = week_label.split()
        if len(parts) == 2 and parts[1].startswith('W'):
            return int(parts[1][1:])
        return None

    def _get_week_year(self, week_label):
        """Extract year from '2026 W12' -> 2026."""
        parts = week_label.split()
        if len(parts) == 2:
            return int(parts[0])
        return None

    def _read_ieccp(self, target_week):
        """Read ie%CCP values from the IECCP tab for the target week and prior week."""
        if 'IECCP' not in self.sheet_names:
            return {}
        ws = self.wb['IECCP']
        # Row 1 = week headers (col 0 = label, cols 1+ = weeks)
        # Rows 15-26 = IECCP values per market (row 26 = MX)
        row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        tw_col = None
        lw_col = None
        for i, v in enumerate(row1):
            if v and str(v).strip() == target_week:
                tw_col = i
                lw_col = i - 1 if i > 0 else None
                break
        if tw_col is None:
            return {}

        # IECCP market rows: 16=US, 17=JP, 18=UK, 19=DE, 20=FR, 21=IT, 22=ES, 23=CA, 24=WW, 25=EU5, 26=MX
        ieccp_rows = {'US': 16, 'JP': 17, 'UK': 18, 'DE': 19, 'FR': 20, 'IT': 21, 'ES': 22, 'CA': 23, 'MX': 26}
        result = {}
        for market, row_num in ieccp_rows.items():
            row = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
            tw_val = safe_float(row[tw_col]) if tw_col < len(row) else None
            lw_val = safe_float(row[lw_col]) if lw_col and lw_col < len(row) else None
            if tw_val:
                result[market] = {'tw': tw_val, 'lw': lw_val}
        return result

    def _read_weekly_tab(self, market, target_week, num_weeks=52):
        """Read the Weekly tab for a market's historical weekly data.

        Returns a list of dicts with weekly metrics going back num_weeks from target_week.
        Each dict has: week, spend, brand_spend, nb_spend, regs, brand_regs, nb_regs,
        cpa, clicks, brand_clicks, nb_clicks, impressions, cpc, cvr, ctr.
        """
        if 'Weekly' not in self.sheet_names:
            return []
        ws = self.wb['Weekly']

        # Map metric block start rows (1-indexed)
        BLOCK_ROWS = {
            'spend': 1, 'regs': 40, 'cpa': 79, 'clicks': 118,
            'impressions': 157, 'cpc': 196, 'cvr': 235, 'ctr': 274,
        }

        # Market row offsets within each block (0-indexed from block start)
        # Each market has 3 rows: total, brand, nb
        MARKET_OFFSETS = {
            'US': 1, 'JP': 4, 'UK': 7, 'DE': 10, 'FR': 13,
            'IT': 16, 'ES': 19, 'CA': 22, 'MX': 25, 'AU': 28,
        }

        if market not in MARKET_OFFSETS:
            return []

        offset = MARKET_OFFSETS[market]

        # Get week column headers from row 1
        row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        # Find target week column
        tw_col = None
        for i, v in enumerate(row1):
            if v and str(v).strip() == target_week:
                tw_col = i
                break
        if tw_col is None:
            return []

        # Determine column range
        start_col = max(1, tw_col - num_weeks + 1)
        week_cols = list(range(start_col, tw_col + 1))

        # Pre-read all needed rows (cache them)
        def get_row(row_num):
            return list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]

        # For each metric, read total/brand/nb rows
        rows_cache = {}
        for metric, block_start in BLOCK_ROWS.items():
            total_row = block_start + offset
            brand_row = block_start + offset + 1
            nb_row = block_start + offset + 2
            rows_cache[f'{metric}_total'] = get_row(total_row)
            rows_cache[f'{metric}_brand'] = get_row(brand_row)
            rows_cache[f'{metric}_nb'] = get_row(nb_row)

        # Build weekly data
        weeks = []
        for col in week_cols:
            week_label = str(row1[col]).strip() if col < len(row1) and row1[col] else ''
            if not week_label:
                continue

            def val(key, c):
                r = rows_cache.get(key)
                if r and c < len(r):
                    return safe_float(r[c])
                return 0.0

            w = {
                'week': week_label,
                'spend': val('spend_total', col),
                'brand_spend': val('spend_brand', col),
                'nb_spend': val('spend_nb', col),
                'regs': safe_int(val('regs_total', col)),
                'brand_regs': safe_int(val('regs_brand', col)),
                'nb_regs': safe_int(val('regs_nb', col)),
                'cpa': val('cpa_total', col) or None,
                'brand_cpa': val('cpa_brand', col) or None,
                'nb_cpa': val('cpa_nb', col) or None,
                'clicks': safe_int(val('clicks_total', col)),
                'brand_clicks': safe_int(val('clicks_brand', col)),
                'nb_clicks': safe_int(val('clicks_nb', col)),
                'impressions': safe_int(val('impressions_total', col)),
                'cpc': val('cpc_total', col) or None,
                'nb_cpc': val('cpc_nb', col) or None,
                'brand_cpc': val('cpc_brand', col) or None,
                'cvr': val('cvr_total', col) or None,
                'brand_cvr': val('cvr_brand', col) or None,
                'nb_cvr': val('cvr_nb', col) or None,
                'ctr': val('ctr_total', col) or None,
            }
            # Skip weeks with no data
            if w['regs'] > 0 or w['spend'] > 0:
                weeks.append(w)

        return weeks


    def _read_monthly_actuals(self):
        """Read 2026 Monthly tab for actual spend/regs by month, with Brand/NB split.

        Layout (rows 19-48):
        Each market has 3 rows: total (X SEM), Brand, Non-Brand.
        Spend actuals: cols 1-12 (Jan-Dec), cols 13-16 (Q1-Q4), col 17 (Total)
        Reg actuals: cols 20-31 (Jan-Dec), cols 32-35 (Q1-Q4)
        """
        if '2026 Monthly' not in self.sheet_names:
            return {}
        ws = self.wb['2026 Monthly']
        monthly_months = [
            '2026 Jan', '2026 Feb', '2026 Mar', '2026 Apr', '2026 May', '2026 Jun',
            '2026 Jul', '2026 Aug', '2026 Sep', '2026 Oct', '2026 Nov', '2026 Dec',
        ]
        # Market total rows (0-indexed from row 19)
        market_rows = {
            'US': 19, 'JP': 22, 'UK': 25, 'DE': 28, 'FR': 31,
            'IT': 34, 'ES': 37, 'CA': 40, 'MX': 43, 'AU': 46,
        }
        actuals = {}
        all_rows = list(ws.iter_rows(min_row=1, max_row=50, max_col=40, values_only=True))
        for market, total_row in market_rows.items():
            if total_row > len(all_rows):
                continue
            total = all_rows[total_row - 1]  # 1-indexed to 0-indexed
            brand = all_rows[total_row]       # next row
            nb = all_rows[total_row + 1]      # row after that
            months = {}
            for j, month in enumerate(monthly_months):
                spend = safe_float(total[1 + j])
                regs = safe_int(total[20 + j])
                brand_spend = safe_float(brand[1 + j])
                brand_regs = safe_int(brand[20 + j])
                nb_spend = safe_float(nb[1 + j])
                nb_regs = safe_int(nb[20 + j])
                # Skip months with no data
                if spend == 0 and regs == 0:
                    continue
                m = {
                    'spend': spend, 'regs': regs,
                    'brand_spend': brand_spend, 'brand_regs': brand_regs,
                    'nb_spend': nb_spend, 'nb_regs': nb_regs,
                }
                # Derived metrics
                m['cpa'] = spend / regs if regs > 0 else None
                m['brand_cpa'] = brand_spend / brand_regs if brand_regs > 0 else None
                m['nb_cpa'] = nb_spend / nb_regs if nb_regs > 0 else None
                months[month] = m
            actuals[market] = months
        return actuals

    def _read_monthly_budget(self):
        """Read 2026 Monthly tab for OP2 budget targets."""
        if '2026 Monthly' not in self.sheet_names:
            return {}
        ws = self.wb['2026 Monthly']
        budgets = {}
        # Layout: Rows 3-12 are OP2 budget (one row per market SEM).
        # Rows 17+ are Actual Spend — we must stop before those.
        # Spend OP2: col 0=Channel, cols 1-12=monthly spend
        # Reg OP2:   col 19=Channel, cols 20-31=monthly regs
        monthly_months = [
            '2026 Jan', '2026 Feb', '2026 Mar', '2026 Apr', '2026 May', '2026 Jun',
            '2026 Jul', '2026 Aug', '2026 Sep', '2026 Oct', '2026 Nov', '2026 Dec',
        ]
        for row in ws.iter_rows(min_row=3, max_row=14, max_col=40, values_only=True):
            market_label = str(row[0]).strip() if row[0] else ''
            if not market_label or 'SEM' not in market_label:
                continue
            market = market_label.replace(' SEM', '').strip()
            spend_vals = {}
            reg_vals = {}
            for j, month in enumerate(monthly_months):
                spend_vals[month] = safe_float(row[1 + j])      # cols 1-12
                reg_vals[month] = safe_int(row[20 + j])          # cols 20-31
            budgets[market] = {'spend_op2': spend_vals, 'regs_op2': reg_vals}
        return budgets

    def _read_wow_yoy_tab(self):
        """Read the pre-computed WoW & YoY summary tab."""
        if 'WoW & YoY' not in self.sheet_names:
            return {}
        ws = self.wb['WoW & YoY']
        data = {}
        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=47, values_only=True))

        # Row 3 has TW/LW week labels
        tw_label = str(rows[2][0]) if rows[2][0] else ''
        lw_label = str(rows[2][1]) if rows[2][1] else ''
        data['tw'] = tw_label
        data['lw'] = lw_label

        # Row 64 has TY/LY labels
        if len(rows) > 63:
            ty_label = str(rows[63][0]) if rows[63][0] else ''
            ly_label = str(rows[63][1]) if rows[63][1] else ''
            data['ty'] = ty_label
            data['ly'] = ly_label

        return data

    def _compute_multi_week_trend(self, days, target_week, num_weeks=6):
        """Compute trend over the last N weeks for key metrics."""
        weeks = self._get_unique_weeks(days)
        if target_week not in weeks:
            return []
        idx = weeks.index(target_week)
        start_idx = max(0, idx - num_weeks + 1)
        trend_weeks = weeks[start_idx:idx + 1]

        trend = []
        for w in trend_weeks:
            agg = self._aggregate_week(days, w)
            if agg:
                trend.append(agg)
        return trend

    def _detect_anomalies(self, trend, current_week_agg):
        """Detect anomalies in the current week vs recent trend."""
        if len(trend) < 3 or not current_week_agg:
            return []

        anomalies = []
        # Use the weeks before current as baseline
        baseline = trend[:-1]

        for metric in ['regs', 'cost', 'cpa', 'cvr', 'brand_regs', 'nb_regs', 'brand_cvr', 'nb_cvr', 'cpc']:
            baseline_vals = [w.get(metric) for w in baseline if w.get(metric) is not None]
            current_val = current_week_agg.get(metric)
            if not baseline_vals or current_val is None:
                continue

            avg = sum(baseline_vals) / len(baseline_vals)
            if avg == 0:
                continue

            deviation = (current_val - avg) / abs(avg)

            # Flag if >20% deviation from recent average
            # Suppress noise: require minimum absolute values for small-volume metrics
            min_thresholds = {'nb_regs': 20, 'brand_regs': 20, 'nb_cvr': 0.005}
            min_val = min_thresholds.get(metric, 0)
            if max(abs(current_val), abs(avg)) < min_val:
                continue

            if abs(deviation) > 0.20:
                direction = 'above' if deviation > 0 else 'below'
                anomalies.append({
                    'metric': metric,
                    'current': current_val,
                    'avg': avg,
                    'deviation_pct': deviation,
                    'direction': direction,
                    'weeks_in_baseline': len(baseline_vals),
                })

        # Check for data lag: if last 2 days of the week have <30% of the daily avg
        # of the first 5 days, flag as potential incomplete data
        daily = current_week_agg.get('num_days', 7)
        if daily == 7:
            # We need the raw daily data — check via daily_patterns if available
            pass  # handled in generate_callout_md via daily_patterns

        return anomalies

    def _compute_daily_patterns(self, days, target_week):
        """Analyze day-of-week patterns within the target week."""
        week_days = [d for d in days if d['week'] == target_week]
        if not week_days:
            return {}

        dow_names = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
        patterns = []
        for d in week_days:
            dt = datetime.datetime.strptime(d['date'], '%Y-%m-%d')
            dow = dt.weekday()
            patterns.append({
                'date': d['date'],
                'dow': dow_names[dow],
                'regs': d['regs'],
                'cost': d['cost'],
                'brand_regs': d['brand_regs'],
                'nb_regs': d['nb_regs'],
            })

        # Find best/worst days
        if patterns:
            best_day = max(patterns, key=lambda x: x['regs'])
            worst_day = min(patterns, key=lambda x: x['regs'])
            weekday_avg = sum(p['regs'] for p in patterns[:5]) / min(5, len([p for p in patterns if p['dow'] not in ['Sat', 'Sun']]) or 1)
            weekend_avg = sum(p['regs'] for p in patterns if p['dow'] in ['Sat', 'Sun']) / max(1, len([p for p in patterns if p['dow'] in ['Sat', 'Sun']]))

            return {
                'daily': patterns,
                'best_day': best_day,
                'worst_day': worst_day,
                'weekday_avg_regs': weekday_avg,
                'weekend_avg_regs': weekend_avg,
            }
        return {}

    def _project_month(self, days, target_week, market, budgets):
        """Project month-end performance based on daily data and remaining days."""
        week_days = [d for d in days if d['week'] == target_week]
        if not week_days:
            return None

        # Determine the month
        month_label = week_days[0]['month']  # e.g., '2026 Mar'
        if not month_label:
            return None

        # Get all days in this month so far
        month_days = [d for d in days if d['month'] == month_label and d['regs'] > 0]
        if not month_days:
            return None

        # Parse month to get total days
        try:
            parts = month_label.split()
            year = int(parts[0])
            month_abbr = parts[1]
            month_map = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
                         'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12}
            month_num = month_map.get(month_abbr, 3)
            import calendar
            total_days_in_month = calendar.monthrange(year, month_num)[1]
        except:
            return None

        days_elapsed = len(month_days)
        days_remaining = total_days_in_month - days_elapsed

        mtd_spend = sum(d['cost'] for d in month_days)
        mtd_regs = sum(d['regs'] for d in month_days)
        mtd_brand_regs = sum(d['brand_regs'] for d in month_days)
        mtd_nb_regs = sum(d['nb_regs'] for d in month_days)

        # Use recent daily average (last 7 days) for projection
        recent = month_days[-7:] if len(month_days) >= 7 else month_days
        daily_avg_spend = sum(d['cost'] for d in recent) / len(recent)
        daily_avg_regs = sum(d['regs'] for d in recent) / len(recent)

        projected_spend = mtd_spend + (daily_avg_spend * days_remaining)
        projected_regs = mtd_regs + (daily_avg_regs * days_remaining)
        projected_cpa = projected_spend / projected_regs if projected_regs > 0 else None

        # Get OP2 targets
        op2_spend = None
        op2_regs = None
        if market in budgets:
            op2_spend = budgets[market]['spend_op2'].get(month_label)
            op2_regs = budgets[market]['regs_op2'].get(month_label)

        return {
            'month': month_label,
            'days_elapsed': days_elapsed,
            'days_remaining': days_remaining,
            'total_days': total_days_in_month,
            'mtd_spend': mtd_spend,
            'mtd_regs': mtd_regs,
            'mtd_brand_regs': mtd_brand_regs,
            'mtd_nb_regs': mtd_nb_regs,
            'projected_spend': projected_spend,
            'projected_regs': int(round(projected_regs)),
            'projected_cpa': projected_cpa,
            'op2_spend': op2_spend,
            'op2_regs': op2_regs,
            'vs_op2_spend_pct': pct(projected_spend, op2_spend) if op2_spend else None,
            'vs_op2_regs_pct': pct(projected_regs, op2_regs) if op2_regs else None,
        }


    def analyze_market(self, market, target_week=None):
        """Full analysis for a single market."""
        days = self._read_daily_tab(market)
        if not days:
            return {'market': market, 'error': f'No data found for {market}'}

        if target_week is None:
            target_week = self._detect_latest_week(days)
        if not target_week:
            return {'market': market, 'error': 'Could not detect target week'}

        # Current week
        tw = self._aggregate_week(days, target_week)

        # Last week
        weeks = self._get_unique_weeks(days)
        tw_idx = weeks.index(target_week) if target_week in weeks else -1
        lw_label = weeks[tw_idx - 1] if tw_idx > 0 else None
        lw = self._aggregate_week(days, lw_label) if lw_label else None

        # YoY week
        yoy_label = self._find_yoy_week(target_week)
        yoy = self._aggregate_week(days, yoy_label) if yoy_label else None

        # YoY of last week (for WoW YoY comparison)
        yoy_lw_label = self._find_yoy_week(lw_label) if lw_label else None
        yoy_lw = self._aggregate_week(days, yoy_lw_label) if yoy_lw_label else None

        # Multi-week trend (last 8 weeks)
        trend = self._compute_multi_week_trend(days, target_week, num_weeks=8)

        # Anomalies
        anomalies = self._detect_anomalies(trend, tw)

        # Daily patterns
        daily_patterns = self._compute_daily_patterns(days, target_week)

        # Monthly projection
        budgets = self._read_monthly_budget()
        projection = self._project_month(days, target_week, market, budgets)

        # ie%CCP data
        ieccp_data = self._read_ieccp(target_week)
        market_ieccp = ieccp_data.get(market)

        # Weekly tab historical data (longer horizon than daily-derived trend)
        weekly_history = self._read_weekly_tab(market, target_week, num_weeks=52)

        # WoW changes
        wow = {}
        if tw and lw:
            for metric in ['cost', 'regs', 'cpa', 'clicks', 'impressions', 'cpc', 'cvr', 'ctr',
                           'brand_cost', 'brand_regs', 'brand_cpa', 'brand_clicks', 'brand_cvr',
                           'nb_cost', 'nb_regs', 'nb_cpa', 'nb_clicks', 'nb_cvr']:
                tw_val = tw.get(metric)
                lw_val = lw.get(metric)
                if tw_val is not None and lw_val is not None:
                    wow[metric] = {
                        'tw': tw_val, 'lw': lw_val,
                        'change': tw_val - lw_val,
                        'pct': pct(tw_val, lw_val),
                    }

        # YoY changes
        yoy_changes = {}
        if tw and yoy:
            for metric in ['cost', 'regs', 'cpa', 'clicks', 'impressions', 'cpc', 'cvr',
                           'brand_cost', 'brand_regs', 'brand_cpa',
                           'nb_cost', 'nb_regs', 'nb_cpa', 'nb_cvr']:
                tw_val = tw.get(metric)
                yoy_val = yoy.get(metric)
                if tw_val is not None and yoy_val is not None:
                    yoy_changes[metric] = {
                        'ty': tw_val, 'ly': yoy_val,
                        'change': tw_val - yoy_val,
                        'pct': pct(tw_val, yoy_val),
                    }

        # WoW comparison: this year's WoW vs last year's WoW
        wow_yoy = {}
        if tw and lw and yoy and yoy_lw:
            tw_wow_regs = pct(tw['regs'], lw['regs']) if lw['regs'] > 0 else None
            ly_wow_regs = pct(yoy['regs'], yoy_lw['regs']) if yoy_lw and yoy_lw['regs'] > 0 else None
            wow_yoy['regs_wow_ty'] = tw_wow_regs
            wow_yoy['regs_wow_ly'] = ly_wow_regs

        return {
            'market': market,
            'target_week': target_week,
            'current_week': tw,
            'last_week': lw,
            'yoy_week': yoy,
            'yoy_lw': yoy_lw,
            'wow': wow,
            'yoy_changes': yoy_changes,
            'wow_yoy': wow_yoy,
            'trend': [{'week': t['week'], 'regs': t['regs'], 'cost': t['cost'],
                       'cpa': t.get('cpa'), 'cvr': t.get('cvr'),
                       'brand_regs': t['brand_regs'], 'nb_regs': t['nb_regs']}
                      for t in trend],
            'anomalies': anomalies,
            'daily_patterns': daily_patterns,
            'projection': projection,
            'ieccp': market_ieccp,
            'weekly_history': weekly_history,
        }

    def generate_data_brief(self, analysis):
        """Generate a structured markdown data brief for analyst agents.

        Organized around the questions analysts need to answer, not raw metrics.
        """
        a = analysis
        if 'error' in a:
            return f"# {a['market']} Data Brief — Error\n\n{a['error']}\n"

        market = a['market']
        tw = a['current_week']
        lw = a['last_week']
        wow = a.get('wow', {})
        yoy = a.get('yoy_changes', {})
        trend = a.get('trend', [])
        anomalies = a.get('anomalies', [])
        daily = a.get('daily_patterns', {})
        proj = a.get('projection')
        wow_yoy = a.get('wow_yoy', {})
        ieccp = a.get('ieccp')

        def rpct(v):
            if v is None: return 'N/A'
            r = round(v * 100)
            if r == 0: return 'flat'
            return f'{r:+d}%'

        def rdollar(v):
            if v is None: return 'N/A'
            return f'${round(v):,}'

        def rk(v):
            if v >= 1000: return f'{v/1000:.1f}K'
            return str(v)

        def rspend(v):
            if v is None: return 'N/A'
            if v >= 1000000: return f'${v/1000000:.1f}M'
            if v >= 1000: return f'${round(v/1000):,}K'
            return f'${round(v):,}'

        lines = []
        wk = tw['week'].split()[-1] if tw else '?'
        lines.append(f'# {market} {wk} Data Brief')
        lines.append('')

        # ── Section 1: Headline numbers ──
        lines.append('## Headline numbers')
        if tw:
            lines.append(f'- Registrations: {tw["regs"]} ({rpct(wow.get("regs",{}).get("pct"))} WoW)')
            lines.append(f'- Spend: {rspend(tw["cost"])} ({rpct(wow.get("cost",{}).get("pct"))} WoW)')
            lines.append(f'- CPA: {rdollar(tw.get("cpa"))} ({rpct(wow.get("cpa",{}).get("pct"))} WoW)')
            lines.append(f'- Brand regs: {tw["brand_regs"]} ({rpct(wow.get("brand_regs",{}).get("pct"))} WoW)')
            lines.append(f'- NB regs: {tw["nb_regs"]} ({rpct(wow.get("nb_regs",{}).get("pct"))} WoW)')
        lines.append('')

        # ── Section 1.5: ie%CCP (if available) ──
        if ieccp:
            lines.append('## ie%CCP')
            tw_pct = round(ieccp['tw'] * 100)
            lines.append(f'- This week: {tw_pct}%')
            if ieccp.get('lw'):
                lw_pct = round(ieccp['lw'] * 100)
                lines.append(f'- Last week: {lw_pct}%')
            lines.append(f'- Target: 100%')
            lines.append('')

        # ── Section 2: Registration driver decomposition ──
        lines.append('## Registration drivers (what caused the WoW change?)')
        if tw and lw:
            # Overall: was it CVR or volume?
            cvr_wow = wow.get('cvr', {}).get('pct')
            clicks_wow = wow.get('clicks', {}).get('pct')
            if cvr_wow is not None and clicks_wow is not None:
                if abs(cvr_wow) > abs(clicks_wow):
                    lines.append(f'Primary driver: CVR ({rpct(cvr_wow)} WoW) more than clicks ({rpct(clicks_wow)} WoW)')
                else:
                    lines.append(f'Primary driver: clicks ({rpct(clicks_wow)} WoW) more than CVR ({rpct(cvr_wow)} WoW)')

            # Brand decomposition
            lines.append('')
            lines.append('Brand:')
            lines.append(f'  Regs: {tw["brand_regs"]} vs {lw["brand_regs"]} LW ({rpct(wow.get("brand_regs",{}).get("pct"))})')
            b_cvr = wow.get('brand_cvr', {}).get('pct')
            b_clicks = wow.get('brand_clicks', {}).get('pct')
            b_cpc = wow.get('brand_cpc', {}).get('pct') if wow.get('brand_cpc') else None
            b_imp = None
            if b_cvr is not None: lines.append(f'  CVR: {tw.get("brand_cvr",0):.2%} vs {lw.get("brand_cvr",0):.2%} ({rpct(b_cvr)})')
            if b_clicks is not None: lines.append(f'  Clicks: {tw["brand_clicks"]} vs {lw["brand_clicks"]} ({rpct(b_clicks)})')
            if b_cpc is not None: lines.append(f'  CPC: {rdollar(tw.get("brand_cpc"))} ({rpct(b_cpc)})')
            lines.append(f'  CPA: {rdollar(tw.get("brand_cpa"))} vs {rdollar(lw.get("brand_cpa"))} ({rpct(wow.get("brand_cpa",{}).get("pct"))})')

            # NB decomposition
            lines.append('')
            lines.append('Non-Brand:')
            lines.append(f'  Regs: {tw["nb_regs"]} vs {lw["nb_regs"]} LW ({rpct(wow.get("nb_regs",{}).get("pct"))})')
            n_cvr = wow.get('nb_cvr', {}).get('pct')
            n_clicks = wow.get('nb_clicks', {}).get('pct')
            n_cpc = wow.get('nb_cpc', {}).get('pct') if wow.get('nb_cpc') else None
            if n_cvr is not None: lines.append(f'  CVR: {tw.get("nb_cvr",0):.2%} vs {lw.get("nb_cvr",0):.2%} ({rpct(n_cvr)})')
            if n_clicks is not None: lines.append(f'  Clicks: {tw["nb_clicks"]} vs {lw["nb_clicks"]} ({rpct(n_clicks)})')
            if n_cpc is not None: lines.append(f'  CPC: {rdollar(tw.get("nb_cpc"))} ({rpct(n_cpc)})')
            lines.append(f'  CPA: {rdollar(tw.get("nb_cpa"))} vs {rdollar(lw.get("nb_cpa"))} ({rpct(wow.get("nb_cpa",{}).get("pct"))})')
        lines.append('')

        # ── Section 3: 8-week trend ──
        lines.append('## 8-week trend')
        if trend:
            lines.append('| Week | Regs | Brand | NB | Cost | CPA | CVR |')
            lines.append('|------|------|-------|-----|------|-----|-----|')
            for t in trend:
                w = t['week'].split()[-1]
                cvr_str = f'{t["cvr"]:.2%}' if t.get('cvr') else 'N/A'
                cpa_str = rdollar(t.get('cpa'))
                lines.append(f'| {w} | {t["regs"]} | {t["brand_regs"]} | {t["nb_regs"]} | {rspend(t["cost"])} | {cpa_str} | {cvr_str} |')
        lines.append('')

        # ── Section 4: YoY comparison ──
        yoy_has_data = yoy and yoy.get('regs', {}).get('ly', 0) > 0
        if yoy_has_data:
            lines.append('## YoY comparison')
            lines.append(f'- Regs: {tw["regs"]} TY vs {yoy["regs"]["ly"]} LY ({rpct(yoy["regs"]["pct"])})')
            lines.append(f'- Spend: {rspend(tw["cost"])} TY vs {rspend(yoy["cost"]["ly"])} LY ({rpct(yoy["cost"]["pct"])})')
            if yoy.get('brand_regs', {}).get('pct') is not None:
                lines.append(f'- Brand regs: {rpct(yoy["brand_regs"]["pct"])} YoY')
            if yoy.get('nb_regs', {}).get('pct') is not None:
                lines.append(f'- NB regs: {rpct(yoy["nb_regs"]["pct"])} YoY')
            if yoy.get('nb_cpa', {}).get('pct') is not None:
                lines.append(f'- NB CPA: {rdollar(yoy["nb_cpa"]["ty"])} vs {rdollar(yoy["nb_cpa"]["ly"])} LY ({rpct(yoy["nb_cpa"]["pct"])})')
            if wow_yoy.get('regs_wow_ty') is not None and wow_yoy.get('regs_wow_ly') is not None:
                lines.append(f'- WoW pattern: TY {rpct(wow_yoy["regs_wow_ty"])} vs LY {rpct(wow_yoy["regs_wow_ly"])} (same week)')
            lines.append('')

        # ── Section 5: Monthly projection inputs ──
        if proj:
            lines.append('## Monthly projection inputs')
            lines.append(f'- Month: {proj["month"]} ({proj["days_elapsed"]}/{proj["total_days"]} days elapsed, {proj["days_remaining"]} remaining)')
            lines.append(f'- MTD actuals: {rspend(proj["mtd_spend"])} spend, {proj["mtd_regs"]} regs ({proj.get("mtd_brand_regs", "?")} Brand, {proj.get("mtd_nb_regs", "?")} NB)')
            if proj.get('op2_spend'):
                lines.append(f'- OP2 targets: {rspend(proj["op2_spend"])} spend, {rk(proj["op2_regs"])} regs')
                pct_through = proj['days_elapsed'] / proj['total_days']
                expected_mtd_regs = proj['op2_regs'] * pct_through
                expected_mtd_spend = proj['op2_spend'] * pct_through
                lines.append(f'- OP2 pace check: at {pct_through:.0%} through the month, linear OP2 pace would be {rk(int(expected_mtd_regs))} regs and {rspend(expected_mtd_spend)} spend')
                lines.append(f'- MTD vs OP2 pace: {rpct(pct(proj["mtd_regs"], expected_mtd_regs))} regs, {rpct(pct(proj["mtd_spend"], expected_mtd_spend))} spend')
            lines.append(f'- Simple linear projection (ingester estimate, not accounting for seasonality/holidays): {rspend(proj["projected_spend"])} spend, {rk(proj["projected_regs"])} regs, {rdollar(proj["projected_cpa"])} CPA')
            lines.append(f'- NOTE: Analyst should produce the actual projection accounting for weekday/weekend mix, holidays, LY patterns, and known upcoming changes.')
            lines.append('')

        # ── Section 6: Anomalies ──
        if anomalies:
            lines.append('## Anomalies (>20% deviation from recent avg)')
            for an in anomalies:
                metric_name = an['metric'].replace('_', ' ')
                lines.append(f'- {metric_name}: {an["direction"]} avg by {abs(an["deviation_pct"]):.0%} (current: {an["current"]:.2f}, avg: {an["avg"]:.2f})')
            lines.append('')

        # ── Section 7: Weekly tab history (longer-term trends + LY comparison) ──
        weekly_history = a.get('weekly_history', [])
        if weekly_history:
            # Split into TY and LY based on week labels
            tw_year = self._get_week_year(a['target_week'])
            tw_wnum = self._get_week_number(a['target_week'])

            if tw_year and tw_wnum:
                # Get last 12 weeks of this year
                ty_weeks = [w for w in weekly_history if self._get_week_year(w['week']) == tw_year]
                ty_recent = ty_weeks[-12:] if len(ty_weeks) > 12 else ty_weeks

                # Get same period last year (target week ± 4 weeks)
                ly_weeks = []
                for w in weekly_history:
                    wy = self._get_week_year(w['week'])
                    wn = self._get_week_number(w['week'])
                    if wy == tw_year - 1 and wn is not None and abs(wn - tw_wnum) <= 4:
                        ly_weeks.append(w)

                if ty_recent:
                    lines.append('## This year weekly trend (last 12 weeks)')
                    lines.append('| Week | Regs | Brand | NB | Spend | CPA | NB CPC | NB CVR |')
                    lines.append('|------|------|-------|-----|-------|-----|--------|--------|')
                    for w in ty_recent:
                        wk = w['week'].split()[-1]
                        nb_cpc_str = f'${w["nb_cpc"]:.2f}' if w.get('nb_cpc') else 'N/A'
                        nb_cvr_str = f'{w["nb_cvr"]:.2%}' if w.get('nb_cvr') else 'N/A'
                        lines.append(f'| {wk} | {w["regs"]} | {w["brand_regs"]} | {w["nb_regs"]} | {rspend(w["spend"])} | {rdollar(w.get("cpa"))} | {nb_cpc_str} | {nb_cvr_str} |')
                    lines.append('')

                if ly_weeks:
                    lines.append(f'## Last year same period (W{tw_wnum-4} to W{tw_wnum+4})')
                    lines.append('| Week | Regs | Brand | NB | Spend | CPA | NB CPC | NB CVR |')
                    lines.append('|------|------|-------|-----|-------|-----|--------|--------|')
                    for w in ly_weeks:
                        wk = w['week'].split()[-1]
                        nb_cpc_str = f'${w["nb_cpc"]:.2f}' if w.get('nb_cpc') else 'N/A'
                        nb_cvr_str = f'{w["nb_cvr"]:.2%}' if w.get('nb_cvr') else 'N/A'
                        lines.append(f'| {wk} | {w["regs"]} | {w["brand_regs"]} | {w["nb_regs"]} | {rspend(w["spend"])} | {rdollar(w.get("cpa"))} | {nb_cpc_str} | {nb_cvr_str} |')
                    lines.append('')

                # Streak detection on key metrics
                streaks = []
                if len(ty_recent) >= 3:
                    # NB CPC streak
                    nb_cpc_vals = [(w['week'].split()[-1], w.get('nb_cpc')) for w in ty_recent if w.get('nb_cpc')]
                    if nb_cpc_vals:
                        cpc_streak = 0
                        cpc_dir = None
                        for i in range(len(nb_cpc_vals) - 1, 0, -1):
                            if nb_cpc_vals[i][1] < nb_cpc_vals[i-1][1]:
                                if cpc_dir is None or cpc_dir == 'declining':
                                    cpc_streak += 1
                                    cpc_dir = 'declining'
                                else:
                                    break
                            elif nb_cpc_vals[i][1] > nb_cpc_vals[i-1][1]:
                                if cpc_dir is None or cpc_dir == 'rising':
                                    cpc_streak += 1
                                    cpc_dir = 'rising'
                                else:
                                    break
                            else:
                                break
                        if cpc_streak >= 3:
                            streaks.append(f'NB CPC {cpc_dir} {cpc_streak} consecutive weeks (${nb_cpc_vals[-cpc_streak-1][1]:.2f} in {nb_cpc_vals[-cpc_streak-1][0]} to ${nb_cpc_vals[-1][1]:.2f} in {nb_cpc_vals[-1][0]})')

                    # CPA streak
                    cpa_vals = [(w['week'].split()[-1], w.get('cpa')) for w in ty_recent if w.get('cpa')]
                    if cpa_vals:
                        cpa_streak = 0
                        cpa_dir = None
                        for i in range(len(cpa_vals) - 1, 0, -1):
                            if cpa_vals[i][1] < cpa_vals[i-1][1]:
                                if cpa_dir is None or cpa_dir == 'declining':
                                    cpa_streak += 1
                                    cpa_dir = 'declining'
                                else:
                                    break
                            elif cpa_vals[i][1] > cpa_vals[i-1][1]:
                                if cpa_dir is None or cpa_dir == 'rising':
                                    cpa_streak += 1
                                    cpa_dir = 'rising'
                                else:
                                    break
                            else:
                                break
                        if cpa_streak >= 3:
                            streaks.append(f'CPA {cpa_dir} {cpa_streak} consecutive weeks ({rdollar(cpa_vals[-cpa_streak-1][1])} in {cpa_vals[-cpa_streak-1][0]} to {rdollar(cpa_vals[-1][1])} in {cpa_vals[-1][0]})')

                    # Regs trend (consecutive up or down)
                    reg_vals = [(w['week'].split()[-1], w['regs']) for w in ty_recent if w['regs'] > 0]
                    if reg_vals:
                        reg_streak = 0
                        reg_dir = None
                        for i in range(len(reg_vals) - 1, 0, -1):
                            if reg_vals[i][1] < reg_vals[i-1][1]:
                                if reg_dir is None or reg_dir == 'declining':
                                    reg_streak += 1
                                    reg_dir = 'declining'
                                else:
                                    break
                            elif reg_vals[i][1] > reg_vals[i-1][1]:
                                if reg_dir is None or reg_dir == 'rising':
                                    reg_streak += 1
                                    reg_dir = 'rising'
                                else:
                                    break
                            else:
                                break
                        if reg_streak >= 3:
                            streaks.append(f'Regs {reg_dir} {reg_streak} consecutive weeks ({reg_vals[-reg_streak-1][1]} in {reg_vals[-reg_streak-1][0]} to {reg_vals[-1][1]} in {reg_vals[-1][0]})')

                if streaks:
                    lines.append('## Detected streaks')
                    for s in streaks:
                        lines.append(f'- {s}')
                    lines.append('')

        # ── Section 8: Daily breakdown ──
        if daily and daily.get('daily'):
            lines.append('## Daily breakdown')
            lines.append('| Day | Regs | Brand | NB | Cost |')
            lines.append('|-----|------|-------|-----|------|')
            for d in daily['daily']:
                lines.append(f'| {d["dow"]} | {d["regs"]} | {d["brand_regs"]} | {d["nb_regs"]} | {rspend(d["cost"])} |')

            # Data lag check
            all_days = daily['daily']
            if len(all_days) >= 5:
                weekday_regs = [d['regs'] for d in all_days[:5]]
                weekday_avg = sum(weekday_regs) / len(weekday_regs)
                last_2 = all_days[-2:]
                last_2_avg = sum(d['regs'] for d in last_2) / len(last_2)
                if weekday_avg > 0 and last_2_avg < weekday_avg * 0.30:
                    lines.append(f'⚠️ DATA LAG: last 2 days avg {last_2_avg:.0f} regs vs weekday avg {weekday_avg:.0f}')
            lines.append('')

        return '\n'.join(lines)

    def generate_callout_md(self, analysis, market_context=None):
        """Generate a markdown callout draft matching the WBR callout style.

        Follows callout-principles.md:
        - Headline: regs, WoW%, spend WoW%, CPA. Monthly projection vs OP2.
        - WoW paragraph: prose, Brand and NB together, attribute to CVR/CPC/clicks.
        - YoY paragraph: spend and regs YoY, Brand vs NB drivers.
        - Note: internal PS context, anomalies, observations.
        - Rounding: whole %, whole dollars, "1.1K" for thousands.
        """
        a = analysis
        if 'error' in a:
            return f"# {a['market']} — Error\n\n{a['error']}\n"

        market = a['market']
        tw = a['current_week']
        lw = a['last_week']
        wow = a['wow']
        yoy = a.get('yoy_changes', {})
        proj = a.get('projection')
        anomalies = a.get('anomalies', [])
        trend = a.get('trend', [])
        daily = a.get('daily_patterns', {})
        wow_yoy_data = a.get('wow_yoy', {})

        # ── Helpers for callout-style rounding ──
        def rpct(v):
            """Round percentage to whole number with sign, e.g. '+33%' or '-7%'."""
            if v is None:
                return 'flat'
            rounded = round(v * 100)
            if rounded == 0:
                return 'flat'
            return f'{rounded:+d}%'

        def rdollar(v):
            """Round to whole dollar, e.g. '$132'."""
            if v is None:
                return 'N/A'
            return f'${round(v):,}'

        def rk(v):
            """Format large numbers as '1.1K', '32.9K', or just the number if <1000."""
            if v >= 10000:
                return f'{v/1000:.1f}K'
            elif v >= 1000:
                return f'{v/1000:.1f}K'
            else:
                return str(v)

        def rspend(v):
            """Format spend as '$139K' or '$68K'."""
            if v is None:
                return 'N/A'
            if v >= 1000000:
                return f'${v/1000000:.1f}M'
            elif v >= 1000:
                return f'${round(v/1000):,}K'
            else:
                return f'${round(v):,}'

        # ── Parse week dates ──
        week_num = tw['week'].split()[-1] if tw else '?'
        date_range_raw = tw.get('date_range', '') if tw else ''
        try:
            parts = date_range_raw.split(' to ')
            d1 = datetime.datetime.strptime(parts[0], '%Y-%m-%d')
            d2 = datetime.datetime.strptime(parts[1], '%Y-%m-%d')
            date_range = f"{d1.strftime('%b')} {d1.day}\u2013{d2.day}"
        except:
            date_range = date_range_raw

        lines = []
        lines.append(f"# {market} \u2014 {week_num} ({date_range})")
        lines.append('')

        # ── Headline paragraph ──
        if tw and wow:
            regs = tw['regs']
            regs_wow_pct = rpct(wow.get('regs', {}).get('pct'))
            spend_wow_pct = rpct(wow.get('cost', {}).get('pct'))
            cpa = tw.get('cpa')

            headline = f"{market} drove {rk(regs)} registrations ({regs_wow_pct} WoW), with {spend_wow_pct} spend WoW"

            # CPA framing
            cpa_wow = wow.get('cpa', {}).get('pct')
            if cpa is not None:
                if cpa_wow is not None and abs(cpa_wow) > 0.03:
                    cpa_direction = 'increased' if cpa_wow > 0 else 'decreased'
                    headline += f". CPA {cpa_direction} to {rdollar(cpa)} ({rpct(cpa_wow)} WoW)"
                else:
                    headline += f". CPA {rdollar(cpa)}"

            headline += '.'

            # Monthly projection
            if proj:
                month_name = proj['month'].split()[-1]
                proj_spend_str = rspend(proj['projected_spend'])
                proj_regs_str = rk(proj['projected_regs'])
                proj_cpa_str = rdollar(proj['projected_cpa']) if proj['projected_cpa'] else 'N/A'

                vs_op2_parts = []
                if proj.get('vs_op2_spend_pct') is not None:
                    vs_op2_parts.append(f"{rpct(proj['vs_op2_spend_pct'])} spend")
                if proj.get('vs_op2_regs_pct') is not None:
                    v = proj['vs_op2_regs_pct']
                    if abs(v) < 0.02:
                        vs_op2_parts.append('flat registrations')
                    else:
                        vs_op2_parts.append(f"{rpct(v)} registrations")

                vs_op2_str = ', '.join(vs_op2_parts)
                headline += f" {month_name} is projected to end at {proj_spend_str} spend and {proj_regs_str} registrations, and {proj_cpa_str} CPA."
                if vs_op2_str:
                    headline += f" (vs. OP2: {vs_op2_str})"

            lines.append(headline)

        # ── WoW paragraph (prose) ──
        if wow:
            br = wow.get('brand_regs', {})
            nb = wow.get('nb_regs', {})
            br_tw, br_lw = br.get('tw', 0), br.get('lw', 0)
            nb_tw, nb_lw = nb.get('tw', 0), nb.get('lw', 0)
            br_pct = br.get('pct')
            nb_pct = nb.get('pct')

            both_up = (br_pct and br_pct > 0.01) and (nb_pct and nb_pct > 0.01)
            both_down = (br_pct and br_pct < -0.01) and (nb_pct and nb_pct < -0.01)

            wow_text = ''
            brand_cvr = wow.get('brand_cvr', {}).get('pct')
            nb_cvr = wow.get('nb_cvr', {}).get('pct')
            clicks_wow = wow.get('clicks', {}).get('pct')
            imp_wow = wow.get('impressions', {}).get('pct')
            nb_cpc_wow = wow.get('nb_cpc', {}).get('pct') if wow.get('nb_cpc') else None

            if both_up:
                if br_tw >= nb_tw:
                    wow_text = f"WoW registrations increased on both Brand ({rpct(br_pct)}) and NB ({rpct(nb_pct)})"
                else:
                    wow_text = f"WoW registrations increased on both NB ({rpct(nb_pct)}) and Brand ({rpct(br_pct)})"
                if brand_cvr is not None and nb_cvr is not None:
                    wow_text += f", driven by CVR increases on both sides ({rpct(brand_cvr)} Brand, {rpct(nb_cvr)} NB)"
                wow_text += '.'
            else:
                # Describe each segment with "On the X side" prose
                seg_parts = []
                # Lead with the segment that has more volume
                segs = [('NB', nb_tw, nb_lw, nb_pct, nb_cvr), ('Brand', br_tw, br_lw, br_pct, brand_cvr)]
                if br_tw >= nb_tw:
                    segs = [('Brand', br_tw, br_lw, br_pct, brand_cvr), ('NB', nb_tw, nb_lw, nb_pct, nb_cvr)]

                for seg_name, s_tw, s_lw, s_pct, s_cvr in segs:
                    verb = 'increased' if s_pct and s_pct > 0 else 'fell' if s_pct and s_pct < -0.01 else 'were flat'
                    part = f"On the {seg_name} side, registrations {verb} {rpct(s_pct)} ({s_lw} to {s_tw})"

                    # Add driver detail
                    drivers = []
                    if s_cvr is not None and abs(s_cvr) > 0.03:
                        drivers.append(f"CVR {rpct(s_cvr)}")
                    seg_clicks = wow.get(f'{"brand" if seg_name == "Brand" else "nb"}_clicks', {}).get('pct')
                    if seg_clicks is not None and abs(seg_clicks) > 0.03:
                        drivers.append(f"clicks {rpct(seg_clicks)}")
                    seg_cpc = wow.get(f'{"brand" if seg_name == "Brand" else "nb"}_cpc', {}).get('pct') if wow.get(f'{"brand" if seg_name == "Brand" else "nb"}_cpc') else None
                    if seg_cpc is not None and abs(seg_cpc) > 0.03:
                        drivers.append(f"CPC {rpct(seg_cpc)}")

                    # Add CPA context
                    seg_cpa_tw = tw.get(f'{"brand" if seg_name == "Brand" else "nb"}_cpa')
                    seg_cpa_lw = lw.get(f'{"brand" if seg_name == "Brand" else "nb"}_cpa') if lw else None
                    if seg_cpa_tw and seg_cpa_lw:
                        seg_cpa_wow = pct(seg_cpa_tw, seg_cpa_lw)
                        if seg_cpa_wow is not None and abs(seg_cpa_wow) > 0.05:
                            cpa_verb = 'rose' if seg_cpa_wow > 0 else 'improved'
                            drivers.append(f"CPA {cpa_verb} to {rdollar(seg_cpa_tw)} ({rpct(seg_cpa_wow)})")

                    if drivers:
                        part += ', ' + ', '.join(drivers[:2])  # limit to 2 drivers per segment

                    seg_parts.append(part)

                wow_text = '. '.join(seg_parts) + '.'

            lines.append(wow_text)

        # ── YoY paragraph (suppress if no meaningful data) ──
        yoy_has_data = yoy and yoy.get('regs', {}).get('ly', 0) > 0
        if yoy_has_data:
            cost_yoy = yoy.get('cost', {})
            regs_yoy = yoy.get('regs', {})

            yoy_text = f"YoY we spent {rpct(cost_yoy.get('pct'))} with {rpct(regs_yoy.get('pct'))} registrations."

            brand_regs_yoy = yoy.get('brand_regs', {})
            nb_regs_yoy = yoy.get('nb_regs', {})
            nb_cpa_yoy = yoy.get('nb_cpa', {})

            breakdown_parts = []
            if brand_regs_yoy.get('pct') is not None and nb_regs_yoy.get('pct') is not None:
                br_yoy_abs = abs(brand_regs_yoy.get('change', 0))
                nb_yoy_abs = abs(nb_regs_yoy.get('change', 0))
                if br_yoy_abs > nb_yoy_abs:
                    breakdown_parts.append(f"Brand regs {rpct(brand_regs_yoy['pct'])} YoY")
                    breakdown_parts.append(f"NB regs {rpct(nb_regs_yoy['pct'])} YoY")
                else:
                    breakdown_parts.append(f"NB regs {rpct(nb_regs_yoy['pct'])} YoY")
                    breakdown_parts.append(f"Brand regs {rpct(brand_regs_yoy['pct'])} YoY")

            if nb_cpa_yoy.get('pct') is not None and nb_cpa_yoy.get('ty') and nb_cpa_yoy.get('ly'):
                breakdown_parts.append(f"NB CPA {rpct(nb_cpa_yoy['pct'])} YoY ({rdollar(nb_cpa_yoy['ly'])} vs {rdollar(nb_cpa_yoy['ty'])})")

            if breakdown_parts:
                yoy_text += ' ' + '. '.join(breakdown_parts) + '.'

            lines.append(yoy_text)

        # ── Note paragraph ──
        note_parts = []

        # NB CVR anomaly
        nb_cvr_tw = tw.get('nb_cvr') if tw else None
        if nb_cvr_tw and trend and len(trend) >= 4:
            recent_nb_cvrs = [t.get('nb_regs', 0) / t.get('nb_clicks', 1) if t.get('nb_clicks', 0) > 0 else 0 for t in trend[:-1]]
            if recent_nb_cvrs:
                avg_nb_cvr = sum(recent_nb_cvrs) / len(recent_nb_cvrs)
                if avg_nb_cvr > 0 and nb_cvr_tw > avg_nb_cvr * 1.20:
                    note_parts.append(f"{week_num} NB CVR of {nb_cvr_tw:.2%} is above the recent average of {avg_nb_cvr:.2%}")
                elif avg_nb_cvr > 0 and nb_cvr_tw < avg_nb_cvr * 0.80:
                    note_parts.append(f"{week_num} NB CVR of {nb_cvr_tw:.2%} is below the recent average of {avg_nb_cvr:.2%}")

        # WoW this year vs last year
        if wow_yoy_data.get('regs_wow_ty') is not None and wow_yoy_data.get('regs_wow_ly') is not None:
            note_parts.append(f"{week_num} WoW last year was {rpct(wow_yoy_data['regs_wow_ly'])} regs")

        # Data lag warning
        if daily and daily.get('daily'):
            all_days = daily.get('daily', [])
            if len(all_days) >= 5:
                weekday_regs = [d['regs'] for d in all_days[:5]]
                weekday_avg = sum(weekday_regs) / len(weekday_regs) if weekday_regs else 0
                last_2 = all_days[-2:]
                if weekday_avg > 0:
                    last_2_avg = sum(d['regs'] for d in last_2) / len(last_2)
                    if last_2_avg < weekday_avg * 0.30:
                        note_parts.append(f"Last 2 days show significantly lower volume (avg {last_2_avg:.0f} vs weekday avg {weekday_avg:.0f}), possible data lag")

        if note_parts:
            lines.append('')
            lines.append('Note: ' + '. '.join(note_parts) + '.')

        lines.append('')

        # ── Supplementary analysis (below the callout) ──
        if len(trend) >= 4:
            lines.append('---')
            lines.append('')
            lines.append('Weekly trend (regs): ' + ' | '.join([f"{t['week'].split()[-1]}: {t['regs']}" for t in trend]))

        if anomalies:
            lines.append('')
            lines.append('Flagged anomalies:')
            for an in anomalies:
                metric_name = an['metric'].replace('_', ' ')
                lines.append(f"- {metric_name} {an['direction']} recent avg by {abs(an['deviation_pct']):.0%}")

        if daily and daily.get('daily'):
            lines.append('')
            lines.append('Daily: ' + ', '.join([f"{d['dow']} {d['regs']}" for d in daily['daily']]))

        lines.append('')
        return '\n'.join(lines)


    def run(self, target_week=None, markets=None, output_dir=None):
        """Main entry point. Analyze all requested markets and produce output."""
        if markets is None:
            markets = ALL_MARKETS

        # Auto-detect week from AU (or first available market)
        if target_week is None:
            for m in markets:
                days = self._read_daily_tab(m)
                if days:
                    target_week = self._detect_latest_week(days)
                    if target_week:
                        break

        if not target_week:
            print("ERROR: Could not detect target week from any market data.")
            return

        print(f"Target week: {target_week}")
        print(f"Markets: {', '.join(markets)}")
        print()

        results = {}
        for market in markets:
            print(f"Analyzing {market}...")
            analysis = self.analyze_market(market, target_week)
            results[market] = analysis

        # Determine output directory
        if output_dir is None:
            week_num = target_week.replace(' ', '-').lower()
            output_dir = f'shared/context/active/callouts'

        os.makedirs(output_dir, exist_ok=True)

        # Write per-market callout drafts
        week_slug = target_week.lower().replace(' ', '-').replace('2026 ', '2026-')
        week_num_only = target_week.split()[-1].lower() if target_week else 'unknown'

        for market in markets:
            analysis = results[market]
            callout_md = self.generate_callout_md(analysis)

            # Write to market subfolder
            market_dir = os.path.join(output_dir, market.lower())
            os.makedirs(market_dir, exist_ok=True)
            callout_path = os.path.join(market_dir, f'{market.lower()}-2026-{week_num_only}.md')
            with open(callout_path, 'w') as f:
                f.write(callout_md)
            print(f"  Wrote {callout_path}")

            # Write per-market data brief for analyst agents
            brief_md = self.generate_data_brief(analysis)
            brief_path = os.path.join(market_dir, f'{market.lower()}-data-brief-2026-{week_num_only}.md')
            with open(brief_path, 'w') as f:
                f.write(brief_md)
            print(f"  Wrote {brief_path}")

        # Read monthly actuals and attach to each market result
        monthly_actuals = self._read_monthly_actuals()
        monthly_budgets = self._read_monthly_budget()
        for market in markets:
            if market in results and 'error' not in results[market]:
                results[market]['monthly_actuals'] = monthly_actuals.get(market, {})
                results[market]['monthly_budget'] = monthly_budgets.get(market, {})

        # Write JSON data extract
        json_dir = os.path.join('shared', 'tools', 'dashboard-ingester', 'data')
        os.makedirs(json_dir, exist_ok=True)
        json_path = os.path.join(json_dir, f'{week_slug}.json')

        # Serialize (handle None and float issues)
        def json_safe(obj):
            if isinstance(obj, float):
                if obj != obj:  # NaN
                    return None
                return round(obj, 6)
            return obj

        with open(json_path, 'w') as f:
            json.dump(results, f, indent=2, default=str)
        print(f"  Wrote {json_path}")

        # Write monthly actuals JSON (standalone, for MBR/flash/reporting use)
        monthly_path = os.path.join(json_dir, 'monthly-actuals.json')
        monthly_combined = {}
        for market in ALL_MARKETS:
            monthly_combined[market] = {
                'actuals': monthly_actuals.get(market, {}),
                'op2': monthly_budgets.get(market, {}),
            }
        with open(monthly_path, 'w') as f:
            json.dump(monthly_combined, f, indent=2, default=str)
        print(f"  Wrote {monthly_path}")

        # Write WW summary
        summary_path = os.path.join(output_dir, f'ww-summary-2026-{week_num_only}.md')
        summary = self._generate_ww_summary(results, target_week)
        with open(summary_path, 'w') as f:
            f.write(summary)
        print(f"  Wrote {summary_path}")

        print("\nDone.")
        return results

    def _generate_ww_summary(self, results, target_week):
        """Generate a WW summary across all markets."""
        lines = []
        week_num = target_week.split()[-1] if target_week else '?'
        lines.append(f"# WW Summary — {week_num}")
        lines.append('')

        # Summary table
        lines.append('| Market | Regs | WoW% | Spend | WoW% | CPA | WoW% | YoY Regs% | Proj Regs | vs OP2 |')
        lines.append('|--------|------|------|-------|------|-----|------|-----------|-----------|--------|')

        ww_regs = 0
        ww_lw_regs = 0
        ww_spend = 0
        ww_lw_spend = 0

        for market in ALL_MARKETS:
            if market not in results:
                continue
            a = results[market]
            if 'error' in a:
                lines.append(f'| {market} | — | — | — | — | — | — | — | — | — |')
                continue

            tw = a.get('current_week', {})
            wow = a.get('wow', {})
            yoy = a.get('yoy_changes', {})
            proj = a.get('projection')

            regs = tw.get('regs', 0)
            spend = tw.get('cost', 0)
            cpa = tw.get('cpa')
            regs_wow = wow.get('regs', {}).get('pct')
            spend_wow = wow.get('cost', {}).get('pct')
            cpa_wow = wow.get('cpa', {}).get('pct')
            regs_yoy = yoy.get('regs', {}).get('pct')
            # Suppress YoY for markets with no meaningful prior year data
            if yoy.get('regs', {}).get('ly', 0) == 0:
                regs_yoy = None
            proj_regs = proj.get('projected_regs') if proj else None
            vs_op2 = proj.get('vs_op2_regs_pct') if proj else None

            ww_regs += regs
            ww_lw_regs += wow.get('regs', {}).get('lw', 0)
            ww_spend += spend
            ww_lw_spend += wow.get('cost', {}).get('lw', 0)

            lines.append(f"| {market} | {fmt_int(regs)} | {fmt_pct(regs_wow)} | {fmt_dollar(spend)} | {fmt_pct(spend_wow)} | "
                         f"{fmt_dollar(cpa) if cpa else 'N/A'} | {fmt_pct(cpa_wow)} | {fmt_pct(regs_yoy)} | "
                         f"{fmt_int(proj_regs) if proj_regs else 'N/A'} | {fmt_pct(vs_op2) if vs_op2 is not None else 'N/A'} |")

        # WW totals
        ww_cpa = ww_spend / ww_regs if ww_regs > 0 else None
        ww_regs_wow = pct(ww_regs, ww_lw_regs)
        ww_spend_wow = pct(ww_spend, ww_lw_spend)
        lines.append(f"| **WW** | **{fmt_int(ww_regs)}** | **{fmt_pct(ww_regs_wow)}** | **{fmt_dollar(ww_spend)}** | "
                     f"**{fmt_pct(ww_spend_wow)}** | **{fmt_dollar(ww_cpa) if ww_cpa else 'N/A'}** | — | — | — | — |")

        lines.append('')

        # Key callouts
        lines.append('## Key Callouts')
        lines.append('')

        # Find biggest movers
        movers = []
        for market in ALL_MARKETS:
            if market not in results or 'error' in results[market]:
                continue
            wow = results[market].get('wow', {})
            regs_pct = wow.get('regs', {}).get('pct')
            if regs_pct is not None:
                movers.append((market, regs_pct))

        movers.sort(key=lambda x: x[1])
        if movers:
            best = movers[-1]
            worst = movers[0]
            lines.append(f"- Biggest gainer: {best[0]} ({fmt_pct(best[1])} regs WoW)")
            lines.append(f"- Biggest decliner: {worst[0]} ({fmt_pct(worst[1])} regs WoW)")

        # Anomalies across markets
        all_anomalies = []
        for market in ALL_MARKETS:
            if market not in results or 'error' in results[market]:
                continue
            for an in results[market].get('anomalies', []):
                all_anomalies.append((market, an))

        if all_anomalies:
            lines.append('')
            lines.append('## Anomalies')
            lines.append('')
            for market, an in all_anomalies:
                metric_name = an['metric'].replace('_', ' ').title()
                lines.append(f"- {market}: {metric_name} {an['direction']} recent avg by {abs(an['deviation_pct']):.0%}")

        lines.append('')
        return '\n'.join(lines)


def main():
    parser = argparse.ArgumentParser(description='WW Dashboard Ingester')
    parser.add_argument('xlsx_path', help='Path to the WW Dashboard xlsx file')
    parser.add_argument('--week', help='Target week (e.g., "2026 W12"). Auto-detects if omitted.')
    parser.add_argument('--markets', help='Comma-separated market list (e.g., AU,MX). All if omitted.')
    parser.add_argument('--output-dir', help='Output directory for callout files.')
    args = parser.parse_args()

    markets = args.markets.split(',') if args.markets else None

    ingester = DashboardIngester(args.xlsx_path)
    ingester.run(target_week=args.week, markets=markets, output_dir=args.output_dir)


if __name__ == '__main__':
    main()
